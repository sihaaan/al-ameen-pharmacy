import csv
import hashlib
import io
import re
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SUPPORTED_OUTSTANDING_EXTENSIONS = {".csv", ".xlsx"}
SUPPORTED_CATEGORY_EXTENSIONS = {".xlsx"}
SUPPORTED_BLOCKLIST_EXTENSIONS = {".csv", ".xlsx"}
REPORT_DATE_RE = re.compile(r"as\s+on\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", re.IGNORECASE)
REPORT_DATE_TO_RE = re.compile(r"\bto\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", re.IGNORECASE)
HEADER_ALIASES = {
    "code": "code",
    "party": "party",
    "place": "place",
    "bill no": "bill_no",
    "bill no.": "bill_no",
    "date": "date",
    "amount": "amount",
    "0-30": "bucket_0_30",
    "0 - 30": "bucket_0_30",
    "30-60": "bucket_30_60",
    "30 - 60": "bucket_30_60",
    "60-90": "bucket_60_90",
    "60 - 90": "bucket_60_90",
    "over 90": "bucket_over_90",
    "total": "total",
    "days": "days",
}
CATEGORY_VALUES = {
    "credit": "credit",
    "insurance": "insurance",
    "clinic": "clinic",
    "branch": "branch",
    "card": "card",
    "misc": "misc",
}


@dataclass
class ParsedInvoiceRow:
    source_row_number: int
    customer_code: str
    customer_name: str
    place: str
    bill_number: str
    invoice_number: str
    lpo_reference: str
    invoice_date: date | None
    amount: Decimal
    bucket_0_30: Decimal
    bucket_30_60: Decimal
    bucket_60_90: Decimal
    bucket_over_90: Decimal
    total: Decimal
    days: int
    raw_data: dict
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParsedOutstanding:
    filename: str
    sha256: str
    size: int
    report_date: date | None
    rows: list[ParsedInvoiceRow]
    skipped_row_count: int
    warnings: list[str]
    parse_meta: dict


@dataclass
class ParsedCategoryMap:
    filename: str
    sha256: str
    entries: dict[str, str]
    code_entries: dict[str, str]
    warnings: list[str]
    parse_meta: dict


@dataclass
class ParsedBlocklist:
    filename: str
    sha256: str
    entries: list[dict]
    warnings: list[str]
    parse_meta: dict


@dataclass
class UploadedAccountingSource:
    filename: str
    extension: str
    data: bytes
    sha256: str

    @property
    def size(self):
        return len(self.data)


def max_upload_bytes():
    return int(getattr(settings, "ACCOUNTING_IMPORT_MAX_UPLOAD_BYTES", 25 * 1024 * 1024))


def max_import_rows():
    return int(getattr(settings, "ACCOUNTING_IMPORT_MAX_ROWS", 100_000))


def max_import_columns():
    return int(getattr(settings, "ACCOUNTING_IMPORT_MAX_COLUMNS", 80))


def normalize_customer_name(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value):
    text = str(value or "").strip().lower()
    text = text.replace(".", "")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if re.fullmatch(r"0\s*30", text):
        return "0-30"
    if re.fullmatch(r"30\s*60", text):
        return "30-60"
    if re.fullmatch(r"60\s*90", text):
        return "60-90"
    return text


def read_upload(uploaded_file, allowed_extensions):
    if not uploaded_file:
        raise ValidationError("Upload a file.")
    filename = Path(uploaded_file.name or "upload").name
    extension = Path(filename).suffix.lower()
    if extension not in allowed_extensions:
        allowed = ", ".join(sorted(allowed_extensions))
        raise ValidationError(f"Unsupported file type. Upload {allowed} only.")
    data = uploaded_file.read()
    if len(data) > max_upload_bytes():
        raise ValidationError(f"File is too large. Maximum size is {max_upload_bytes() // (1024 * 1024)} MB.")
    if extension == ".xlsx" and (not data.startswith(b"PK") or not zipfile.is_zipfile(io.BytesIO(data))):
        raise ValidationError("Invalid Excel file. Upload a valid .xlsx workbook.")
    if extension == ".csv" and data.startswith(b"PK"):
        raise ValidationError("Invalid CSV file.")
    return filename, extension, data, hashlib.sha256(data).hexdigest()


def read_outstanding_source(uploaded_file):
    filename, extension, data, sha256 = read_upload(uploaded_file, SUPPORTED_OUTSTANDING_EXTENSIONS)
    return UploadedAccountingSource(filename=filename, extension=extension, data=data, sha256=sha256)


def _looks_like_blocklist_header(value):
    normalized = normalize_header(value)
    return normalized in {
        "name",
        "company",
        "company name",
        "customer",
        "customer name",
        "party",
        "cat",
        "category",
        "type",
        "blocked",
        "blocklist",
    }


def _blocklist_entry_from_cells(cells):
    values = [str(cell or "").strip() for cell in cells]
    values = [value for value in values if value]
    if not values:
        return None
    name = values[0]
    if _looks_like_blocklist_header(name):
        return None
    normalized = normalize_customer_name(name)
    if not normalized or len(normalized) < 2 or normalized.isdigit():
        return None
    return {
        "name": name,
        "normalized_name": normalized,
        "category_hint": values[1] if len(values) > 1 and not _looks_like_blocklist_header(values[1]) else "",
    }


def parse_blocklist_upload(uploaded_file):
    filename, extension, data, sha256 = read_upload(uploaded_file, SUPPORTED_BLOCKLIST_EXTENSIONS)
    warnings = []
    entries_by_name = {}

    if extension == ".csv":
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1", errors="replace")
        for row in csv.reader(io.StringIO(text)):
            entry = _blocklist_entry_from_cells(row)
            if entry:
                entries_by_name.setdefault(entry["normalized_name"], entry)
    else:
        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValidationError("Could not read the blocklist workbook.") from exc
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                entry = _blocklist_entry_from_cells(row)
                if entry:
                    entries_by_name.setdefault(entry["normalized_name"], entry)

    if not entries_by_name:
        raise ValidationError("No company names were found in the blocklist file.")

    return ParsedBlocklist(
        filename=filename,
        sha256=sha256,
        entries=list(entries_by_name.values()),
        warnings=warnings,
        parse_meta={"entry_count": len(entries_by_name), "source_retention": "parsed_and_discarded"},
    )


def parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            parsed = timezone.datetime.strptime(text, fmt).date()
            if parsed.year < 1950:
                parsed = parsed.replace(year=parsed.year + 100)
            return parsed
        except ValueError:
            continue
    return None


def parse_report_date(cells):
    for cell in cells:
        text = str(cell or "")
        match = REPORT_DATE_RE.search(text) or REPORT_DATE_TO_RE.search(text)
        if match:
            return parse_date(match.group(1))
    return None


def parse_decimal(value):
    text = str(value or "").strip()
    if not text or text in {"-", "--"}:
        return Decimal("0.00")
    negative = text.startswith("(") or text.endswith("-")
    text = text.replace("AED", "").replace(",", "").replace("(", "").replace(")", "").replace("-", "")
    text = re.sub(r"[^0-9.]", "", text)
    if not text:
        return Decimal("0.00")
    try:
        number = Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        raise ValueError(f"Invalid amount: {value}")
    return -number if negative else number


def parse_int(value):
    try:
        return int(parse_decimal(value))
    except Exception:
        return 0


def split_bill_reference(value):
    text = str(value or "").strip()
    if not text:
        return "", ""
    match = re.match(r"^(\d{6})(.*)$", text)
    if not match:
        return re.sub(r"-+", "-", text).strip("- "), ""
    invoice_number = match.group(1)
    reference = (match.group(2) or "").strip()
    reference = re.sub(r"^-+", "", reference)
    reference = re.sub(r"-+", "-", reference)
    reference = reference.strip("- ")
    return invoice_number, reference


def find_data_start(row):
    normalized = [normalize_header(cell) for cell in row]
    for index in range(0, max(1, len(normalized) - 10)):
        window = normalized[index : index + 12]
        roles = [HEADER_ALIASES.get(item, item) for item in window]
        if roles[:6] == ["code", "party", "place", "bill_no", "date", "amount"] and "days" in roles:
            return index + 12
    return None


def clean_row(row):
    return ["" if cell is None else str(cell).strip() for cell in row]


def row_is_trailer(row, data_start):
    if data_start is None or len(row) <= data_start:
        return True
    text = " ".join(str(cell or "").strip().lower() for cell in row[data_start : data_start + 12])
    if not text:
        return True
    return text.startswith("cus total") or text.startswith("grand total")


def parse_invoice_row(row, row_number, report_date):
    row = clean_row(row)
    data_start = find_data_start(row)
    if data_start is None:
        return None, "No invoice header/data section detected."
    if len(row) < data_start + 12 or row_is_trailer(row, data_start):
        return None, "Skipped non-invoice row."

    cells = row[data_start : data_start + 12]
    customer_code, customer_name, place, bill_number, invoice_date_raw = cells[:5]
    if not customer_code and not customer_name:
        return None, "Missing customer."
    if not bill_number and normalize_header(customer_code) == "code":
        return None, "Skipped repeated header."
    invoice_number, lpo_reference = split_bill_reference(bill_number)

    warnings = []
    invoice_date = parse_date(invoice_date_raw)
    if not invoice_date:
        warnings.append("Invalid or missing invoice date.")

    try:
        amount = parse_decimal(cells[5])
        bucket_0_30 = parse_decimal(cells[6])
        bucket_30_60 = parse_decimal(cells[7])
        bucket_60_90 = parse_decimal(cells[8])
        bucket_over_90 = parse_decimal(cells[9])
        total = parse_decimal(cells[10])
    except ValueError as exc:
        return None, str(exc)

    days = parse_int(cells[11])
    if days == 0 and invoice_date:
        anchor = report_date or timezone.localdate()
        days = max((anchor - invoice_date).days, 0)
        if not report_date:
            warnings.append("Report date missing; calculated days using upload date.")
    elif not cells[11]:
        warnings.append("Days column missing; calculated days if invoice date was available.")

    return ParsedInvoiceRow(
        source_row_number=row_number,
        customer_code=customer_code.strip(),
        customer_name=customer_name.strip(),
        place=place.strip(),
        bill_number=bill_number.strip(),
        invoice_number=invoice_number,
        lpo_reference=lpo_reference,
        invoice_date=invoice_date,
        amount=amount,
        bucket_0_30=bucket_0_30,
        bucket_30_60=bucket_30_60,
        bucket_60_90=bucket_60_90,
        bucket_over_90=bucket_over_90,
        total=total,
        days=days,
        raw_data={
            "source_row": row_number,
            "raw_bill_number": bill_number,
        },
        warnings=warnings,
    ), ""


def validate_row_shape(row, row_number):
    if len(row) > max_import_columns():
        raise ValidationError(
            f"Row {row_number} has too many columns for the Accounting import parser. "
            "Please export the agewise outstanding report as a standard CSV/XLSX file."
        )


def parse_csv_rows(data):
    text = data.decode("utf-8-sig", errors="replace")
    rows = []
    try:
        for row_number, row in enumerate(csv.reader(io.StringIO(text)), start=1):
            if row_number > max_import_rows():
                raise ValidationError(f"File has too many rows. Maximum supported rows: {max_import_rows()}.")
            validate_row_shape(row, row_number)
            rows.append(row)
    except csv.Error as exc:
        raise ValidationError(f"Invalid CSV file: {exc}") from exc
    return rows


def parse_xlsx_rows(data):
    workbook = load_xlsx_workbook(data)
    rows = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if len(rows) + 1 > max_import_rows():
                raise ValidationError(f"File has too many rows. Maximum supported rows: {max_import_rows()}.")
            cleaned = list(row)
            validate_row_shape(cleaned, len(rows) + 1)
            rows.append(cleaned)
    return rows


def load_xlsx_workbook(data):
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
        raise ValidationError("Invalid Excel workbook. Please upload a valid .xlsx file.") from exc


def parse_outstanding_upload(uploaded_file=None, *, source=None):
    source = source or read_outstanding_source(uploaded_file)
    rows = parse_csv_rows(source.data) if source.extension == ".csv" else parse_xlsx_rows(source.data)
    report_date = None
    parsed_rows = []
    warnings = []
    skip_reasons = Counter()

    for row_number, row in enumerate(rows, start=1):
        if report_date is None:
            report_date = parse_report_date(row)
        parsed, reason = parse_invoice_row(row, row_number, report_date)
        if parsed:
            parsed_rows.append(parsed)
        else:
            skip_reasons[reason or "Skipped row."] += 1

    if not parsed_rows:
        raise ValidationError("No usable invoice rows were found in this file.")
    if report_date is None:
        report_date = timezone.localdate()
        warnings.append("Report date was not found; upload date was used.")

    for reason, count in skip_reasons.most_common(8):
        if reason and reason not in {"Skipped non-invoice row.", "No invoice header/data section detected."}:
            warnings.append(f"{count} rows skipped: {reason}")

    return ParsedOutstanding(
        filename=source.filename,
        sha256=source.sha256,
        size=source.size,
        report_date=report_date,
        rows=parsed_rows,
        skipped_row_count=sum(skip_reasons.values()),
        warnings=warnings,
        parse_meta={
            "extension": source.extension,
            "total_input_rows": len(rows),
            "skip_reasons": dict(skip_reasons.most_common(20)),
        },
    )


def normalize_category(value):
    text = normalize_customer_name(value)
    return CATEGORY_VALUES.get(text, "unknown")


def parse_category_upload(uploaded_file):
    if not uploaded_file:
        return None
    filename, extension, data, sha256 = read_upload(uploaded_file, SUPPORTED_CATEGORY_EXTENSIONS)
    workbook = load_xlsx_workbook(data)
    entries = {}
    code_entries = {}
    warnings = []
    sheet_meta = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_row = None
        code_index = None
        name_index = None
        category_index = None
        for index, row in enumerate(rows[:30]):
            normalized = [normalize_header(cell) for cell in row]
            for col, header in enumerate(normalized):
                if header in {"code", "customer code", "cust code", "account code"}:
                    code_index = col
                if header in {"cust name", "customer name", "name", "party"}:
                    name_index = col
                if header in {"cat", "category", "type"}:
                    category_index = col
            if name_index is not None and category_index is not None:
                header_row = index
                break
        if header_row is None:
            sheet_meta.append({"sheet": sheet.title, "selected": False, "reason": "No customer/category header."})
            continue
        selected_count = 0
        for row in rows[header_row + 1 :]:
            code = row[code_index] if code_index is not None and len(row) > code_index else ""
            name = row[name_index] if len(row) > name_index else ""
            category = row[category_index] if len(row) > category_index else ""
            normalized_code = str(code or "").strip().lower()
            normalized_name = normalize_customer_name(name)
            normalized_category = normalize_category(category)
            if not normalized_name and not normalized_code:
                continue
            if normalized_category == "unknown":
                warnings.append(f"Unknown category '{category}' for {name}.")
            if normalized_code:
                code_entries[normalized_code] = normalized_category
            if normalized_name:
                entries[normalized_name] = normalized_category
            selected_count += 1
        sheet_meta.append({"sheet": sheet.title, "selected": True, "header_row": header_row + 1, "rows": selected_count})

    return ParsedCategoryMap(
        filename=filename,
        sha256=sha256,
        entries=entries,
        code_entries=code_entries,
        warnings=warnings[:50],
        parse_meta={
            "sheets": sheet_meta,
            "entry_count": len(entries),
            "code_entry_count": len(code_entries),
            "extension": extension,
        },
    )
