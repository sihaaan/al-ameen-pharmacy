import csv
from io import BytesIO, StringIO
from unittest.mock import patch
from zipfile import ZipFile

from django.contrib.auth.models import Group, Permission, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from rest_framework.test import APITestCase

from .models import AccountCustomer, AccountingBlocklistedCustomer, AccountingImport, AccountingImportCustomer
from .parsers import normalize_customer_name, parse_outstanding_upload, split_bill_reference
from .permissions import accounting_permissions_queryset, set_user_accounting_access


def make_agewise_row(code, party, bill_no, invoice_date, amount, b0, b30, b60, b90, total, days):
    return [
        "AL AMEEN PHARMACY",
        "Frij Murar, Somali St, Deira, Dubai.",
        "Agewise Outstanding as on 25/05/26",
        "Page -1 of 1",
        "Code",
        "Party",
        "Place",
        "Bill No.",
        "Date",
        "Amount",
        "0-30",
        "30 - 60",
        "60 - 90",
        "Over 90",
        "TOTAL",
        "Days",
        code,
        party,
        "",
        bill_no,
        invoice_date,
        amount,
        b0,
        b30,
        b60,
        b90,
        total,
        days,
        "37071549.94",
        "586650.57",
    ]


def make_repeated_header_export_row(code, party, bill_no, invoice_date, amount, b0, b30, b60, b90, total, days):
    return [
        "AL AMEEN PHARMACY",
        "Frij Murar, Somali St, Deira, Dubai. Tel: 04 271 3695, Fax: 273 1737.",
        "Agewise Outstanding from 01/01/20 to 01/06/26",
        "Page -1 of 1",
        "Code",
        "Party",
        "Place",
        "Bill No.",
        "Date",
        "Amount",
        "0-30",
        "30 - 60",
        "60 - 90",
        "Over 90",
        "TOTAL",
        "Days",
        code,
        party,
        "",
        bill_no,
        invoice_date,
        amount,
        b0,
        b30,
        b60,
        b90,
        total,
        days,
        "27312116.19",
        "544868.44",
        "474621.80",
        "444934.45",
        "25646829.75",
        "27111254.44",
        "",
    ]


def make_agewise_upload(name="ageoutcode test.csv", marker=""):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(make_agewise_row("083", "MILLENNIUM AIRPORT HOTEL", "571920-UA3IJ2-", "03/02/2026", "1557.70", "0.00", "0.00", "1557.70", "0.00", "1557.70", "80"))
    writer.writerow(make_agewise_row("084", "CARD CUSTOMER", f"571921-{marker}", "20/05/2026", "100.00", "100.00", "0.00", "0.00", "0.00", "100.00", "5"))
    writer.writerow(make_agewise_row("085", "CREDIT NOTE CUSTOMER", "CR-1", "03/02/2026", "(126.00", "0.00", "0.00", "(126.00", "0.00", "(126.00", "80"))
    return SimpleUploadedFile(name, buffer.getvalue().encode("utf-8"), content_type="text/csv")


def make_repeated_header_export_upload(name="ageoutcode.csv"):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(make_repeated_header_export_row("001", "EMRILL CO", "569112-PO165273-", "17/01/2026", "62.50", "0.00", "0.00", "0.00", "62.40", "62.40", "135.00"))
    writer.writerow(make_repeated_header_export_row("001", "EMRILL CO", "579225-PO173953-", "27/02/2026", "275.00", "0.00", "0.00", "0.00", "275.00", "275.00", "94.00"))
    writer.writerow(make_repeated_header_export_row("PL", "P/LABEL", "C.Card-", "04/02/2020", "300.00", "0.00", "0.00", "0.00", "300.00", "300.00", "2309.00"))
    return SimpleUploadedFile(name, buffer.getvalue().encode("utf-8"), content_type="text/csv")


def make_ledger_upload(name="ledger.csv"):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(make_agewise_row("083", "MILLENNIUM AIRPORT HOTEL", "570170-PO-100-", "01/01/2026", "100.00", "0.00", "100.00", "0.00", "0.00", "100.00", "145"))
    writer.writerow(make_agewise_row("083", "MILLENNIUM AIRPORT HOTEL", "570171-PO-200-", "15/02/2026", "500.00", "0.00", "0.00", "500.00", "0.00", "500.00", "99"))
    writer.writerow(make_agewise_row("083", "MILLENNIUM AIRPORT HOTEL", "570172-CN-25-", "20/02/2026", "(25.00", "0.00", "0.00", "(25.00", "0.00", "(25.00", "94"))
    writer.writerow(make_agewise_row("084", "CARD CUSTOMER", "571921-", "20/05/2026", "100.00", "100.00", "0.00", "0.00", "0.00", "100.00", "5"))
    return SimpleUploadedFile(name, buffer.getvalue().encode("utf-8"), content_type="text/csv")


def make_category_upload():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cat"
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append([None, "Cust Name", "Cat"])
    sheet.append([None, "MILLENNIUM AIRPORT HOTEL", "Credit"])
    sheet.append([None, "CARD CUSTOMER", "Card"])
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile("ageoutcode may 2026.xlsx", buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def make_code_category_upload():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cat"
    sheet.append(["Customer Code", "Cust Name", "Cat"])
    sheet.append(["A01", "DUPLICATE NAME LLC", "Credit"])
    sheet.append(["B02", "DUPLICATE NAME LLC", "Card"])
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile("categories-by-code.xlsx", buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def make_blocklist_upload():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "blocklist"
    sheet.append(["Name", "Cat"])
    sheet.append(["MILLENNIUM AIRPORT HOTEL", "Misc"])
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile("blocklist.xlsx", buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def make_serial_blocklist_upload():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "blocklist"
    sheet.append(["S. No.", "Company Name", "Category"])
    sheet.append([1, "INTERMASS", "Branch"])
    sheet.append([2, "MILLENNIUM AIRPORT HOTEL", "Misc"])
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile("blocklist-with-serial.xlsx", buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class AccountingParserTests(TestCase):
    def test_parse_sample_style_csv_and_accounting_negatives(self):
        parsed = parse_outstanding_upload(make_agewise_upload())
        self.assertEqual(parsed.report_date.isoformat(), "2026-05-25")
        self.assertEqual(len(parsed.rows), 3)
        first = parsed.rows[0]
        self.assertEqual(first.customer_code, "083")
        self.assertEqual(first.customer_name, "MILLENNIUM AIRPORT HOTEL")
        self.assertEqual(first.bill_number, "571920-UA3IJ2-")
        self.assertEqual(first.invoice_number, "571920")
        self.assertEqual(first.lpo_reference, "UA3IJ2")
        self.assertEqual(first.days, 80)
        self.assertEqual(str(first.bucket_60_90), "1557.70")
        second = parsed.rows[1]
        self.assertEqual(second.invoice_number, "571921")
        self.assertEqual(second.lpo_reference, "")
        credit = parsed.rows[2]
        self.assertEqual(str(credit.total), "-126.00")

    def test_parse_repeated_header_export_csv_shape(self):
        parsed = parse_outstanding_upload(make_repeated_header_export_upload())
        self.assertEqual(parsed.report_date.isoformat(), "2026-06-01")
        self.assertEqual(len(parsed.rows), 3)
        self.assertEqual(parsed.skipped_row_count, 0)
        self.assertEqual(parsed.rows[0].customer_code, "001")
        self.assertEqual(parsed.rows[0].customer_name, "EMRILL CO")
        self.assertEqual(parsed.rows[0].invoice_number, "569112")
        self.assertEqual(parsed.rows[0].lpo_reference, "PO165273")
        self.assertEqual(parsed.rows[0].days, 135)
        self.assertEqual(parsed.rows[2].customer_name, "P/LABEL")
        self.assertNotIn("Report date was not found", " ".join(parsed.warnings))

    def test_bill_number_reference_split_is_clean(self):
        examples = {
            "570170-284750-0-": ("570170", "284750-0"),
            "320815--": ("320815", ""),
            "571920-UA3IJ2-": ("571920", "UA3IJ2"),
            "571921-": ("571921", ""),
        }
        for raw, expected in examples.items():
            with self.subTest(raw=raw):
                self.assertEqual(split_bill_reference(raw), expected)


class AccountingAPITests(APITestCase):
    def setUp(self):
        self.customer = User.objects.create_user(username="customer", password="pass")
        self.staff = User.objects.create_user(username="staff", password="pass", is_staff=True)
        self.accountant = User.objects.create_user(username="accountant", password="pass", is_staff=True)
        self.superuser = User.objects.create_superuser(username="owner", password="pass", email="owner@example.com")
        perms = Permission.objects.filter(
            content_type__app_label="accounting",
            codename__in=[
                "view_accounting_module",
                "upload_accounting_statement",
                "generate_accounting_statement",
                "edit_accounting_customer",
                "download_accounting_statement",
            ],
        )
        self.accountant.user_permissions.add(*perms)

    def upload_import(self):
        self.client.force_authenticate(self.accountant)
        return self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_agewise_upload(), "category_file": make_category_upload()},
            format="multipart",
        )

    def upload_ledger_import(self):
        self.client.force_authenticate(self.accountant)
        return self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_ledger_upload(), "category_file": make_category_upload()},
            format="multipart",
        )

    def test_accounting_permissions(self):
        dashboard_url = reverse("accounting-dashboard")
        anon = self.client.get(dashboard_url)
        self.assertEqual(anon.status_code, 401)

        self.client.force_authenticate(self.customer)
        self.assertEqual(self.client.get(dashboard_url).status_code, 403)

        self.client.force_authenticate(self.staff)
        self.assertEqual(self.client.get(dashboard_url).status_code, 403)

        self.client.force_authenticate(self.accountant)
        self.assertEqual(self.client.get(dashboard_url).status_code, 200)

        self.client.force_authenticate(self.superuser)
        self.assertEqual(self.client.get(dashboard_url).status_code, 200)

    def test_accounting_group_staff_user_allowed(self):
        group_user = User.objects.create_user(username="group_accountant", password="pass", is_staff=True)
        group = Group.objects.create(name="Accounting")
        group_user.groups.add(group)
        self.client.force_authenticate(group_user)
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 200)

    def test_accounting_access_toggle_manages_only_accounting_access(self):
        user = User.objects.create_user(username="toggle_user", password="pass", is_staff=True)
        unrelated_group = Group.objects.create(name="Inventory")
        unrelated_permission = Permission.objects.filter(codename="view_user").first()
        user.groups.add(unrelated_group)
        if unrelated_permission:
            user.user_permissions.add(unrelated_permission)

        self.client.force_authenticate(user)
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 403)

        set_user_accounting_access(user, True)
        user = User.objects.get(pk=user.pk)
        self.client.force_authenticate(user)
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 200)

        set_user_accounting_access(user, False)
        user = User.objects.get(pk=user.pk)
        self.client.force_authenticate(user)
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 403)
        self.assertTrue(user.groups.filter(name="Inventory").exists())
        if unrelated_permission:
            self.assertTrue(user.user_permissions.filter(pk=unrelated_permission.pk).exists())
        self.assertFalse(user.groups.filter(name="Accounting").exists())
        self.assertFalse(user.user_permissions.filter(pk__in=accounting_permissions_queryset().values("pk")).exists())

    def test_existing_permission_group_still_allows_accounting(self):
        user = User.objects.create_user(username="custom_group_user", password="pass", is_staff=True)
        group = Group.objects.create(name="Finance Team")
        group.permissions.add(Permission.objects.get(codename="view_accounting_module", content_type__app_label="accounting"))
        user.groups.add(group)
        self.client.force_authenticate(User.objects.get(pk=user.pk))
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 200)

    def test_accounting_access_still_requires_staff_unless_superuser(self):
        user = User.objects.create_user(username="nonstaff_accounting", password="pass", is_staff=False)
        set_user_accounting_access(user, True)
        self.client.force_authenticate(User.objects.get(pk=user.pk))
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 403)

        self.client.force_authenticate(self.superuser)
        self.assertEqual(self.client.get(reverse("accounting-dashboard")).status_code, 200)

    def test_django_user_admin_shows_accounting_access_checkbox(self):
        self.client.force_login(self.superuser)
        response = self.client.get(reverse("admin:auth_user_change", args=[self.staff.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Accounting access")

    def test_upload_groups_customers_matches_categories_and_persists_email(self):
        response = self.upload_import()
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["parsed_row_count"], 3)
        self.assertEqual(response.data["customer_count"], 3)
        self.assertEqual(response.data["due_customer_count"], 2)
        self.assertEqual(response.data["report_date_display"], "25/05/2026")

        customer = AccountCustomer.objects.get(customer_code="083")
        self.assertEqual(customer.category, "credit")

        summary = AccountingImportCustomer.objects.get(customer=customer)
        patch = self.client.patch(
            reverse("accounting-import-customer-detail", args=[summary.id]),
            {"email": "accounts@hotel.example", "category": "clinic", "is_ignored": False},
            format="json",
        )
        self.assertEqual(patch.status_code, 200)
        customer.refresh_from_db()
        self.assertEqual(customer.email, "accounts@hotel.example")
        self.assertEqual(customer.category, "clinic")

        second = self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_agewise_upload("new_month.csv", marker="new")},
            format="multipart",
        )
        self.assertEqual(second.status_code, 201)
        self.assertEqual(
            AccountingImportCustomer.objects.filter(customer__customer_code="083").order_by("-id").first().email,
            "accounts@hotel.example",
        )

    def test_duplicate_upload_can_apply_category_workbook_to_existing_import(self):
        self.client.force_authenticate(self.accountant)
        first = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")
        self.assertEqual(first.status_code, 201)
        self.assertEqual(AccountCustomer.objects.get(customer_code="083").category, "unknown")

        second = self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_agewise_upload(), "category_file": make_category_upload()},
            format="multipart",
        )

        self.assertEqual(second.status_code, 200)
        self.assertTrue(second.data["duplicate"])
        self.assertEqual(AccountingImport.objects.count(), 1)
        self.assertGreaterEqual(second.data["category_update"]["matched"], 2)
        self.assertIn("unchanged", second.data["category_update"])
        self.assertEqual(AccountCustomer.objects.get(customer_code="083").category, "credit")
        self.assertEqual(AccountingImportCustomer.objects.get(customer__customer_code="084").category, "card")

        third = self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_agewise_upload(), "category_file": make_category_upload()},
            format="multipart",
        )
        self.assertEqual(third.status_code, 200)
        self.assertEqual(third.data["category_update"]["updated"], 0)
        self.assertGreaterEqual(third.data["category_update"]["unchanged"], 2)
        self.assertIn("already up to date", third.data["category_update_message"])

    def test_category_workbook_can_be_applied_to_existing_import(self):
        self.client.force_authenticate(self.accountant)
        response = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")
        self.assertEqual(response.status_code, 201)

        apply_response = self.client.post(
            reverse("accounting-import-apply-categories", args=[response.data["id"]]),
            {"category_file": make_category_upload()},
            format="multipart",
        )

        self.assertEqual(apply_response.status_code, 200)
        self.assertEqual(AccountCustomer.objects.get(customer_code="083").category, "credit")

    def test_blocklist_upload_excludes_matching_customers_from_due_counts(self):
        self.client.force_authenticate(self.accountant)
        response = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["due_customer_count"], 2)

        block_response = self.client.post(
            reverse("accounting-import-apply-blocklist"),
            {"file": make_blocklist_upload()},
            format="multipart",
        )

        self.assertEqual(block_response.status_code, 200)
        self.assertEqual(AccountingBlocklistedCustomer.objects.count(), 1)
        summary = AccountingImportCustomer.objects.get(customer__customer_code="083")
        summary.refresh_from_db()
        summary.customer.refresh_from_db()
        summary.accounting_import.refresh_from_db()
        self.assertTrue(summary.customer.is_ignored)
        self.assertTrue(summary.is_ignored)
        self.assertFalse(summary.is_due)
        self.assertEqual(summary.status, AccountingImportCustomer.STATUS_IGNORED)
        self.assertEqual(summary.accounting_import.due_customer_count, 1)

    def test_blocklist_entries_can_be_managed_and_apply_to_existing_imports(self):
        self.client.force_authenticate(self.accountant)
        upload_response = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")
        self.assertEqual(upload_response.status_code, 201)

        create_response = self.client.post(
            reverse("accounting-blocklist-list"),
            {"name": "MILLENNIUM AIRPORT HOTEL", "category_hint": "Internal branch"},
            format="json",
        )

        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(create_response.data["blocklist_update"]["matched_customers"], 1)
        self.assertEqual(AccountingBlocklistedCustomer.objects.filter(is_active=True).count(), 1)
        summary = AccountingImportCustomer.objects.get(customer__name="MILLENNIUM AIRPORT HOTEL")
        summary.accounting_import.refresh_from_db()
        self.assertTrue(summary.is_ignored)
        self.assertEqual(summary.accounting_import.due_customer_count, 1)

        update_response = self.client.patch(
            reverse("accounting-blocklist-detail", args=[create_response.data["id"]]),
            {"category_hint": "Do not send"},
            format="json",
        )

        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data["category_hint"], "Do not send")

        delete_response = self.client.delete(reverse("accounting-blocklist-detail", args=[create_response.data["id"]]))
        self.assertEqual(delete_response.status_code, 204)
        self.assertFalse(AccountingBlocklistedCustomer.objects.get(pk=create_response.data["id"]).is_active)

    def test_blocklist_applies_to_clear_customer_name_variants(self):
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(make_agewise_row("901", "INTERMASS TRADING LLC - BRANCH 01", "INV-1", "03/02/2026", "100.00", "0.00", "100.00", "0.00", "0.00", "100.00", "80"))
        writer.writerow(make_agewise_row("902", "GENERIC CLINIC CUSTOMER", "INV-2", "03/02/2026", "200.00", "0.00", "200.00", "0.00", "0.00", "200.00", "80"))
        upload = SimpleUploadedFile("blocklist-variants.csv", buffer.getvalue().encode("utf-8"), content_type="text/csv")

        self.client.force_authenticate(self.accountant)
        upload_response = self.client.post(reverse("accounting-import-upload"), {"file": upload}, format="multipart")
        self.assertEqual(upload_response.status_code, 201)
        self.assertEqual(upload_response.data["due_customer_count"], 2)

        block_response = self.client.post(
            reverse("accounting-blocklist-list"),
            {"name": "INTERMASS", "category_hint": "Do not send"},
            format="json",
        )

        self.assertEqual(block_response.status_code, 201)
        self.assertEqual(block_response.data["blocklist_update"]["matched_customers"], 1)
        blocked = AccountingImportCustomer.objects.get(customer__customer_code="901")
        allowed = AccountingImportCustomer.objects.get(customer__customer_code="902")
        blocked.accounting_import.refresh_from_db()
        self.assertTrue(blocked.is_ignored)
        self.assertFalse(blocked.is_due)
        self.assertFalse(allowed.is_ignored)
        self.assertTrue(allowed.is_due)
        self.assertEqual(blocked.accounting_import.due_customer_count, 1)

    def test_generic_single_word_blocklist_does_not_overmatch_variants(self):
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(make_agewise_row("903", "DUBAI CLINIC CUSTOMER", "INV-1", "03/02/2026", "100.00", "0.00", "100.00", "0.00", "0.00", "100.00", "80"))
        upload = SimpleUploadedFile("generic-blocklist.csv", buffer.getvalue().encode("utf-8"), content_type="text/csv")

        self.client.force_authenticate(self.accountant)
        upload_response = self.client.post(reverse("accounting-import-upload"), {"file": upload}, format="multipart")
        self.assertEqual(upload_response.status_code, 201)

        block_response = self.client.post(
            reverse("accounting-blocklist-list"),
            {"name": "CLINIC", "category_hint": "Too broad"},
            format="json",
        )

        self.assertEqual(block_response.status_code, 201)
        self.assertEqual(block_response.data["blocklist_update"]["matched_customers"], 0)
        summary = AccountingImportCustomer.objects.get(customer__customer_code="903")
        self.assertFalse(summary.is_ignored)
        self.assertTrue(summary.is_due)

    def test_existing_blocklist_variant_applies_to_future_imports(self):
        self.client.force_authenticate(self.accountant)
        self.client.post(
            reverse("accounting-blocklist-list"),
            {"name": "INTERMASS", "category_hint": "Do not send"},
            format="json",
        )
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(make_agewise_row("904", "INTERMASS TRADING LLC - BRANCH 02", "INV-1", "03/02/2026", "100.00", "0.00", "100.00", "0.00", "0.00", "100.00", "80"))
        upload = SimpleUploadedFile("future-blocklist.csv", buffer.getvalue().encode("utf-8"), content_type="text/csv")

        upload_response = self.client.post(reverse("accounting-import-upload"), {"file": upload}, format="multipart")

        self.assertEqual(upload_response.status_code, 201)
        self.assertEqual(upload_response.data["due_customer_count"], 0)
        summary = AccountingImportCustomer.objects.get(customer__customer_code="904")
        self.assertTrue(summary.is_ignored)
        self.assertFalse(summary.is_due)

    def test_serial_number_blocklist_workbook_applies_to_future_imports(self):
        self.client.force_authenticate(self.accountant)
        block_response = self.client.post(
            reverse("accounting-import-apply-blocklist"),
            {"file": make_serial_blocklist_upload()},
            format="multipart",
        )
        self.assertEqual(block_response.status_code, 200)
        self.assertEqual(block_response.data["blocklist_update"]["loaded"], 2)
        self.assertTrue(AccountingBlocklistedCustomer.objects.filter(normalized_name="intermass").exists())

        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(make_agewise_row("904", "INTERMASS TRADING LLC - BRANCH 02", "INV-1", "03/02/2026", "100.00", "0.00", "100.00", "0.00", "0.00", "100.00", "80"))
        upload = SimpleUploadedFile("future-serial-blocklist.csv", buffer.getvalue().encode("utf-8"), content_type="text/csv")

        upload_response = self.client.post(reverse("accounting-import-upload"), {"file": upload}, format="multipart")

        self.assertEqual(upload_response.status_code, 201)
        self.assertEqual(upload_response.data["due_customer_count"], 0)
        summary = AccountingImportCustomer.objects.get(customer__customer_code="904")
        self.assertTrue(summary.is_ignored)
        self.assertEqual(summary.status, AccountingImportCustomer.STATUS_IGNORED)

    def test_duplicate_import_reapplies_current_active_blocklist(self):
        self.client.force_authenticate(self.accountant)
        first = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.data["due_customer_count"], 2)
        AccountingBlocklistedCustomer.objects.create(
            name="MILLENNIUM AIRPORT HOTEL",
            normalized_name=normalize_customer_name("MILLENNIUM AIRPORT HOTEL"),
            category_hint="Block after original import",
            is_active=True,
        )

        duplicate = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")

        self.assertEqual(duplicate.status_code, 200)
        self.assertTrue(duplicate.data["duplicate"])
        self.assertEqual(duplicate.data["due_customer_count"], 1)
        self.assertEqual(duplicate.data["blocklist_update"]["matched_import_customers"], 1)
        summary = AccountingImportCustomer.objects.get(customer__customer_code="083")
        self.assertTrue(summary.is_ignored)
        self.assertFalse(summary.is_due)

    def test_same_customer_name_with_different_codes_stays_separate(self):
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(make_agewise_row("A01", "DUPLICATE NAME LLC", "INV-1", "03/02/2026", "10.00", "0.00", "10.00", "0.00", "0.00", "10.00", "80"))
        writer.writerow(make_agewise_row("B02", "DUPLICATE NAME LLC", "INV-2", "03/02/2026", "20.00", "0.00", "20.00", "0.00", "0.00", "20.00", "80"))
        upload = SimpleUploadedFile("duplicate-names.csv", buffer.getvalue().encode("utf-8"), content_type="text/csv")

        self.client.force_authenticate(self.accountant)
        response = self.client.post(reverse("accounting-import-upload"), {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["customer_count"], 2)
        self.assertEqual(AccountCustomer.objects.filter(name="DUPLICATE NAME LLC").count(), 2)
        self.assertEqual(AccountingImportCustomer.objects.count(), 2)

    def test_category_matching_uses_customer_code_before_name(self):
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(make_agewise_row("A01", "DUPLICATE NAME LLC", "INV-1", "03/02/2026", "10.00", "0.00", "10.00", "0.00", "0.00", "10.00", "80"))
        writer.writerow(make_agewise_row("B02", "DUPLICATE NAME LLC", "INV-2", "03/02/2026", "20.00", "0.00", "20.00", "0.00", "0.00", "20.00", "80"))
        upload = SimpleUploadedFile("duplicate-names.csv", buffer.getvalue().encode("utf-8"), content_type="text/csv")

        self.client.force_authenticate(self.accountant)
        response = self.client.post(
            reverse("accounting-import-upload"),
            {"file": upload, "category_file": make_code_category_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(AccountCustomer.objects.get(customer_code="A01").category, "credit")
        self.assertEqual(AccountCustomer.objects.get(customer_code="B02").category, "card")

    def test_csv_outstanding_and_category_workbook_upload_paths_do_not_500(self):
        self.client.force_authenticate(self.accountant)
        upload_response = self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_agewise_upload("monthly.csv"), "category_file": make_category_upload()},
            format="multipart",
        )
        self.assertEqual(upload_response.status_code, 201)
        self.assertEqual(upload_response.data["category_update"]["matched"], 2)
        self.assertEqual(upload_response.data["category_update"]["unchanged"], 2)
        self.assertEqual(upload_response.data["category_update"]["updated"], 0)
        self.assertIn("Category workbook applied.", upload_response.data["category_update_message"])
        self.assertEqual(AccountCustomer.objects.get(customer_code="083").category, "credit")

        apply_response = self.client.post(
            reverse("accounting-import-apply-categories", args=[upload_response.data["id"]]),
            {"category_file": make_category_upload()},
            format="multipart",
        )
        self.assertEqual(apply_response.status_code, 200)
        self.assertIn("category_update", apply_response.data)

        bad_workbook = SimpleUploadedFile(
            "bad-category.xlsx",
            b"not a real xlsx file",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        bad_response = self.client.post(
            reverse("accounting-import-apply-categories", args=[upload_response.data["id"]]),
            {"category_file": bad_workbook},
            format="multipart",
        )
        self.assertEqual(bad_response.status_code, 400)

        malformed_csv = SimpleUploadedFile("bad.csv", b"not,the,agewise,export\njust,noise", content_type="text/csv")
        malformed_response = self.client.post(
            reverse("accounting-import-upload"),
            {"file": malformed_csv},
            format="multipart",
        )
        self.assertEqual(malformed_response.status_code, 400)
        self.assertIn("No usable invoice rows", str(malformed_response.data["detail"]))

    @override_settings(ACCOUNTING_IMPORT_MAX_ROWS=2)
    def test_oversized_row_count_returns_clean_validation_error(self):
        self.client.force_authenticate(self.accountant)
        response = self.client.post(
            reverse("accounting-import-upload"),
            {"file": make_repeated_header_export_upload()},
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("too many rows", str(response.data["detail"]).lower())

    def test_duplicate_upload_returns_previous_import_without_creating_another(self):
        first = self.upload_import()
        self.assertEqual(first.status_code, 201)
        with patch("accounting.services.parse_outstanding_upload") as parse_mock:
            second = self.client.post(reverse("accounting-import-upload"), {"file": make_agewise_upload()}, format="multipart")
        self.assertEqual(second.status_code, 200)
        self.assertTrue(second.data["duplicate"])
        self.assertEqual(AccountingImport.objects.count(), 1)
        parse_mock.assert_not_called()

    def test_statement_pdfs_and_zip_download_are_staff_only_and_customer_facing(self):
        response = self.upload_import()
        import_id = response.data["id"]
        summary = AccountingImportCustomer.objects.get(customer__customer_code="083")

        self.client.force_authenticate(self.staff)
        self.assertEqual(self.client.get(reverse("accounting-import-customer-statement-pdf", args=[summary.id])).status_code, 403)

        self.client.force_authenticate(self.accountant)
        pdf = self.client.get(reverse("accounting-import-customer-statement-pdf", args=[summary.id]))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertGreater(len(pdf.content), 500)
        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages)
        self.assertIn("Invoice No.", text)
        self.assertIn("LPO / Reference No.", text)
        self.assertIn("Debit", text)
        self.assertIn("Credit", text)
        self.assertIn("Balance", text)
        self.assertIn("Days", text)
        self.assertIn("25/05/2026", text)
        self.assertNotIn("PDC", text)
        self.assertNotIn("PDC Value", text)
        self.assertIn("571920", text)
        self.assertIn("UA3IJ2", text)
        self.assertNotIn("0-30", text)
        self.assertNotIn("30-60", text)
        self.assertNotIn("60-90", text)
        self.assertNotIn("Over 90", text)
        self.assertNotIn("Email missing", text)
        self.assertNotIn("Category", text)

        zip_response = self.client.get(reverse("accounting-import-statements-zip", args=[import_id]))
        self.assertEqual(zip_response.status_code, 200)
        self.assertEqual(zip_response["Content-Type"], "application/zip")
        with ZipFile(BytesIO(zip_response.content)) as archive:
            self.assertGreaterEqual(len(archive.namelist()), 1)

    def test_ledger_running_balance_and_date_range_outputs(self):
        response = self.upload_ledger_import()
        import_id = response.data["id"]
        summary = AccountingImportCustomer.objects.get(customer__customer_code="083")
        self.client.force_authenticate(self.accountant)

        detail = self.client.get(reverse("accounting-import-customer-detail", args=[summary.id]))
        self.assertEqual(detail.status_code, 200)
        balances = [row["balance"] for row in detail.data["ledger_rows"]]
        self.assertEqual(balances, ["100.00", "600.00", "575.00"])
        self.assertEqual(detail.data["ledger_rows"][-1]["credit"], "25.00")
        self.assertEqual(detail.data["ledger_rows"][0]["invoice_date"], "2026-01-01")
        self.assertEqual(detail.data["ledger_rows"][0]["invoice_date_display"], "01/01/2026")
        self.assertEqual(detail.data["ledger_rows"][0]["days"], 144)
        self.assertEqual(detail.data["ledger_rows"][1]["days"], 99)
        self.assertEqual(detail.data["total_outstanding"], "575.00")
        self.assertEqual(detail.data["statement_period"]["display_from"], "01/01/2026")
        self.assertEqual(detail.data["statement_period"]["display_to"], "20/02/2026")
        self.assertEqual(detail.data["statement_period"]["display"], "01/01/2026 to 20/02/2026")

        full_pdf = self.client.get(reverse("accounting-import-customer-statement-pdf", args=[summary.id]))
        self.assertEqual(full_pdf.status_code, 200)
        full_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(full_pdf.content)).pages)
        self.assertIn("01/01/2026 to 20/02/2026", full_text)
        self.assertIn("25/05/2026", full_text)
        self.assertIn("Days", full_text)
        self.assertIn("Credit", full_text)
        self.assertIn("AED 25.00", full_text)
        self.assertNotIn("PDC", full_text)

        filtered_detail = self.client.get(
            reverse("accounting-import-customer-detail", args=[summary.id]),
            {"date_from": "2026-02-01", "date_to": "2026-02-28"},
        )
        self.assertEqual(filtered_detail.status_code, 200)
        self.assertEqual(len(filtered_detail.data["ledger_rows"]), 2)
        self.assertEqual(filtered_detail.data["ledger_rows"][-1]["balance"], "475.00")
        self.assertEqual(filtered_detail.data["total_outstanding"], "475.00")
        self.assertEqual(filtered_detail.data["statement_period"]["display_from"], "01/02/2026")
        self.assertEqual(filtered_detail.data["statement_period"]["display_to"], "28/02/2026")
        self.assertEqual(filtered_detail.data["statement_period"]["display"], "01/02/2026 to 28/02/2026")

        one_day_detail = self.client.get(
            reverse("accounting-import-customer-detail", args=[summary.id]),
            {"date_from": "2026-02-15", "date_to": "2026-02-15"},
        )
        self.assertEqual(one_day_detail.status_code, 200)
        self.assertEqual(len(one_day_detail.data["ledger_rows"]), 1)
        self.assertEqual(one_day_detail.data["statement_period"]["display_from"], "15/02/2026")
        self.assertEqual(one_day_detail.data["statement_period"]["display_to"], "15/02/2026")
        self.assertEqual(one_day_detail.data["statement_period"]["display"], "15/02/2026")

        filtered_list = self.client.get(
            reverse("accounting-import-customer-list"),
            {"import_id": import_id, "date_from": "01/02/2026", "date_to": "28/02/2026", "due_only": "true"},
        )
        self.assertEqual(filtered_list.status_code, 200)
        hotel = next(item for item in filtered_list.data if item["customer_code"] == "083")
        self.assertEqual(hotel["invoice_count"], 2)
        self.assertEqual(hotel["total_outstanding"], "475.00")
        self.assertEqual(hotel["overdue_amount"], "475.00")

        dash_date_filter = self.client.get(
            reverse("accounting-import-customer-list"),
            {"import_id": import_id, "date_from": "01-02-2026", "date_to": "28-02-2026", "due_only": "true"},
        )
        self.assertEqual(dash_date_filter.status_code, 200)
        self.assertEqual(next(item for item in dash_date_filter.data if item["customer_code"] == "083")["invoice_count"], 2)

        pdf = self.client.get(
            reverse("accounting-import-customer-statement-pdf", args=[summary.id]),
            {"date_from": "2026-02-01", "date_to": "2026-02-28"},
        )
        self.assertEqual(pdf.status_code, 200)
        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages)
        self.assertIn("01/02/2026 to 28/02/2026", text)
        self.assertIn("570171", text)
        self.assertIn("570172", text)
        self.assertNotIn("570170", text)
        self.assertIn("AED 475.00", text)
        self.assertIn("Days", text)
        self.assertNotIn("PDC", text)

        zip_response = self.client.get(
            reverse("accounting-import-statements-zip", args=[import_id]),
            {"date_from": "2026-02-01", "date_to": "2026-02-28"},
        )
        self.assertEqual(zip_response.status_code, 200)
        with ZipFile(BytesIO(zip_response.content)) as archive:
            self.assertTrue(any("MILLENNIUM" in name for name in archive.namelist()))
            pdf_text = "\n".join(
                page.extract_text() or ""
                for page in PdfReader(BytesIO(archive.read(next(name for name in archive.namelist() if "MILLENNIUM" in name)))).pages
            )
        self.assertIn("01/02/2026 to 28/02/2026", pdf_text)
        self.assertIn("570171", pdf_text)
        self.assertNotIn("570170", pdf_text)
        self.assertNotIn("PDC", pdf_text)

        excel = self.client.get(
            reverse("accounting-import-customer-statement-excel", args=[summary.id]),
            {"date_from": "2026-02-01", "date_to": "2026-02-28"},
        )
        self.assertEqual(excel.status_code, 200)
        workbook = load_workbook(BytesIO(excel.content), data_only=True)
        sheet = workbook.active
        header_row = next(
            row[0].row
            for row in sheet.iter_rows()
            if [cell.value for cell in row[:8]] == ["Invoice Date", "Doc Type", "Invoice No.", "LPO / Reference No.", "Debit", "Credit", "Balance", "Days"]
        )
        header_values = [sheet.cell(header_row, column).value for column in range(1, 9)]
        self.assertEqual(header_values, ["Invoice Date", "Doc Type", "Invoice No.", "LPO / Reference No.", "Debit", "Credit", "Balance", "Days"])
        self.assertEqual(sheet.cell(header_row + 1, 1).number_format, "DD/MM/YYYY")
        self.assertEqual(sheet.cell(header_row + 1, 8).value, 99)
        all_values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertNotIn("PDC", all_values)
        self.assertNotIn("0-30", all_values)
        self.assertNotIn("30-60", all_values)
        self.assertIn("01/02/2026 to 28/02/2026", all_values)
        self.assertEqual(sheet.freeze_panes, f"A{header_row + 1}")
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertEqual(len(sheet.tables), 0)
        with ZipFile(BytesIO(excel.content)) as workbook_archive:
            self.assertFalse(any(name.startswith("xl/tables/") for name in workbook_archive.namelist()))

        excel_zip = self.client.get(
            reverse("accounting-import-statements-excel-zip", args=[import_id]),
            {"date_from": "2026-02-01", "date_to": "2026-02-28"},
        )
        self.assertEqual(excel_zip.status_code, 200)
        with ZipFile(BytesIO(excel_zip.content)) as archive:
            excel_names = archive.namelist()
            self.assertTrue(any(name.endswith(".xlsx") and "MILLENNIUM" in name for name in excel_names))

    def test_ignored_customers_are_excluded_from_statement_zip(self):
        response = self.upload_import()
        import_id = response.data["id"]
        summary = AccountingImportCustomer.objects.get(customer__customer_code="083")
        self.client.force_authenticate(self.accountant)
        self.client.patch(
            reverse("accounting-import-customer-detail", args=[summary.id]),
            {"is_ignored": True},
            format="json",
        )

        zip_response = self.client.get(reverse("accounting-import-statements-zip", args=[import_id]))
        self.assertEqual(zip_response.status_code, 200)
        with ZipFile(BytesIO(zip_response.content)) as archive:
            names = archive.namelist()
        self.assertFalse(any("MILLENNIUM" in name for name in names))

    @override_settings(ACCOUNTING_STATEMENT_ZIP_SYNC_LIMIT=1)
    def test_large_statement_zip_is_batched_and_selected_zip_still_works(self):
        response = self.upload_import()
        import_id = response.data["id"]
        summary = AccountingImportCustomer.objects.get(customer__customer_code="083")
        self.client.force_authenticate(self.accountant)

        batched = self.client.get(reverse("accounting-import-statements-zip", args=[import_id]))
        self.assertEqual(batched.status_code, 200)
        self.assertEqual(batched["X-Accounting-Zip-Batched"], "true")
        with ZipFile(BytesIO(batched.content)) as archive:
            part_names = archive.namelist()
            self.assertGreaterEqual(len(part_names), 2)
            self.assertTrue(all(name.endswith(".zip") for name in part_names))
            first_part = archive.read(part_names[0])
        with ZipFile(BytesIO(first_part)) as part_archive:
            self.assertEqual(len(part_archive.namelist()), 1)

        selected = self.client.get(
            reverse("accounting-import-statements-zip", args=[import_id]),
            {"customer_ids": str(summary.id), "style": "professional"},
        )
        self.assertEqual(selected.status_code, 200)
        self.assertEqual(selected["Content-Type"], "application/zip")
        with ZipFile(BytesIO(selected.content)) as archive:
            names = archive.namelist()
        self.assertEqual(len(names), 1)
        self.assertTrue(any("MILLENNIUM" in name for name in names))

    def test_ageing_filters_and_ordering(self):
        self.upload_import()
        self.client.force_authenticate(self.accountant)
        list_url = reverse("accounting-import-customer-list")

        over_60 = self.client.get(list_url, {"ageing": "over_60"})
        self.assertEqual(over_60.status_code, 200)
        self.assertGreaterEqual(len(over_60.data), 1)
        self.assertTrue(all(item["max_days"] > 60 for item in over_60.data))

        over_90 = self.client.get(list_url, {"ageing": "over_90"})
        self.assertEqual(over_90.status_code, 200)
        self.assertTrue(all(item["max_days"] > 90 for item in over_90.data))

        has_60_90 = self.client.get(list_url, {"ageing": "has_60_90"})
        self.assertEqual(has_60_90.status_code, 200)
        self.assertTrue(all(float(item["bucket_60_90"]) > 0 for item in has_60_90.data))

        ordered = self.client.get(list_url, {"ordering": "-max_days"})
        self.assertEqual(ordered.status_code, 200)
        max_days = [item["max_days"] for item in ordered.data]
        self.assertEqual(max_days, sorted(max_days, reverse=True))
