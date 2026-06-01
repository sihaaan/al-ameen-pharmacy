from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .services import statement_filename, statement_ledger


PRIMARY = "0F766E"
PRIMARY_DARK = "115E59"
ACCENT = "D6A84F"
LIGHT = "E6FFFA"
SOFT = "F9FAFB"
SOFT_GREEN = "F0FDFA"
BORDER = "D1D5DB"
TEXT = "111827"
MUTED = "6B7280"
WHITE = "FFFFFF"


def money(value):
    return float(value or 0)


def period_text(ledger):
    start = ledger.get("period_start")
    end = ledger.get("period_end")
    if start and end:
        if start == end:
            return start.isoformat()
        return f"{start.isoformat()} to {end.isoformat()}"
    if start:
        return start.isoformat()
    if end:
        return end.isoformat()
    return "No invoice rows"


def statement_excel_filename(import_customer):
    return statement_filename(import_customer).replace(".pdf", ".xlsx")


def apply_common_styles(sheet):
    thin = Side(style="thin", color=BORDER)
    subtle = Side(style="hair", color="E5E7EB")
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=subtle, right=subtle, top=subtle, bottom=subtle)
    return thin


def build_statement_workbook(import_customer, *, date_from=None, date_to=None):
    ledger = statement_ledger(import_customer, date_from=date_from, date_to=date_to)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statement"

    # Header mirrors the customer-facing PDF hierarchy while staying Excel-friendly.
    sheet.merge_cells("A1:D2")
    sheet["A1"] = "AL AMEEN PHARMACY LLC"
    sheet["A1"].font = Font(name="Calibri", size=18, bold=True, color=PRIMARY_DARK)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor=SOFT_GREEN)

    sheet.merge_cells("A3:D3")
    sheet["A3"] = "Dubai, United Arab Emirates | Phone: +971 50 545 6388 | Email: pharmacydxb@gmail.com"
    sheet["A3"].font = Font(name="Calibri", size=9, color=MUTED)
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet["A3"].fill = PatternFill("solid", fgColor=SOFT_GREEN)

    sheet.merge_cells("E1:G1")
    sheet["E1"] = "STATEMENT OF ACCOUNT"
    sheet["E1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    sheet["E1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["E1"].fill = PatternFill("solid", fgColor=PRIMARY)
    sheet.merge_cells("E2:G2")
    sheet["E2"] = "Overdue Payment Statement"
    sheet["E2"].font = Font(name="Calibri", size=10, bold=True, color=PRIMARY_DARK)
    sheet["E2"].alignment = Alignment(horizontal="center")
    sheet["E2"].fill = PatternFill("solid", fgColor=LIGHT)
    sheet.merge_cells("E3:G3")
    sheet["E3"] = f"Statement Date: {import_customer.accounting_import.report_date or '-'}"
    sheet["E3"].font = Font(name="Calibri", size=9, color=TEXT)
    sheet["E3"].alignment = Alignment(horizontal="center")
    sheet["E3"].fill = PatternFill("solid", fgColor=LIGHT)

    info_rows = [
        ("Customer", import_customer.customer_name, "Currency", "AED"),
        ("Account No.", import_customer.customer_code or "-", "Statement Period", period_text(ledger)),
        ("Final Balance", money(ledger["final_balance"]), "Invoices", ledger["invoice_count"]),
    ]
    start_row = 5
    for index, row in enumerate(info_rows, start=start_row):
        sheet.cell(index, 1, row[0])
        sheet.cell(index, 2, row[1])
        sheet.cell(index, 5, row[2])
        sheet.cell(index, 6, row[3])

    header_row = 10
    headers = ["Invoice Date", "Doc Type", "Invoice No.", "LPO / Reference No.", "Debit", "Credit", "Balance"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col, header)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = header_row + 1
    for line in ledger["lines"]:
        invoice = line["row"]
        values = [
            invoice.invoice_date,
            line["doc_type"],
            invoice.invoice_number or invoice.bill_number,
            invoice.lpo_reference or "-",
            money(line["debit"]),
            money(line["credit"]),
            money(line["balance"]),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(current_row, col, value)
            if col in {5, 6, 7}:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == 1:
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
        current_row += 1

    if not ledger["lines"]:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        sheet.cell(current_row, 1, "No invoice rows found for this statement period.")
        current_row += 1

    table_end_row = max(header_row + 1, current_row - 1)
    if ledger["lines"]:
        table = Table(displayName=f"StatementRows{import_customer.id}", ref=f"A{header_row}:G{table_end_row}")
        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)

    totals_start = current_row + 2
    totals = [
        ("Total Debit", ledger["total_debit"]),
        ("Total Credit", ledger["total_credit"]),
        ("Net Value / Total Outstanding", ledger["net_value"]),
        ("Final Balance", ledger["final_balance"]),
    ]
    for index, (label, value) in enumerate(totals, start=totals_start):
        sheet.merge_cells(start_row=index, start_column=4, end_row=index, end_column=5)
        sheet.merge_cells(start_row=index, start_column=6, end_row=index, end_column=7)
        sheet.cell(index, 4, label)
        sheet.cell(index, 6, money(value))
        sheet.cell(index, 6).number_format = '#,##0.00'
        sheet.cell(index, 4).font = Font(name="Calibri", size=10, bold=True, color=TEXT)
        sheet.cell(index, 6).font = Font(name="Calibri", size=10, bold=True, color=TEXT)
        sheet.cell(index, 4).alignment = Alignment(horizontal="left")
        sheet.cell(index, 6).alignment = Alignment(horizontal="right")

    note_row = totals_start + len(totals) + 2
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=7)
    sheet.cell(
        note_row,
        1,
        "Please verify this statement and clear the outstanding amount at the earliest. "
        "If payment has recently been made, please accept our thanks and ignore this reminder.",
    )
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(note_row, 1).font = Font(name="Calibri", size=10, italic=True, color=PRIMARY_DARK)
    sheet.cell(note_row, 1).fill = PatternFill("solid", fgColor=SOFT_GREEN)

    footer_row = note_row + 4
    sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=7)
    sheet.cell(footer_row, 1, "Al Ameen Pharmacy LLC | This statement is computer generated for account reconciliation.")
    sheet.cell(footer_row, 1).font = Font(name="Calibri", size=9, color=MUTED)
    sheet.cell(footer_row, 1).alignment = Alignment(horizontal="center")

    thin = apply_common_styles(sheet)
    for col in range(1, 8):
        sheet.column_dimensions[get_column_letter(col)].width = [16, 14, 16, 34, 15, 15, 17][col - 1]
    for row_number in range(1, footer_row + 1):
        sheet.row_dimensions[row_number].height = 22
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[note_row].height = 42
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:G{max(header_row, current_row - 1)}"

    for row_number in range(start_row, start_row + len(info_rows)):
        sheet.cell(row_number, 1).fill = PatternFill("solid", fgColor=SOFT)
        sheet.cell(row_number, 5).fill = PatternFill("solid", fgColor=SOFT)
        sheet.cell(row_number, 1).font = Font(name="Calibri", size=10, bold=True, color=TEXT)
        sheet.cell(row_number, 5).font = Font(name="Calibri", size=10, bold=True, color=TEXT)
        if row_number == start_row + 2:
            sheet.cell(row_number, 2).number_format = '#,##0.00'
    for row_number in range(totals_start, totals_start + len(totals)):
        sheet.cell(row_number, 4).fill = PatternFill("solid", fgColor=LIGHT)
        sheet.cell(row_number, 6).fill = PatternFill("solid", fgColor=LIGHT)
        sheet.cell(row_number, 4).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet.cell(row_number, 6).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.cell(totals_start + len(totals) - 1, 4).fill = PatternFill("solid", fgColor=PRIMARY)
    sheet.cell(totals_start + len(totals) - 1, 6).fill = PatternFill("solid", fgColor=PRIMARY)
    sheet.cell(totals_start + len(totals) - 1, 4).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    sheet.cell(totals_start + len(totals) - 1, 6).font = Font(name="Calibri", size=10, bold=True, color=WHITE)

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = f"${header_row}:${header_row}"
    sheet.print_area = f"A1:G{footer_row}"
    sheet.oddFooter.center.text = "Page &P of &N"
    sheet.oddFooter.right.text = "Statement of Account"
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
