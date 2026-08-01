from io import BytesIO
from decimal import Decimal
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase, override_settings
from openpyxl import Workbook
from reportlab.pdfgen import canvas

from .historical_import_parsers import parse_historical_pdf_upload
from .price_reference import (
    _reference_rows_from_preview,
    parse_price_reference_source,
    parse_price_reference_workbook,
)
from .models import QuotationSettings


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def workbook_upload(workbook, name="reference.xlsx"):
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(name, buffer.getvalue(), content_type=XLSX_MIME)


def reference_sheet(workbook, title="Reference"):
    sheet = workbook.create_sheet(title)
    sheet.append(["Sl No", "Items", "unit", "qty", "uprice"])
    return sheet


class PriceReferenceAttachmentFidelityTests(SimpleTestCase):
    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=1,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=10,
        QUOTATION_PRICE_REFERENCE_MAX_EXCEL_ROWS=10,
    )
    def test_visible_sheet_is_selected_before_cap_and_inspection_meta_is_returned(self):
        workbook = Workbook()
        hidden = workbook.active
        hidden.title = "Hidden archive"
        hidden.sheet_state = "hidden"
        hidden.append(["Sl No", "Items", "unit", "qty", "uprice"])
        hidden.append([1, "Old archived item", "PCS", 9, 99])
        for index in range(2, 11):
            archived = workbook.create_sheet(f"Hidden archive {index}")
            archived.sheet_state = "hidden"
        current = reference_sheet(workbook, "Current")
        current.append([1, "Current bandages", "PCS", 2, 12.5])
        current["G2"] = "=1+1"

        rows, meta = parse_price_reference_workbook(workbook_upload(workbook))

        self.assertEqual([row.item_name for row in rows], ["Current bandages"])
        self.assertEqual(rows[0].unit_price, Decimal("12.50"))
        self.assertEqual(meta["attachment_safety"]["validated_format"], "xlsx")
        self.assertEqual(meta["spreadsheet_fidelity"]["hidden_sheet_count"], 10)
        self.assertEqual(meta["spreadsheet_fidelity"]["formula_cell_count"], 1)
        self.assertEqual(
            [sheet["sheet_name"] for sheet in meta["sheet_metadata"]],
            ["Current"],
        )
        self.assertFalse(meta["spreadsheet_limits"]["sheet_limit_reached"])

    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=10,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=5,
        QUOTATION_PRICE_REFERENCE_MAX_EXCEL_ROWS=2,
    )
    def test_row_and_column_limits_warn_only_when_source_exceeds_them(self):
        exact_workbook = Workbook()
        exact_workbook.remove(exact_workbook.active)
        exact = reference_sheet(exact_workbook)
        exact.append([1, "Exact-limit item", "PCS", 1, 10])

        exact_rows, exact_meta = parse_price_reference_workbook(
            workbook_upload(exact_workbook, "exact.xlsx")
        )

        self.assertEqual(len(exact_rows), 1)
        self.assertFalse(exact_meta["sheet_metadata"][0]["row_limit_reached"])
        self.assertFalse(
            any("Stopped reading price reference sheet" in warning for warning in exact_meta["warnings"])
        )

        truncated_workbook = Workbook()
        truncated_workbook.remove(truncated_workbook.active)
        truncated = reference_sheet(truncated_workbook)
        truncated.append([1, "Included item", "PCS", 1, 10, "ignored column"])
        truncated.append([2, "Truncated item", "PCS", 2, 20, "ignored column"])

        truncated_rows, truncated_meta = parse_price_reference_workbook(
            workbook_upload(truncated_workbook, "truncated.xlsx")
        )

        self.assertEqual([row.item_name for row in truncated_rows], ["Included item"])
        self.assertTrue(truncated_meta["sheet_metadata"][0]["row_limit_reached"])
        self.assertTrue(truncated_meta["sheet_metadata"][0]["column_limit_reached"])
        self.assertTrue(
            any("contains additional rows" in warning for warning in truncated_meta["warnings"])
        )
        self.assertTrue(
            any("after 5 columns" in warning for warning in truncated_meta["warnings"])
        )
        self.assertEqual(len(truncated_rows[0].raw_values), 5)

    def test_preview_based_reference_keeps_attachment_fidelity_metadata(self):
        rows, meta = _reference_rows_from_preview(
            {
                "source_type": "pdf",
                "source_filename": "reference.pdf",
                "lines": [
                    {
                        "raw_name": "Bandages",
                        "quantity": "2",
                        "unit": "PCS",
                        "unit_price": "12.50",
                    }
                ],
                "meta": {
                    "attachment_safety": {"validated_format": "pdf"},
                    "pdf_fidelity": {"form_field_markers": True},
                },
            }
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(meta["attachment_safety"]["validated_format"], "pdf")
        self.assertTrue(meta["pdf_fidelity"]["form_field_markers"])


class PriceReferenceAIWarningRetentionTests(TestCase):
    class Provider:
        def clean_rows(self, **_kwargs):
            return {
                "rows": [
                    {
                        "item_name": "Bandages",
                        "quantity": "2",
                        "unit": "PCS",
                        "unit_price": "12.50",
                        "vat_rate": "",
                        "vat_amount": "",
                        "line_total": "25.00",
                        "pack_info": "",
                        "notes": "",
                        "raw_source_text": "Bandages | 2 | PCS | 12.50",
                        "page_number": "1",
                        "confidence": 0.95,
                        "parse_status": "parsed",
                        "reason": "",
                    }
                ],
                "warnings": ["AI warning remains visible."],
                "document_notes": "",
            }, {"input_tokens": 10, "output_tokens": 10}

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    @patch("quotations.price_reference.parse_file_preview")
    def test_pdf_ai_cleanup_retains_deterministic_attachment_warnings(self, parse_file):
        QuotationSettings.objects.update_or_create(
            pk=1,
            defaults={"ai_parsing_enabled": True},
        )
        parse_file.return_value = {
            "source_type": "pdf",
            "source_filename": "reference.pdf",
            "source_mime_type": "application/pdf",
            "source_sha256": "d" * 64,
            "source_file_ref": "",
            "source_file_size": 123,
            "parse_method": "pymupdf_text_v2",
            "original_text": "Bandages | 2 | PCS | 12.50",
            "lines": [
                {
                    "raw_name": "Bandages",
                    "quantity": "2",
                    "unit": "PCS",
                    "unit_price": "12.50",
                    "raw_line": "Bandages | 2 | PCS | 12.50",
                }
            ],
            "warnings": [
                "PDF contains form or XFA fields; verify the document manually."
            ],
            "meta": {
                "attachment_safety": {"validated_format": "pdf"},
                "pdf_fidelity": {"form_field_markers": True},
            },
        }
        upload = SimpleUploadedFile(
            "reference.pdf",
            b"parser-is-mocked",
            content_type="application/pdf",
        )

        with patch(
            "quotations.ai_parsing.get_ai_parse_provider",
            return_value=self.Provider(),
        ):
            rows, meta = parse_price_reference_source(upload, use_ai=True)

        self.assertEqual(len(rows), 1)
        self.assertTrue(rows[0].item_name.startswith("Bandages"))
        self.assertIn(
            "PDF contains form or XFA fields; verify the document manually.",
            meta["warnings"],
        )
        self.assertIn("AI warning remains visible.", meta["warnings"])
        self.assertTrue(meta["pdf_fidelity"]["form_field_markers"])


class HistoricalPdfAttachmentFidelityTests(SimpleTestCase):
    def make_pdf_upload(self, *, content_type="application/pdf", markers=b""):
        buffer = BytesIO()
        document = canvas.Canvas(buffer)
        document.drawString(72, 760, "Historical quotation reference")
        document.showPage()
        document.save()
        return SimpleUploadedFile(
            "historical.pdf",
            buffer.getvalue() + markers,
            content_type=content_type,
        )

    @patch(
        "quotations.historical_import_parsers.store_import_source",
        return_value="historical_sources/test.pdf",
    )
    def test_shared_pdf_warnings_and_metadata_are_added_without_changing_extraction(self, _store):
        upload = self.make_pdf_upload(
            content_type="text/plain",
            markers=b"\n% /JavaScript /EmbeddedFile /AcroForm\n",
        )

        preview = parse_historical_pdf_upload(upload)

        self.assertEqual(preview["lines"], [])
        self.assertTrue(
            any("declared content type" in warning for warning in preview["warnings"])
        )
        self.assertTrue(
            any("active-content markers" in warning for warning in preview["warnings"])
        )
        self.assertTrue(
            any("embedded-file markers" in warning for warning in preview["warnings"])
        )
        self.assertTrue(any("form or XFA fields" in warning for warning in preview["warnings"]))
        self.assertTrue(preview["meta"]["attachment_safety"]["active_content_markers"])
        self.assertTrue(preview["meta"]["attachment_safety"]["embedded_file_markers"])
        self.assertTrue(preview["meta"]["pdf_fidelity"]["form_field_markers"])

    @patch("quotations.historical_import_parsers._extract_pdf_text")
    def test_malformed_pdf_is_rejected_by_shared_inspection_before_extraction(self, extract_text):
        upload = SimpleUploadedFile(
            "historical.pdf",
            b"%PDF-1.4\nnot-a-complete-document",
            content_type="application/pdf",
        )

        with self.assertRaises(ValidationError):
            parse_historical_pdf_upload(upload)

        extract_text.assert_not_called()
