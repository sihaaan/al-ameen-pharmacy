import base64
import hashlib
import html
import json
import os
import re
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from django.core import signing
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.signing import BadSignature, SignatureExpired, TimestampSigner
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_parsing import AIParseError, get_ai_parse_provider, settings_ai_status
from .import_parsers import parse_file_preview
from .import_rules import parse_inquiry_line
from .matching import suggest_product_for_text
from .models import (
    Company,
    ContractIntelligenceItem,
    ContractIntelligenceRun,
    ContractIntelligenceSource,
    GmailOAuthConnection,
    QuotationAuditLog,
    QuotationSettings,
    normalize_label,
)
from .services import audit_log

try:
    from cryptography.fernet import Fernet, InvalidToken
except ImportError:  # pragma: no cover - production installs this transitively; dev can still run checks.
    Fernet = None
    InvalidToken = Exception


GMAIL_READONLY_SCOPE = "https://www.googleapis.com/auth/gmail.readonly"
GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GMAIL_API_BASE = "https://gmail.googleapis.com/gmail/v1/users/me"

MAX_ANALYSIS_CHARS = 18000
MAX_ATTACHMENT_BYTES = 5 * 1024 * 1024
SUPPORTED_ATTACHMENT_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".xlsb"}
DEFAULT_DISCOVERY_BATCH_SIZE = 25
MAX_DISCOVERY_BATCH_SIZE = 100
MAX_CONTRACT_MESSAGES = 5000
DEFAULT_AI_ANALYSIS_BATCH_SIZE = 5


CONTRACT_AI_JSON_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "classification": {
            "type": "string",
            "enum": [
                ContractIntelligenceSource.CLASS_INQUIRY,
                ContractIntelligenceSource.CLASS_QUOTATION,
                ContractIntelligenceSource.CLASS_LPO,
                ContractIntelligenceSource.CLASS_FOLLOWUP,
                ContractIntelligenceSource.CLASS_IRRELEVANT,
                ContractIntelligenceSource.CLASS_UNKNOWN,
            ],
        },
        "confidence": {"type": "number"},
        "items": {
            "type": "array",
            "maxItems": 120,
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "item_name": {"type": "string"},
                    "suggested_item_name": {"type": "string"},
                    "quantity": {"type": "string"},
                    "unit": {"type": "string"},
                    "unit_price": {"type": "string"},
                    "currency": {"type": "string"},
                    "project": {"type": "string"},
                    "contact_text": {"type": "string"},
                    "source_text": {"type": "string"},
                    "source_filename": {"type": "string"},
                    "source_page": {"type": "string"},
                    "confidence": {"type": "number"},
                    "reason": {"type": "string"},
                },
                "required": [
                    "item_name",
                    "suggested_item_name",
                    "quantity",
                    "unit",
                    "unit_price",
                    "currency",
                    "project",
                    "contact_text",
                    "source_text",
                    "source_filename",
                    "source_page",
                    "confidence",
                    "reason",
                ],
            },
        },
        "warnings": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["classification", "confidence", "items", "warnings"],
}


def gmail_oauth_configured():
    return bool(
        getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "")
        and getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "")
    )


def gmail_oauth_redirect_uri(request=None):
    configured = getattr(settings, "GOOGLE_OAUTH_REDIRECT_URI", "")
    if configured:
        return configured
    if request is not None:
        return request.build_absolute_uri("/api/quotations/gmail/oauth/callback/")
    return ""


def gmail_frontend_redirect_url(status="connected"):
    base = getattr(settings, "FRONTEND_URL", "") or getattr(settings, "SITE_URL", "")
    if base:
        return f"{base.rstrip('/')}/admin?quotation_tab=contract-intelligence&gmail={urllib.parse.quote(status)}"
    return f"/admin?quotation_tab=contract-intelligence&gmail={urllib.parse.quote(status)}"


def _fernet():
    if Fernet is None:
        return None
    digest = hashlib.sha256(str(settings.SECRET_KEY).encode("utf-8")).digest()
    return Fernet(base64.urlsafe_b64encode(digest))


def encrypt_token(value):
    if not value:
        return ""
    fernet = _fernet()
    if fernet is None:
        return signing.dumps({"token": str(value)}, salt="quotation-gmail-token")
    return fernet.encrypt(str(value).encode("utf-8")).decode("utf-8")


def decrypt_token(value):
    if not value:
        return ""
    fernet = _fernet()
    if fernet is None:
        try:
            payload = signing.loads(str(value), salt="quotation-gmail-token")
            return payload.get("token", "")
        except BadSignature:
            return ""
    try:
        return fernet.decrypt(str(value).encode("utf-8")).decode("utf-8")
    except InvalidToken:
        return ""


def build_gmail_auth_url(user, request=None):
    if not gmail_oauth_configured():
        raise ValueError("Gmail OAuth is not configured. Add GOOGLE_OAUTH_CLIENT_ID and GOOGLE_OAUTH_CLIENT_SECRET.")
    signer = TimestampSigner(salt="quotation-gmail-oauth")
    state = signer.sign(str(user.id))
    params = {
        "client_id": settings.GOOGLE_OAUTH_CLIENT_ID,
        "redirect_uri": gmail_oauth_redirect_uri(request),
        "response_type": "code",
        "scope": GMAIL_READONLY_SCOPE,
        "access_type": "offline",
        "prompt": "consent",
        "include_granted_scopes": "true",
        "state": state,
    }
    return f"{GOOGLE_AUTH_URL}?{urllib.parse.urlencode(params)}"


def parse_gmail_oauth_state(state, max_age=600):
    signer = TimestampSigner(salt="quotation-gmail-oauth")
    try:
        return int(signer.unsign(state, max_age=max_age))
    except (BadSignature, SignatureExpired, TypeError, ValueError):
        return None


def _json_request(url, *, method="GET", data=None, token=None, timeout=60):
    encoded = None
    headers = {"Accept": "application/json"}
    if data is not None:
        encoded = json.dumps(data).encode("utf-8")
        headers["Content-Type"] = "application/json"
    if token:
        headers["Authorization"] = f"Bearer {token}"
    request = urllib.request.Request(url, data=encoded, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google API request failed with HTTP {exc.code}: {detail[:400]}") from exc


def _form_request(url, data, timeout=60):
    encoded = urllib.parse.urlencode(data).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=encoded,
        headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google OAuth request failed with HTTP {exc.code}: {detail[:400]}") from exc


def exchange_gmail_code(user, code, request=None):
    token_payload = _form_request(
        GOOGLE_TOKEN_URL,
        {
            "code": code,
            "client_id": settings.GOOGLE_OAUTH_CLIENT_ID,
            "client_secret": settings.GOOGLE_OAUTH_CLIENT_SECRET,
            "redirect_uri": gmail_oauth_redirect_uri(request),
            "grant_type": "authorization_code",
        },
    )
    access_token = token_payload.get("access_token", "")
    if not access_token:
        raise RuntimeError("Google OAuth did not return an access token.")
    gmail_profile = _json_request(f"{GMAIL_API_BASE}/profile", token=access_token)
    expires_in = int(token_payload.get("expires_in") or 3600)
    connection, _ = GmailOAuthConnection.objects.get_or_create(user=user)
    existing_refresh = decrypt_token(connection.refresh_token_encrypted)
    refresh_token = token_payload.get("refresh_token") or existing_refresh
    connection.email = gmail_profile.get("emailAddress", "") or connection.email
    connection.access_token_encrypted = encrypt_token(access_token)
    connection.refresh_token_encrypted = encrypt_token(refresh_token)
    connection.token_expiry = timezone.now() + timedelta(seconds=max(expires_in - 60, 60))
    connection.scopes = token_payload.get("scope", GMAIL_READONLY_SCOPE).split()
    connection.status = GmailOAuthConnection.STATUS_CONNECTED
    connection.last_error = ""
    connection.connected_at = timezone.now()
    connection.disconnected_at = None
    connection.save()
    return connection


def disconnect_gmail(connection):
    connection.access_token_encrypted = ""
    connection.refresh_token_encrypted = ""
    connection.status = GmailOAuthConnection.STATUS_DISCONNECTED
    connection.disconnected_at = timezone.now()
    connection.save(update_fields=["access_token_encrypted", "refresh_token_encrypted", "status", "disconnected_at", "updated_at"])
    return connection


def get_valid_access_token(connection):
    if connection.status != GmailOAuthConnection.STATUS_CONNECTED:
        raise RuntimeError("Gmail is not connected.")
    access_token = decrypt_token(connection.access_token_encrypted)
    if access_token and connection.token_expiry and connection.token_expiry > timezone.now() + timedelta(seconds=30):
        return access_token
    refresh_token = decrypt_token(connection.refresh_token_encrypted)
    if not refresh_token:
        raise RuntimeError("Gmail refresh token is missing. Reconnect Gmail.")
    payload = _form_request(
        GOOGLE_TOKEN_URL,
        {
            "client_id": settings.GOOGLE_OAUTH_CLIENT_ID,
            "client_secret": settings.GOOGLE_OAUTH_CLIENT_SECRET,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        },
    )
    access_token = payload.get("access_token", "")
    if not access_token:
        raise RuntimeError("Could not refresh Gmail access token.")
    expires_in = int(payload.get("expires_in") or 3600)
    connection.access_token_encrypted = encrypt_token(access_token)
    connection.token_expiry = timezone.now() + timedelta(seconds=max(expires_in - 60, 60))
    connection.last_error = ""
    connection.save(update_fields=["access_token_encrypted", "token_expiry", "last_error", "updated_at"])
    return access_token


def build_contract_gmail_query(run):
    raw_query = (run.gmail_query or "").strip()
    if raw_query:
        return raw_query
    company_terms = [run.target_company_name or "", getattr(run.company, "name", "")]
    company_term = next((term for term in company_terms if term), "ALEC")
    domain_hint = (getattr(run, "sender_domain_hint", "") or "").strip().lstrip("@")
    if domain_hint:
        # For contract intelligence, domain hints should widen discovery rather
        # than over-filter it. AI/basic analysis will classify relevance later.
        company_search_parts = [
            f"from:{domain_hint}",
            f"to:{domain_hint}",
            f"cc:{domain_hint}",
            f'"{domain_hint}"',
            f'"{company_term}"',
        ]
        parts = ["(" + " OR ".join(company_search_parts) + ")"]
    else:
        parts = [
            f'"{company_term}"',
            "(inquiry OR enquiry OR RFQ OR quotation OR quote OR LPO OR purchase order)",
        ]
    if run.date_from:
        parts.append(f"after:{run.date_from.strftime('%Y/%m/%d')}")
    if run.date_to:
        parts.append(f"before:{(run.date_to + timedelta(days=1)).strftime('%Y/%m/%d')}")
    return " ".join(parts)


def gmail_search_messages(connection, query, max_messages=100, page_token=""):
    token = get_valid_access_token(connection)
    max_messages = min(max(int(max_messages or DEFAULT_DISCOVERY_BATCH_SIZE), 1), MAX_DISCOVERY_BATCH_SIZE)
    params = {
        "q": query,
        "maxResults": max_messages,
    }
    if page_token:
        params["pageToken"] = page_token
    url = f"{GMAIL_API_BASE}/messages?{urllib.parse.urlencode(params)}"
    payload = _json_request(url, token=token)
    return {
        "messages": payload.get("messages") or [],
        "next_page_token": payload.get("nextPageToken") or "",
        "result_size_estimate": payload.get("resultSizeEstimate"),
    }


def _decode_gmail_data(data):
    if not data:
        return b""
    padded = data + ("=" * (-len(data) % 4))
    return base64.urlsafe_b64decode(padded.encode("utf-8"))


def _strip_html(value):
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", value or "")
    text = re.sub(r"(?is)<br\s*/?>", "\n", text)
    text = re.sub(r"(?is)</p\s*>", "\n", text)
    text = re.sub(r"(?is)<.*?>", " ", text)
    return html.unescape(re.sub(r"[ \t]+", " ", text)).strip()


def _walk_parts(payload):
    yield payload
    for part in payload.get("parts") or []:
        yield from _walk_parts(part)


def _header(headers, name):
    name = name.lower()
    for header in headers or []:
        if str(header.get("name", "")).lower() == name:
            return header.get("value", "")
    return ""


def _message_datetime(message):
    internal = message.get("internalDate")
    if internal:
        try:
            return datetime.fromtimestamp(int(internal) / 1000, tz=timezone.get_current_timezone())
        except (TypeError, ValueError, OSError):
            pass
    return None


def gmail_fetch_message(connection, message_id, *, include_attachments=True):
    token = get_valid_access_token(connection)
    payload = _json_request(
        f"{GMAIL_API_BASE}/messages/{urllib.parse.quote(message_id)}?format=full",
        token=token,
    )
    headers = payload.get("payload", {}).get("headers") or []
    text_parts = []
    attachment_refs = []
    for part in _walk_parts(payload.get("payload") or {}):
        mime = part.get("mimeType") or ""
        body = part.get("body") or {}
        filename = part.get("filename") or ""
        data = body.get("data")
        if data and mime == "text/plain":
            text_parts.append(_decode_gmail_data(data).decode("utf-8", errors="replace"))
        elif data and mime == "text/html":
            text_parts.append(_strip_html(_decode_gmail_data(data).decode("utf-8", errors="replace")))
        if filename and body.get("attachmentId"):
            attachment_refs.append(
                {
                    "filename": filename,
                    "mime_type": mime,
                    "size": body.get("size") or 0,
                    "attachment_id": body.get("attachmentId"),
                }
            )

    parsed_attachments = []
    if include_attachments:
        for attachment in attachment_refs[:10]:
            extension = os.path.splitext(attachment["filename"])[1].lower()
            if extension not in SUPPORTED_ATTACHMENT_EXTENSIONS:
                parsed_attachments.append({**attachment, "status": "skipped", "reason": "Unsupported attachment type for V1."})
                continue
            if int(attachment.get("size") or 0) > MAX_ATTACHMENT_BYTES:
                parsed_attachments.append({**attachment, "status": "skipped", "reason": "Attachment too large for synchronous V1 parsing."})
                continue
            try:
                attachment_payload = _json_request(
                    f"{GMAIL_API_BASE}/messages/{urllib.parse.quote(message_id)}/attachments/{urllib.parse.quote(attachment['attachment_id'])}",
                    token=token,
                )
                content = _decode_gmail_data(attachment_payload.get("data", ""))
                upload = SimpleUploadedFile(attachment["filename"], content, content_type=attachment.get("mime_type") or "application/octet-stream")
                preview = parse_file_preview(upload)
                parsed_attachments.append(
                    {
                        **attachment,
                        "status": "parsed",
                        "source_file_ref": preview.get("source_file_ref", ""),
                        "line_count": len(preview.get("lines") or []),
                        "lines": preview.get("lines") or [],
                        "warnings": preview.get("warnings") or [],
                    }
                )
            except Exception as exc:
                parsed_attachments.append({**attachment, "status": "failed", "reason": str(exc)[:250]})

    body_text = "\n".join(part for part in text_parts if part).strip()
    return {
        "gmail_message_id": payload.get("id", message_id),
        "gmail_thread_id": payload.get("threadId", ""),
        "subject": _header(headers, "Subject"),
        "sender": _header(headers, "From"),
        "recipients": ", ".join(filter(None, [_header(headers, "To"), _header(headers, "Cc")])),
        "sent_at": _message_datetime(payload),
        "snippet": payload.get("snippet", ""),
        "body_text": body_text,
        "attachments": parsed_attachments or attachment_refs,
    }


def gmail_fetch_message_metadata(connection, message_id):
    token = get_valid_access_token(connection)
    params = [
        ("format", "metadata"),
        ("metadataHeaders", "Subject"),
        ("metadataHeaders", "From"),
        ("metadataHeaders", "To"),
        ("metadataHeaders", "Cc"),
    ]
    payload = _json_request(
        f"{GMAIL_API_BASE}/messages/{urllib.parse.quote(message_id)}?{urllib.parse.urlencode(params)}",
        token=token,
    )
    headers = payload.get("payload", {}).get("headers") or []
    return {
        "gmail_message_id": payload.get("id", message_id),
        "gmail_thread_id": payload.get("threadId", ""),
        "subject": _header(headers, "Subject"),
        "sender": _header(headers, "From"),
        "recipients": ", ".join(filter(None, [_header(headers, "To"), _header(headers, "Cc")])),
        "sent_at": _message_datetime(payload),
        "snippet": payload.get("snippet", ""),
    }


def hydrate_contract_source(source, connection, *, include_attachments=True):
    if source.body_text or any((attachment or {}).get("status") == "parsed" for attachment in source.attachments or []):
        return source
    payload = gmail_fetch_message(connection, source.gmail_message_id, include_attachments=include_attachments)
    source.subject = payload.get("subject", "")[:500] or source.subject
    source.sender = payload.get("sender", "")[:500] or source.sender
    source.recipients = payload.get("recipients", "") or source.recipients
    source.sent_at = payload.get("sent_at") or source.sent_at
    source.snippet = payload.get("snippet", "") or source.snippet
    source.body_text = payload.get("body_text", "")
    source.attachments = payload.get("attachments") or []
    source.source_sha256 = hashlib.sha256(
        "\n".join([source.subject or "", source.sender or "", source.body_text or ""]).encode("utf-8")
    ).hexdigest()
    source.status = "fetched"
    source.error = ""
    source.save(
        update_fields=[
            "subject",
            "sender",
            "recipients",
            "sent_at",
            "snippet",
            "body_text",
            "attachments",
            "source_sha256",
            "status",
            "error",
            "updated_at",
        ]
    )
    return source


def discover_contract_sources(run, user, *, batch_size=None, reset_cursor=False):
    try:
        connection = user.quotation_gmail_connection
    except GmailOAuthConnection.DoesNotExist:
        connection = None
    if not connection or connection.status != GmailOAuthConnection.STATUS_CONNECTED:
        raise RuntimeError("Connect Gmail read-only before running discovery.")

    run.status = ContractIntelligenceRun.STATUS_DISCOVERING
    run.started_at = timezone.now()
    run.warnings = []
    if reset_cursor:
        run.discovery_page_token = ""
        run.discovery_exhausted = False
    run.save(update_fields=["status", "started_at", "warnings", "discovery_page_token", "discovery_exhausted", "updated_at"])
    query = build_contract_gmail_query(run)
    existing_count = ContractIntelligenceSource.objects.filter(run=run).exclude(gmail_message_id="").count()
    remaining = max(min(int(run.max_messages or 0), MAX_CONTRACT_MESSAGES) - existing_count, 0)
    if remaining <= 0:
        run.status = ContractIntelligenceRun.STATUS_READY
        run.discovery_exhausted = True
        run.completed_at = timezone.now()
        refresh_contract_run_summary(run)
        run.save(update_fields=["status", "discovery_exhausted", "completed_at", "summary", "updated_at"])
        return {
            "query": query,
            "created": 0,
            "reused": 0,
            "failed": 0,
            "warnings": ["Run message limit reached. Increase max emails or create a narrower run if more discovery is needed."],
            "next_page_token": "",
            "discovery_exhausted": True,
            "result_size_estimate": run.discovery_result_estimate,
        }

    effective_batch_size = min(
        max(int(batch_size or run.discovery_batch_size or DEFAULT_DISCOVERY_BATCH_SIZE), 1),
        MAX_DISCOVERY_BATCH_SIZE,
        remaining,
    )
    search_result = gmail_search_messages(connection, query, effective_batch_size, page_token=run.discovery_page_token)
    messages = search_result["messages"]
    created = 0
    reused = 0
    failed = 0
    warnings = []
    for message in messages:
        message_id = message.get("id")
        if not message_id:
            continue
        if ContractIntelligenceSource.objects.filter(run=run, gmail_message_id=message_id).exists():
            reused += 1
            continue
        try:
            payload = gmail_fetch_message_metadata(connection, message_id)
            source_hash = hashlib.sha256(
                "\n".join([payload.get("subject", ""), payload.get("sender", ""), payload.get("snippet", "")]).encode("utf-8")
            ).hexdigest()
            ContractIntelligenceSource.objects.create(
                run=run,
                gmail_message_id=payload.get("gmail_message_id", ""),
                gmail_thread_id=payload.get("gmail_thread_id", ""),
                subject=payload.get("subject", "")[:500],
                sender=payload.get("sender", "")[:500],
                recipients=payload.get("recipients", ""),
                sent_at=payload.get("sent_at"),
                snippet=payload.get("snippet", ""),
                body_text="",
                source_sha256=source_hash,
                attachments=[],
                status="candidate",
            )
            created += 1
        except Exception as exc:
            failed += 1
            warnings.append(f"{message_id}: {str(exc)[:180]}")
    run.discovery_page_token = search_result.get("next_page_token") or ""
    run.discovery_exhausted = not bool(run.discovery_page_token) or (existing_count + created + reused) >= int(run.max_messages or 0)
    run.discovery_result_estimate = search_result.get("result_size_estimate") or run.discovery_result_estimate
    run.status = ContractIntelligenceRun.STATUS_READY if failed < len(messages or []) else ContractIntelligenceRun.STATUS_FAILED
    run.completed_at = timezone.now()
    run.warnings = warnings
    refresh_contract_run_summary(run)
    run.save(
        update_fields=[
            "status",
            "completed_at",
            "warnings",
            "summary",
            "discovery_page_token",
            "discovery_exhausted",
            "discovery_result_estimate",
            "updated_at",
        ]
    )
    audit_log(
        user,
        QuotationAuditLog.ACTION_IMPORTED,
        run,
        company=run.company,
        message="Discovered contract intelligence emails from Gmail.",
        changes={
            "query": query,
            "batch_size": effective_batch_size,
            "created": created,
            "reused": reused,
            "failed": failed,
            "discovery_exhausted": run.discovery_exhausted,
        },
    )
    return {
        "query": query,
        "created": created,
        "reused": reused,
        "failed": failed,
        "warnings": warnings,
        "next_page_token": run.discovery_page_token,
        "discovery_exhausted": run.discovery_exhausted,
        "result_size_estimate": run.discovery_result_estimate,
    }


def _decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip()).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _quantity(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip()).quantize(Decimal("0.001"))
    except (InvalidOperation, ValueError):
        return None


def _confidence(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        text = str(value).strip()
        if text.endswith("%"):
            return float(text[:-1].strip()) / 100
        parsed = float(text)
        return parsed / 100 if parsed > 1 else parsed
    except (TypeError, ValueError):
        return default


def _source_date(source):
    return source.sent_at.date() if source.sent_at else None


def _classify_source_deterministically(source):
    text = normalize_label(" ".join([source.subject, source.snippet, source.body_text[:1000]]))
    if any(token in text for token in ["purchase order", " lpo ", "local purchase order", " po "]):
        return ContractIntelligenceSource.CLASS_LPO, 0.75
    if any(token in text for token in ["quotation", "quote", "quoted"]):
        return ContractIntelligenceSource.CLASS_QUOTATION, 0.72
    if any(token in text for token in ["inquiry", "enquiry", "rfq", "request for quotation", "please quote", "kindly quote"]):
        return ContractIntelligenceSource.CLASS_INQUIRY, 0.80
    if any(token in text for token in ["follow up", "following up", "reminder"]):
        return ContractIntelligenceSource.CLASS_FOLLOWUP, 0.65
    return ContractIntelligenceSource.CLASS_UNKNOWN, 0.40


NOISE_PREFIXES = (
    "address",
    "dear ",
    "thanks",
    "thank you",
    "regards",
    "best regards",
    "kind regards",
    "please find",
    "kindly find",
    "attached",
    "subject",
    "from ",
    "to ",
    "cc ",
    "tel",
    "phone",
    "mobile",
    "email",
    "e-mail",
    "website",
    "project",
    "manuf",
    "manufacturer",
    "model",
    "remarks",
    "remark",
    "rfq",
    "r.f.q",
    "last date",
    "procurement",
    "chief executive",
    "prepared by",
    "approved by",
)


NOISE_PHRASES = (
    "alec logo",
    "al ameen invoice",
    "quotation is attached",
    "revised quotation is attached",
    "please find the attached",
    "kindly find the attached",
    "confidential",
    "terms and conditions",
    "follow us",
    "frij murar",
    "p.o. box",
    "po box",
    "against lpo",
    "need soa",
    "overdue payment",
    "invoice #",
    "last date of submission",
    "procurement specialist",
    "chief executive officer",
    "payment terms",
    "days credit",
    "external email",
    "vector file rgb",
    "human rights",
    "sustainability",
    "esg report",
    "publishes",
    "al ameen pharmacy group",
    "al ameen pharmacy llc",
    "authorized signature",
    "company stamp",
)


NOISE_LABELS = {
    "item",
    "items",
    "item descr",
    "item description",
    "description",
    "qty",
    "quantity",
    "uom",
    "unit",
    "unit price",
    "price",
    "total",
    "vat",
    "brand",
    "comments",
    "remark",
    "remarks",
    "rfq",
    "rfq ref",
    "rfq reference",
    "r f q",
    "project",
    "manuf",
    "manufacturer",
    "model",
    "last date of submission",
    "procurement specialist",
    "contact",
    "contact person",
    "mobile",
    "telephone",
    "phone",
    "email",
    "date",
    "subject",
    "from",
    "to",
    "cc",
    "s no",
    "sl no",
    "sno",
    "particular",
    "particulars",
}


def _has_low_text_signal(text):
    compact = re.sub(r"\s+", "", str(text or ""))
    if not compact:
        return True
    alpha_count = sum(char.isalpha() for char in compact)
    alnum_count = sum(char.isalnum() for char in compact)
    if alpha_count < 3:
        return True
    if re.fullmatch(r"[\W\d_]+", compact):
        return True
    leading = compact[:1]
    if leading in {"!", "$", "&", "#", "{", "[", "]", "|", "\\"} and not re.match(r"^\d{5,}[A-Za-z ]{3,}", compact):
        return True
    symbol_count = len(compact) - alnum_count
    if len(compact) <= 14 and symbol_count >= alpha_count:
        return True
    return False


def _clean_contract_item_name(value):
    text = html.unescape(str(value or ""))
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^[>\-*•\s]+", "", text).strip()
    text = text.strip("|:;- ")
    text = re.sub(r"^\d+\s*[.)|-]\s*", "", text).strip()
    text = re.sub(r"^\d{5,}\s+", "", text).strip()
    text = re.sub(r"^(?:item\s+descr(?:iption)?|description)\s*[:*-]\s*", "", text, flags=re.I).strip()

    comment_text = ""
    comment_match = re.search(r"\bcomments?\s*:\s*(.+)$", text, flags=re.I)
    if comment_match:
        comment_text = comment_match.group(1).strip()
        text = text[: comment_match.start()].strip()

    text = re.sub(
        r"\bbrand\s*:\s*(?:brand\s+as\s+quot(?:ed|e)?|as\s+quot(?:ed|e)?|brand\s+quoted|not\s+specified|n/?a|-)?\b",
        "",
        text,
        flags=re.I,
    )
    text = re.sub(r"\bbrand\s*:\s*$", "", text, flags=re.I).strip()
    text = re.sub(r"\s+", " ", text).strip(" |:;-")

    comment_text = re.sub(
        r"\bbrand\s*:\s*(?:brand\s+as\s+quot(?:ed|e)?|as\s+quot(?:ed|e)?|not\s+specified|n/?a|-)?\b",
        "",
        comment_text,
        flags=re.I,
    )
    comment_text = re.sub(r"\s+", " ", comment_text).strip(" |:;-")
    if comment_text and not _is_contract_item_noise_basic(comment_text):
        base_norm = normalize_label(text)
        comment_norm = normalize_label(comment_text)
        if comment_norm and comment_norm != base_norm and comment_norm not in base_norm:
            text = f"{text} - {comment_text}" if text else comment_text

    return text[:500]


def _is_contract_item_noise_basic(value):
    text = " ".join(str(value or "").split())
    if not text:
        return True
    lowered = text.lower().strip()
    normalized = normalize_label(text)
    if not normalized or len(normalized) < 3:
        return True
    if normalized in NOISE_LABELS:
        return True
    if _has_low_text_signal(text):
        return True
    if re.search(r"\bcid\s*:", lowered) or re.search(r"\[[^\]]*(?:cid|logo|external email)[^\]]*\]", lowered):
        return True
    if lowered.startswith(NOISE_PREFIXES):
        return True
    if any(phrase in lowered for phrase in NOISE_PHRASES):
        return True
    if re.search(r"\b(?:s\s*\.?\s*no|sl\s*\.?\s*no|partic(?:ulars?|ulr|lr)|prtc?l)\b", lowered):
        return True
    if re.search(r"https?://|www\.|@[\w.-]+", lowered):
        return True
    unwrapped_text = text.strip("<> ")
    if re.fullmatch(r"(?:[mt]\s*)?[+()0-9\s./-]{7,}(?:\s*\|\s*(?:[mt]\s*)?[+()0-9\s./-]{7,})*", unwrapped_text, flags=re.I):
        return True
    if re.fullmatch(r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}", lowered):
        return True
    if re.fullmatch(r"[a-z][a-z .'-]{2,40}", lowered) and any(
        title in lowered for title in ["suleman", "mahesh", "mohammad", "muhammad", "rengan", "rensan"]
    ):
        return True
    if re.search(r"\+?\d[\d\s()./-]{5,}", lowered) and any(
        title in lowered for title in ["suleman", "mahesh", "mohammad", "muhammad", "rengan", "rensan"]
    ):
        return True
    return False


def _is_contract_item_noise(value):
    text = _clean_contract_item_name(value)
    unwrapped_text = text.strip("<> ")
    normalized = normalize_label(text)
    if not normalized or len(normalized) < 3:
        return True
    lowered = text.lower().strip()
    if normalized in NOISE_LABELS:
        return True
    if _has_low_text_signal(text):
        return True
    if re.search(r"\bcid\s*:", lowered) or re.search(r"\[[^\]]*(?:cid|logo|external email)[^\]]*\]", lowered):
        return True
    if lowered.startswith(NOISE_PREFIXES):
        return True
    if any(phrase in lowered for phrase in NOISE_PHRASES):
        return True
    if re.search(r"\b(?:s\s*\.?\s*no|sl\s*\.?\s*no|partic(?:ulars?|ulr|lr)|prtc?l)\b", lowered):
        return True
    if re.search(r"https?://|www\.|@[\w.-]+", lowered):
        return True
    if re.fullmatch(r"(?:[mt]\s*)?[+()0-9\s./-]{7,}(?:\s*\|\s*(?:[mt]\s*)?[+()0-9\s./-]{7,})*", unwrapped_text, flags=re.I):
        return True
    if re.fullmatch(r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}", lowered):
        return True
    if re.fullmatch(r"[a-z][a-z .'-]{2,40}", lowered) and any(
        title in lowered for title in ["suleman", "mahesh", "mohammad", "muhammad", "rengan", "rensan"]
    ):
        return True
    if re.search(r"\+?\d[\d\s()./-]{5,}", lowered) and any(
        title in lowered for title in ["suleman", "mahesh", "mohammad", "muhammad", "rengan", "rensan"]
    ):
        return True
    # Long prose lines without any item-like quantity or price signal are almost
    # always email body/signature text, not product demand.
    if len(text.split()) >= 10 and not re.search(r"\b\d+(?:\.\d+)?\b", text):
        return True
    return False


def _deterministic_items_from_text(source):
    items = []
    lines = [line.strip() for line in re.split(r"[\r\n]+", source.body_text or "") if line.strip()]
    for line in lines[:300]:
        if _is_contract_item_noise(line):
            continue
        parsed = parse_inquiry_line(line, base_confidence=0.45)
        if not parsed:
            continue
        name = parsed.get("requested_item_name") or parsed.get("raw_name") or ""
        clean_name = _clean_contract_item_name(name)
        if _is_contract_item_noise(clean_name):
            continue
        items.append(
            {
                "item_name": clean_name,
                "suggested_item_name": clean_name,
                "quantity": parsed.get("quantity"),
                "unit": parsed.get("unit") or "",
                "unit_price": parsed.get("unit_price"),
                "currency": "AED",
                "project": "",
                "contact_text": source.sender,
                "source_text": line,
                "source_filename": "",
                "source_page": "",
                "confidence": float(parsed.get("parse_confidence") or 0.45),
                "reason": "Detected from email body text.",
            }
        )
    return items


def _deterministic_items_from_attachments(source):
    items = []
    for attachment in source.attachments or []:
        if attachment.get("status") != "parsed":
            continue
        for row in attachment.get("lines") or []:
            name = row.get("requested_item_name") or row.get("item_name") or row.get("raw_name") or ""
            clean_name = _clean_contract_item_name(name)
            if _is_contract_item_noise(clean_name):
                continue
            items.append(
                {
                    "item_name": clean_name,
                    "suggested_item_name": clean_name,
                    "quantity": row.get("quantity"),
                    "unit": row.get("unit") or "",
                    "unit_price": row.get("unit_price"),
                    "currency": "AED",
                    "project": "",
                    "contact_text": source.sender,
                    "source_text": row.get("raw_line") or row.get("raw_source_text") or name,
                    "source_filename": attachment.get("filename") or "",
                    "source_page": str(row.get("source_page") or row.get("page_number") or ""),
                    "confidence": float(row.get("parse_confidence") or 0.65),
                    "reason": "Detected from Gmail attachment.",
                }
            )
    return items


def _ai_contract_instructions(run):
    return (
        "You extract contract intelligence for a pharmacy yearly supply opportunity. "
        "Classify the source email as inquiry, quotation, lpo, followup, irrelevant, or unknown. "
        "Extract only real requested/quoted/orderable product lines. Ignore signatures, totals, headers, greetings, legal text, and metadata. "
        "Normalize item names conservatively without inventing products. Preserve brand/spec/pack details when present. "
        "Return quantities, units, unit prices, project/contact/source details when available. "
        "This is review-only. Do not suggest creating records silently."
    )


def _ai_items_for_source(source, run):
    settings_obj = QuotationSettings.get_solo()
    status = settings_ai_status(settings_obj)
    if status["status"] != "ai_available":
        raise AIParseError(status["label"])

    attachment_lines = []
    for attachment in source.attachments or []:
        for row in (attachment.get("lines") or [])[:80]:
            attachment_lines.append(
                {
                    "filename": attachment.get("filename"),
                    "item": row.get("requested_item_name") or row.get("item_name") or row.get("raw_name"),
                    "quantity": row.get("quantity"),
                    "unit": row.get("unit"),
                    "unit_price": row.get("unit_price"),
                    "raw": row.get("raw_line") or row.get("raw_source_text"),
                }
            )
    text_context = json.dumps(
        {
            "target_company": run.target_company_name,
            "subject": source.subject,
            "sender": source.sender,
            "recipients": source.recipients,
            "sent_at": source.sent_at.isoformat() if source.sent_at else "",
            "snippet": source.snippet,
            "body_text": (source.body_text or "")[:MAX_ANALYSIS_CHARS],
            "attachment_rows": attachment_lines[:120],
        },
        default=str,
    )
    provider_name = status["availability"]["provider"]
    model = status["availability"]["text_model"]
    provider = get_ai_parse_provider(provider_name)
    parsed, _usage = provider.clean_rows(
        mode="contract_intelligence",
        model=model,
        instructions=_ai_contract_instructions(run),
        text_context=text_context[:MAX_ANALYSIS_CHARS],
        image_data_urls=[],
        json_schema=CONTRACT_AI_JSON_SCHEMA,
        schema_name="contract_intelligence_items",
    )
    return parsed


def _create_contract_item(run, source, payload, *, requested_date=None):
    if not isinstance(payload, dict):
        return None
    raw_name = (payload.get("suggested_item_name") or payload.get("item_name") or "").strip()
    name = _clean_contract_item_name(raw_name)
    if _is_contract_item_noise(name):
        return None
    try:
        product_match = suggest_product_for_text(name, run.company).product if run.company_id else suggest_product_for_text(name, None).product
    except Exception:
        product_match = None
    return ContractIntelligenceItem.objects.create(
        run=run,
        source=source,
        product=product_match,
        original_item_name=_clean_contract_item_name(payload.get("item_name") or raw_name or name)[:500],
        suggested_item_name=name[:500],
        quantity=_quantity(payload.get("quantity")),
        unit=(payload.get("unit") or "")[:80],
        unit_price=_decimal(payload.get("unit_price")),
        currency=(payload.get("currency") or "AED")[:3],
        requested_date=requested_date,
        project=(payload.get("project") or "")[:255],
        contact_text=(payload.get("contact_text") or source.sender or "")[:255],
        source_text=payload.get("source_text") or "",
        source_filename=(payload.get("source_filename") or "")[:255],
        source_page=str(payload.get("source_page") or "")[:30],
        confidence=max(0.0, min(_confidence(payload.get("confidence")), 1.0)),
        ai_reason=payload.get("reason") or "",
        status=ContractIntelligenceItem.STATUS_SUGGESTED,
    )


def analyze_contract_run(run, user, *, use_ai=True, source_limit=None):
    created = 0
    failed = 0
    warnings = []
    requested_limit = min(
        max(int(source_limit or getattr(settings, "QUOTATION_CONTRACT_ANALYZE_BATCH_SIZE", 25)), 1),
        100,
    )
    if use_ai:
        effective_limit = min(
            requested_limit,
            max(int(getattr(settings, "QUOTATION_CONTRACT_AI_ANALYZE_BATCH_SIZE", DEFAULT_AI_ANALYSIS_BATCH_SIZE)), 1),
        )
    else:
        effective_limit = requested_limit
    source_queryset = ContractIntelligenceSource.objects.filter(run=run)
    pending_queryset = source_queryset.exclude(status__in=["analyzed", "failed"])
    sources = list(pending_queryset.order_by("-sent_at", "-created_at")[:effective_limit])
    if not sources:
        total_items = ContractIntelligenceItem.objects.filter(run=run).count()
        pending_sources = pending_queryset.count()
        if total_items:
            active_items = ContractIntelligenceItem.objects.filter(run=run).exclude(
                status=ContractIntelligenceItem.STATUS_REJECTED
            ).count()
            run.status = ContractIntelligenceRun.STATUS_REVIEW if active_items else ContractIntelligenceRun.STATUS_READY
            cleanup_result = {"skipped": True, "reason": "No pending sources. Existing extracted rows were kept unchanged."}
        else:
            cleanup_result = None
        if not total_items and source_queryset.exists():
            run.status = ContractIntelligenceRun.STATUS_READY
        elif not total_items:
            run.status = ContractIntelligenceRun.STATUS_DRAFT
        run.ai_status = "no_pending_sources"
        run.completed_at = timezone.now()
        run.warnings = []
        refresh_contract_run_summary(run)
        run.save(update_fields=["status", "ai_status", "completed_at", "warnings", "summary", "updated_at"])
        return {
            "items_created": 0,
            "sources_failed": source_queryset.filter(status="failed").count(),
            "sources_analyzed": 0,
            "sources_processed": 0,
            "pending_sources": pending_sources,
            "warnings": [],
            "no_pending_sources": True,
            "cleanup": cleanup_result,
        }

    run.status = ContractIntelligenceRun.STATUS_ANALYZING
    run.ai_status = "running"
    run.started_at = timezone.now()
    run.save(update_fields=["status", "ai_status", "started_at", "updated_at"])
    try:
        connection = user.quotation_gmail_connection
    except GmailOAuthConnection.DoesNotExist:
        connection = None
    for source in sources:
        source.status = "analyzing"
        source.save(update_fields=["status", "updated_at"])
        if not source.body_text and source.gmail_message_id:
            if not connection or connection.status != GmailOAuthConnection.STATUS_CONNECTED:
                failed += 1
                source.status = "failed"
                source.error = "Connect Gmail read-only to fetch message content before analysis."
                source.save(update_fields=["status", "error", "updated_at"])
                continue
            try:
                source = hydrate_contract_source(source, connection, include_attachments=run.include_attachments)
            except Exception as exc:
                failed += 1
                source.status = "failed"
                source.error = f"Could not fetch Gmail source content: {str(exc)[:420]}"
                source.save(update_fields=["status", "error", "updated_at"])
                warnings.append(f"{source.subject or source.gmail_message_id}: {source.error[:140]}")
                continue
        ContractIntelligenceItem.objects.filter(run=run, source=source).delete()
        classification, confidence = _classify_source_deterministically(source)
        payload = None
        if use_ai:
            try:
                payload = _ai_items_for_source(source, run)
                classification = payload.get("classification") or classification
                confidence = float(payload.get("confidence") or confidence)
            except Exception as exc:
                error_text = str(exc)
                if "timed out" in error_text.lower() or "timeout" in error_text.lower():
                    error_text = (
                        "AI request timed out; deterministic extraction used. "
                        "Try fewer sources per AI batch or increase QUOTATION_AI_PARSE_TIMEOUT_SECONDS. "
                        f"{error_text}"
                    )
                else:
                    error_text = f"AI unavailable or failed; deterministic extraction used. {error_text}"
                warnings.append(f"{source.subject or source.gmail_message_id}: {error_text[:220]}")
        try:
            source.classification = classification
            source.confidence = max(0.0, min(confidence, 1.0))
            item_payloads = (payload or {}).get("items") if payload else None
            item_payloads = item_payloads or (_deterministic_items_from_attachments(source) + _deterministic_items_from_text(source))
            item_errors = []
            for item_payload in item_payloads:
                try:
                    if _create_contract_item(run, source, item_payload, requested_date=_source_date(source)):
                        created += 1
                except Exception as item_exc:
                    item_errors.append(str(item_exc)[:180])
            if item_errors:
                warnings.append(
                    f"{source.subject or source.gmail_message_id}: skipped {len(item_errors)} malformed extracted row(s)."
                )
            source.status = "analyzed"
            source.error = ""
            source.save(update_fields=["classification", "confidence", "status", "error", "updated_at"])
        except Exception as exc:
            failed += 1
            source.status = "failed"
            source.error = str(exc)[:500]
            source.save(update_fields=["status", "error", "updated_at"])
    cleanup_result = clean_contract_run_items(run, source_ids=[source.id for source in sources])
    total_items = ContractIntelligenceItem.objects.filter(run=run).exclude(status=ContractIntelligenceItem.STATUS_REJECTED).count()
    pending_sources = ContractIntelligenceSource.objects.filter(run=run).exclude(status__in=["analyzed", "failed"]).count()
    if total_items:
        run.status = ContractIntelligenceRun.STATUS_REVIEW
    elif pending_sources and sources:
        run.status = ContractIntelligenceRun.STATUS_READY
    else:
        run.status = ContractIntelligenceRun.STATUS_FAILED
    run.ai_status = "completed_with_warnings" if warnings else "completed"
    run.completed_at = timezone.now()
    run.warnings = warnings
    refresh_contract_run_summary(run)
    run.save(update_fields=["status", "ai_status", "completed_at", "warnings", "summary", "updated_at"])
    audit_log(
        user,
        QuotationAuditLog.ACTION_IMPORTED,
        run,
        company=run.company,
        message="Analyzed contract intelligence source emails.",
        changes={
            "items_created": created,
            "sources_failed": failed,
            "sources_analyzed_this_batch": len(sources) - failed,
            "pending_sources": pending_sources,
            "warnings": warnings[:10],
            "cleanup": cleanup_result,
        },
    )
    return {
        "items_created": created,
        "sources_failed": failed,
        "sources_analyzed": len(sources) - failed,
        "sources_processed": len(sources),
        "pending_sources": pending_sources,
        "warnings": warnings,
        "cleanup": cleanup_result,
    }


def clean_contract_run_items(run, *, source_ids=None, limit=None, save_summary=True):
    """Clean existing extracted rows without creating durable catalog data."""
    total = 0
    updated = 0
    noise_rejected = 0
    already_clean = 0
    skipped_approved = 0
    queryset = ContractIntelligenceItem.objects.filter(run=run).select_related("source", "product")
    if source_ids is not None:
        queryset = queryset.filter(source_id__in=list(source_ids))
    if limit:
        queryset = queryset.order_by("id")[: int(limit)]
    for item in queryset:
        total += 1
        if item.status == ContractIntelligenceItem.STATUS_APPROVED:
            skipped_approved += 1
            continue

        current_name = item.suggested_item_name or item.original_item_name
        cleaned_name = _clean_contract_item_name(current_name)
        is_noise = _is_contract_item_noise(cleaned_name or current_name)
        update_fields = []

        if is_noise:
            if item.status != ContractIntelligenceItem.STATUS_REJECTED:
                item.status = ContractIntelligenceItem.STATUS_REJECTED
                update_fields.append("status")
                noise_rejected += 1
            else:
                already_clean += 1
            note = "Marked as metadata/noise by contract intelligence cleanup."
            if note not in (item.ai_reason or ""):
                item.ai_reason = f"{item.ai_reason}\n{note}".strip() if item.ai_reason else note
                update_fields.append("ai_reason")
        elif cleaned_name and cleaned_name != item.suggested_item_name:
            item.suggested_item_name = cleaned_name[:500]
            item.normalized_item_name = normalize_label(cleaned_name)
            update_fields.extend(["suggested_item_name", "normalized_item_name"])
            updated += 1
        else:
            already_clean += 1

        if update_fields:
            item.save(update_fields=sorted(set(update_fields + ["updated_at"])))

    if save_summary:
        refresh_contract_run_summary(run)
        run.save(update_fields=["summary", "updated_at"])
    return {
        "total": total,
        "updated": updated,
        "noise_rejected": noise_rejected,
        "already_clean": already_clean,
        "skipped_approved": skipped_approved,
    }


def refresh_contract_run_summary(run):
    sources = ContractIntelligenceSource.objects.filter(run=run)
    items = ContractIntelligenceItem.objects.filter(run=run)
    active_items = items.exclude(status=ContractIntelligenceItem.STATUS_REJECTED)
    rejected_noise_count = items.filter(status=ContractIntelligenceItem.STATUS_REJECTED).count()
    grouped = defaultdict(lambda: {"count": 0, "latest_date": "", "last_price": "", "units": set(), "source_count": 0})
    source_ids_by_item = defaultdict(set)
    for item in active_items:
        item_name = (item.suggested_item_name or item.original_item_name or "").strip()
        key = normalize_label(item_name)
        if not key:
            continue
        row = grouped[key]
        row["item_name"] = item_name
        row["count"] += 1
        if item.unit:
            row["units"].add(item.unit)
        if item.source_id:
            source_ids_by_item[key].add(item.source_id)
        if item.requested_date and (not row["latest_date"] or item.requested_date.isoformat() > row["latest_date"]):
            row["latest_date"] = item.requested_date.isoformat()
            row["last_price"] = str(item.unit_price) if item.unit_price is not None else ""
    top_items = []
    for key, row in grouped.items():
        top_items.append(
            {
                "normalized": key,
                "item_name": row["item_name"],
                "count": row["count"],
                "source_count": len(source_ids_by_item[key]),
                "latest_date": row["latest_date"],
                "last_price": row["last_price"],
                "units": sorted(row["units"])[:5],
            }
        )
    top_items.sort(key=lambda row: (-row["count"], row["item_name"]))
    run.summary = {
        "sources": sources.count(),
        "sources_candidate": sources.filter(status="candidate").count(),
        "sources_fetched": sources.filter(status="fetched").count(),
        "sources_analyzed": sources.filter(status="analyzed").count(),
        "sources_failed": sources.filter(status="failed").count(),
        "discovery_exhausted": run.discovery_exhausted,
        "discovery_result_estimate": run.discovery_result_estimate,
        "discovery_batch_size": run.discovery_batch_size,
        "domain_hint": run.sender_domain_hint,
        "inquiries": sources.filter(classification=ContractIntelligenceSource.CLASS_INQUIRY).count(),
        "quotations": sources.filter(classification=ContractIntelligenceSource.CLASS_QUOTATION).count(),
        "lpos": sources.filter(classification=ContractIntelligenceSource.CLASS_LPO).count(),
        "items": active_items.count(),
        "raw_items": items.count(),
        "rejected_noise_items": rejected_noise_count,
        "unique_items": len(grouped),
        "matched_products": active_items.exclude(product__isnull=True).count(),
        "needs_review": active_items.filter(status=ContractIntelligenceItem.STATUS_NEEDS_REVIEW).count(),
        "top_items": top_items[:25],
    }
    return run.summary


def build_contract_intelligence_export(run):
    refresh_contract_run_summary(run)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Items"
    ws.append([f"Contract Intelligence: {run.target_company_name}"])
    ws.append(["Generated", timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")])
    ws.append(["Sources", run.summary.get("sources", 0), "Extracted rows", run.summary.get("items", 0), "Unique items", run.summary.get("unique_items", 0)])
    ws.append([])
    headers = [
        "Item",
        "Normalized",
        "Qty",
        "Unit",
        "Unit Price",
        "Currency",
        "Date",
        "Source",
        "Email Subject",
        "Product Match",
        "Confidence",
        "Reason",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(horizontal="center")
    for item in (
        run.items.exclude(status=ContractIntelligenceItem.STATUS_REJECTED)
        .select_related("source", "product")
        .order_by("normalized_item_name", "-requested_date", "-id")
    ):
        ws.append(
            [
                item.suggested_item_name or item.original_item_name,
                item.normalized_item_name,
                item.quantity,
                item.unit,
                item.unit_price,
                item.currency,
                item.requested_date.strftime("%d/%m/%Y") if item.requested_date else "",
                item.source_filename or (item.source.sender if item.source else ""),
                item.source.subject if item.source else "",
                item.product.name if item.product else "",
                item.confidence,
                item.ai_reason,
            ]
        )
    widths = [36, 30, 10, 12, 14, 10, 12, 28, 44, 34, 12, 50]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:L{ws.max_row}"

    summary = wb.create_sheet("Summary")
    summary.append(["Item", "Mentions", "Sources", "Latest Date", "Last Price", "Units"])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
    for row in run.summary.get("top_items", []):
        summary.append(
            [
                row["item_name"],
                row["count"],
                row["source_count"],
                row["latest_date"],
                row["last_price"],
                ", ".join(row["units"]),
            ]
        )
    for index, width in enumerate([42, 12, 12, 14, 14, 24], start=1):
        summary.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
