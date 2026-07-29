from datetime import timedelta
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import QuotationLine
from .pdf_config import get_quotation_pdf_config


PRIMARY = "0F766E"
BORDER = "D1D5DB"
TEXT = "111827"
MUTED = "6B7280"


def _local_date(value):
    if not value:
        return timezone.localdate()
    if hasattr(value, "date"):
        return timezone.localtime(value).date()
    return value


def _valid_until(quotation, config):
    if quotation.valid_until:
        return quotation.valid_until
    return _local_date(quotation.created_at) + timedelta(days=config.validity_days)


def _payment_terms(quotation, config):
    if getattr(quotation, "payment_terms", ""):
        return quotation.get_payment_terms_display()
    return config.payment_terms


def _safe_number(value):
    return float(value or 0)


def _optional_text(value):
    value = "" if value is None else str(value).strip()
    if value.lower() in {"", "-", "—", "n/a", "na", "none", "null"}:
        return ""
    return value


def build_quotation_excel(quotation):
    config = get_quotation_pdf_config()
    quote_date = _local_date(quotation.created_at)
    valid_until = _valid_until(quotation, config)
    show_brand_column = bool(getattr(quotation, "show_brand_column", False))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet.sheet_view.showGridLines = True

    thin = Side(style="thin", color=BORDER)
    header_border = Border(bottom=thin)
    total_border = Border(top=thin)
    header_fill = PatternFill("solid", fgColor="E5E7EB")

    sheet["A1"] = "Quotation Export"
    sheet["A1"].font = Font(bold=True, size=14, color=TEXT)
    sheet["A2"] = config.company_name or "Al Ameen Pharmacy"
    sheet["A2"].font = Font(bold=True, color=PRIMARY)

    contact = quotation.contact
    info_rows = [
        ("Customer", quotation.company.name),
        ("Customer Address", _optional_text(getattr(quotation.company, "billing_address", ""))),
        ("Customer TRN", _optional_text(getattr(quotation.company, "trn", ""))),
        ("Attention", _optional_text(contact.name if contact else "")),
        ("Contact No.", _optional_text(getattr(contact, "phone", "") if contact else "")),
        ("Contact Email", _optional_text(getattr(contact, "email", "") if contact else "")),
        ("Quotation #", quotation.quotation_number),
        ("Date", quote_date),
        ("Valid Until", valid_until),
        ("Currency", quotation.currency),
        ("Status", quotation.get_status_display()),
        ("Payment Terms", _payment_terms(quotation, config) or "-"),
        ("Prepared By", quotation.created_by.username if quotation.created_by else "-"),
    ]
    info_rows = [(label, value) for label, value in info_rows if str(value or "").strip()]
    for row_offset, (label, value) in enumerate(info_rows, start=4):
        label_cell = sheet.cell(row=row_offset, column=1, value=label)
        value_cell = sheet.cell(row=row_offset, column=2, value=value)
        label_cell.font = Font(bold=True, color=MUTED)
        label_cell.alignment = Alignment(vertical="top")
        value_cell.alignment = Alignment(vertical="top")
        if hasattr(value, "year"):
            value_cell.number_format = "dd/mm/yyyy"

    table_start = 18
    columns = [
        ("serial", "S. No."),
        ("item", "Item Description"),
    ]
    if show_brand_column:
        columns.append(("brand", "Brand"))
    columns.extend(
        [
            ("quantity", "Qty"),
            ("unit", "Unit"),
            ("unit_price", "Unit Price"),
            ("vat_rate", "VAT %"),
            ("vat_amount", "VAT Amount"),
            ("line_total", "Line Total"),
        ]
    )
    column_numbers = {
        key: column
        for column, (key, _header) in enumerate(columns, start=1)
    }
    headers = [header for _key, header in columns]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_start, column=column, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    money_format = "#,##0.00"
    unit_price_format = "#,##0.00#"
    lines = quotation.lines.exclude(match_status=QuotationLine.MATCH_IGNORED).order_by("sort_order", "id")
    for index, line in enumerate(lines, start=1):
        row = table_start + index
        values_by_key = {
            "serial": index,
            "item": line.item_name_snapshot,
            "brand": str(getattr(line, "brand_name_snapshot", "") or "").strip() or "-",
            "quantity": _safe_number(line.quantity),
            "unit": line.unit or "-",
            "unit_price": _safe_number(line.unit_price),
            "vat_rate": _safe_number(line.vat_rate),
            "vat_amount": _safe_number(line.vat_amount),
            "line_total": _safe_number(line.line_total),
        }
        for column, (key, _header) in enumerate(columns, start=1):
            value = values_by_key[key]
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(key in {"item", "brand"}))
            if key == "serial":
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key in {"quantity", "unit_price", "vat_rate", "vat_amount", "line_total"}:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            if key == "quantity":
                cell.number_format = "#,##0.000"
            elif key == "vat_rate":
                cell.number_format = "0.00"
            elif key == "unit_price":
                cell.number_format = unit_price_format
            elif key in {"vat_amount", "line_total"}:
                cell.number_format = money_format

    last_line_row = table_start + max(lines.count(), 1)
    totals_start = last_line_row + 2
    totals = [
        ("Subtotal", _safe_number(quotation.subtotal)),
        ("VAT", _safe_number(quotation.vat_total)),
        ("Grand Total", _safe_number(quotation.total)),
    ]
    totals_label_column = column_numbers["vat_amount"]
    totals_value_column = column_numbers["line_total"]
    for row_offset, (label, value) in enumerate(totals, start=totals_start):
        label_cell = sheet.cell(row=row_offset, column=totals_label_column, value=label)
        value_cell = sheet.cell(row=row_offset, column=totals_value_column, value=value)
        for cell in (label_cell, value_cell):
            cell.font = Font(bold=True, color=PRIMARY if label == "Grand Total" else TEXT)
            cell.alignment = Alignment(horizontal="right")
            if label == "Subtotal":
                cell.border = total_border
        value_cell.number_format = money_format

    footer_start = totals_start + len(totals) + 2
    sheet.cell(row=footer_start, column=1, value="Terms").font = Font(bold=True, color=MUTED)
    sheet.cell(row=footer_start, column=2, value=config.default_terms or "")
    sheet.cell(row=footer_start + 1, column=1, value="Payment Terms").font = Font(bold=True, color=MUTED)
    sheet.cell(row=footer_start + 1, column=2, value=_payment_terms(quotation, config) or "-")

    if show_brand_column:
        widths = {
            "A": 11,
            "B": 38,
            "C": 24,
            "D": 12,
            "E": 16,
            "F": 14,
            "G": 10,
            "H": 14,
            "I": 14,
        }
    else:
        widths = {
            "A": 11,
            "B": 48,
            "C": 12,
            "D": 16,
            "E": 14,
            "F": 10,
            "G": 14,
            "H": 14,
        }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
