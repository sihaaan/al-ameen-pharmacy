import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from .ai_parsing import AIParseError, clean_preview_with_ai
from .import_parsers import read_upload_bytes
from .import_parsers import parse_file_preview, parse_text_preview
from .models import normalize_label


PRICE_REFERENCE_WORKBOOK_EXTENSIONS = {".xlsx"}
PRICE_REFERENCE_PREVIEW_EXTENSIONS = {".pdf", ".xls", ".xlsb"}
PRICE_REFERENCE_FILE_EXTENSIONS = PRICE_REFERENCE_WORKBOOK_EXTENSIONS | PRICE_REFERENCE_PREVIEW_EXTENSIONS
PRICE_HEADER_ALIASES = {
    "serial": {"sno", "s no", "sl no", "sr no", "serial", "#"},
    "item": {"item", "items", "item description", "item desc", "description", "material description"},
    "unit": {"unit", "uom", "uom ", "uom.", "pack", "packing"},
    "quantity": {"qty", "quantity", "req quantity", "req qty"},
    "unit_price": {"uprice", "u price", "u/price", "unit price", "rate", "al ameen", "u/p"},
    "total": {"total", "amount", "net value", "net", "gross total", "g total", "total "},
    "vat": {"vat", "vat amount", "tax"},
}


@dataclass
class PriceReferenceRow:
    item_name: str
    unit: str
    quantity: str
    unit_price: Decimal
    vat_rate: Decimal
    vat_amount: Decimal | None
    total: Decimal | None
    sheet_name: str
    row_number: int
    sequence: int
    raw_values: list[str]


def _clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_key(value):
    value = normalize_label(_clean_cell(value))
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def _money(value):
    text = _clean_cell(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _quantity(value):
    text = _clean_cell(value).replace(",", "")
    if not text:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return match.group(0) if match else text[:50]


def _is_numeric_text(value):
    text = _clean_cell(value).replace(",", "")
    if not text:
        return False
    try:
        Decimal(text)
        return True
    except InvalidOperation:
        return False


def _looks_like_item(value):
    text = _clean_cell(value)
    key = normalize_label(text)
    if not key or len(key) < 3:
        return False
    if key in {"total", "subtotal", "vat", "gross total", "net value"}:
        return False
    if re.search(r"\b(total|subtotal|vat|email|phone|quotation|invoice)\b", key):
        return False
    return bool(re.search(r"[A-Za-z]", text))


def _map_header(row):
    mapped = {}
    for index, value in enumerate(row):
        key = _header_key(value)
        if not key:
            continue
        for role, aliases in PRICE_HEADER_ALIASES.items():
            if key in aliases and role not in mapped:
                mapped[role] = index
                break
    if "item" in mapped and "unit_price" in mapped:
        return mapped
    return {}


def _infer_no_header_columns(row):
    values = [_clean_cell(value) for value in row]
    if len(values) < 5:
        return {}
    if _money(values[4]) is not None and _looks_like_item(values[1]):
        return {"serial": 0, "item": 1, "unit": 2, "quantity": 3, "unit_price": 4, "total": 5, "vat": 6}
    return {}


def _value(row, columns, role):
    index = columns.get(role)
    if index is None or index >= len(row):
        return ""
    return row[index]


def _parse_reference_row(row, columns, *, sheet_name, row_number, sequence):
    item_name = _clean_cell(_value(row, columns, "item"))
    if not _looks_like_item(item_name):
        return None
    unit_price = _money(_value(row, columns, "unit_price"))
    if unit_price is None:
        return None
    unit = _clean_cell(_value(row, columns, "unit"))[:50]
    if _is_numeric_text(unit):
        unit = ""
    quantity = _quantity(_value(row, columns, "quantity"))
    vat_amount = _money(_value(row, columns, "vat"))
    total = _money(_value(row, columns, "total"))
    vat_rate = Decimal("0.00")
    if vat_amount is not None and unit_price and quantity:
        try:
            subtotal = Decimal(quantity) * unit_price
            if subtotal > 0 and vat_amount > 0:
                rate = (vat_amount / subtotal) * Decimal("100")
                vat_rate = Decimal("5.00") if Decimal("3.00") <= rate <= Decimal("7.00") else Decimal("0.00")
        except (InvalidOperation, ValueError):
            vat_rate = Decimal("0.00")
    return PriceReferenceRow(
        item_name=item_name[:255],
        unit=unit,
        quantity=quantity,
        unit_price=unit_price,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        total=total,
        sheet_name=sheet_name,
        row_number=row_number,
        sequence=sequence,
        raw_values=[_clean_cell(value) for value in row],
    )


def parse_price_reference_workbook(uploaded_file):
    data = read_upload_bytes(uploaded_file)
    filename = Path(uploaded_file.name or "").name
    extension = Path(filename).suffix.lower()
    if extension not in PRICE_REFERENCE_WORKBOOK_EXTENSIONS:
        raise ValidationError("Upload an .xlsx price reference workbook.")
    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationError(f"Could not read price reference workbook: {exc}") from exc

    rows = []
    warnings = []
    sequence = 0
    try:
        for sheet in workbook.worksheets[:10]:
            if getattr(sheet, "sheet_state", "visible") != "visible":
                continue
            current_columns = {}
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = tuple(row or [])
                header_columns = _map_header(values)
                if header_columns:
                    current_columns = header_columns
                    continue
                columns = current_columns or _infer_no_header_columns(values)
                if not columns:
                    continue
                parsed = _parse_reference_row(
                    values,
                    columns,
                    sheet_name=sheet.title,
                    row_number=row_number,
                    sequence=sequence,
                )
                sequence += 1
                if parsed:
                    rows.append(parsed)
    finally:
        workbook.close()
    if not rows:
        warnings.append("No item price rows were detected in the reference workbook.")
    return rows, {"filename": filename, "row_count": len(rows), "warnings": warnings}


def _reference_rows_from_preview(preview):
    rows = []
    warnings = list(preview.get("warnings") or [])
    sequence = 0
    for index, line in enumerate(preview.get("lines") or [], start=1):
        if (line.get("parse_status") or line.get("status")) == "ignored":
            continue
        item_name = line.get("raw_name") or line.get("item_name") or ""
        unit_price = _money(line.get("unit_price"))
        if not _looks_like_item(item_name) or unit_price is None:
            continue
        rows.append(
            PriceReferenceRow(
                item_name=_clean_cell(item_name)[:255],
                unit=_clean_cell(line.get("unit"))[:50],
                quantity=_quantity(line.get("quantity")),
                unit_price=unit_price,
                vat_rate=_money(line.get("vat_rate")) or Decimal("0.00"),
                vat_amount=_money(line.get("vat_amount")),
                total=_money(line.get("line_total") or line.get("total")),
                sheet_name=preview.get("source_filename") or preview.get("source_type") or "Price reference",
                row_number=index,
                sequence=sequence,
                raw_values=[
                    _clean_cell(item_name),
                    _clean_cell(line.get("quantity")),
                    _clean_cell(line.get("unit")),
                    _clean_cell(line.get("unit_price")),
                    _clean_cell(line.get("line_total") or line.get("total")),
                ],
            )
        )
        sequence += 1
    if not rows:
        warnings.append("No item price rows were detected in the price reference source.")
    return rows, {
        "filename": preview.get("source_filename") or preview.get("source_type") or "Pasted price reference",
        "row_count": len(rows),
        "warnings": warnings,
        "parse_method": preview.get("parse_method", ""),
        "source_type": preview.get("source_type", ""),
        "ai_status": preview.get("ai_status", ""),
        "ai_status_label": preview.get("ai_status_label", ""),
    }


def parse_price_reference_source(uploaded_file=None, *, raw_text="", raw_html="", use_ai=False, actor=None):
    if uploaded_file:
        filename = Path(uploaded_file.name or "").name
        extension = Path(filename).suffix.lower()
        if extension in PRICE_REFERENCE_WORKBOOK_EXTENSIONS:
            return parse_price_reference_workbook(uploaded_file)
        if extension not in PRICE_REFERENCE_FILE_EXTENSIONS:
            raise ValidationError("Upload an .xlsx workbook, .xls/.xlsb workbook, or .pdf price reference file.")
        preview = parse_file_preview(uploaded_file)
    elif str(raw_text or "").strip():
        preview = parse_text_preview(raw_text, raw_html=raw_html)
    else:
        raise ValidationError("Upload a price reference file or paste price reference text.")

    ai_warning = ""
    if use_ai:
        try:
            preview = clean_preview_with_ai(preview, actor=actor, requested_mode="auto", allow_vision=True)
        except AIParseError as exc:
            ai_warning = f"AI price reference parsing was unavailable; deterministic parse was used. Detail: {exc}"

    rows, meta = _reference_rows_from_preview(preview)
    if ai_warning:
        meta.setdefault("warnings", []).append(ai_warning)
    return rows, meta


def _tokens(value):
    return {token for token in normalize_label(value).split() if len(token) >= 2}


def _match_score(inquiry_name, reference_name):
    left = normalize_label(inquiry_name)
    right = normalize_label(reference_name)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        shorter = min(len(left), len(right))
        longer = max(len(left), len(right))
        return max(0.88, min(0.98, shorter / longer + 0.12))
    left_tokens = _tokens(left)
    right_tokens = _tokens(right)
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens)
    containment = overlap / min(len(left_tokens), len(right_tokens))
    jaccard = overlap / len(left_tokens | right_tokens)
    return max(containment * 0.88, jaccard)


def _row_payload(row, score):
    return {
        "item_name": row.item_name,
        "unit": row.unit,
        "quantity": row.quantity,
        "unit_price": str(row.unit_price),
        "vat_rate": str(row.vat_rate),
        "vat_amount": str(row.vat_amount) if row.vat_amount is not None else "",
        "total": str(row.total) if row.total is not None else "",
        "sheet_name": row.sheet_name,
        "row_number": row.row_number,
        "confidence": round(score, 3),
        "match_label": "Exact" if score >= 0.999 else "Likely",
    }


def _is_blank(value):
    return value is None or str(value).strip() == ""


def apply_price_reference_to_preview(preview, reference_rows):
    updated_lines = []
    matched_count = 0
    review_count = 0
    unmatched_count = 0
    for line in preview.get("lines", []) or []:
        original_line = dict(line)
        line = dict(line)
        raw_name = line.get("raw_name") or line.get("item_name") or ""
        best = None
        best_score = 0.0
        for row in reference_rows:
            score = _match_score(raw_name, row.item_name)
            if score > best_score or (score == best_score and best and row.sequence > best.sequence):
                best = row
                best_score = score
        matched_reference = best if best and best_score >= 0.82 else None
        if matched_reference:
            line["unit_price"] = str(matched_reference.unit_price)
            line["vat_rate"] = str(matched_reference.vat_rate)
            line["price_reference_match"] = _row_payload(matched_reference, best_score)
            line["price_reference_status"] = "matched"
            line["parse_status"] = line.get("parse_status") or "parsed"
            matched_count += 1
        elif best and best_score >= 0.62:
            line["price_reference_match"] = _row_payload(best, best_score)
            line["price_reference_status"] = "needs_review"
            review_count += 1
        else:
            line["price_reference_status"] = "unmatched"
            unmatched_count += 1
        line["quantity"] = original_line.get("quantity", "")
        if matched_reference and _is_blank(original_line.get("unit")) and matched_reference.unit:
            line["unit"] = matched_reference.unit
        else:
            line["unit"] = original_line.get("unit", "")
        updated_lines.append(line)
    return {
        **preview,
        "lines": updated_lines,
        "price_reference_summary": {
            "matched_count": matched_count,
            "needs_review_count": review_count,
            "unmatched_count": unmatched_count,
            "reference_row_count": len(reference_rows),
        },
    }
