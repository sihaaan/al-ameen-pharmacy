import hashlib
import re
import zipfile
from io import BytesIO
from pathlib import Path

import filetype
import pdfplumber
from django.conf import settings
from django.core.exceptions import ValidationError
from openpyxl import load_workbook as load_openpyxl_workbook
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from python_calamine import load_workbook as load_calamine_workbook

try:
    import fitz
except Exception:  # pragma: no cover - optional runtime dependency guard
    fitz = None

try:
    import magic
except Exception:  # pragma: no cover - libmagic is optional and platform dependent
    magic = None

from .import_rules import (
    classify_header_cell,
    detect_header_row,
    parse_inquiry_line,
    parse_html_table_lines,
    parse_structured_row,
    parse_text_table_lines,
    parse_text_lines,
    split_quantity_unit,
    row_to_text,
    is_noise_line,
    summarize_lines,
)
from .ocr import OCRProviderUnavailable, get_ocr_provider
from .private_storage import store_import_source


ALLOWED_EXTENSIONS = {".xlsx", ".xlsb", ".xls", ".pdf"}
ZIP_EXCEL_EXTENSIONS = {".xlsx", ".xlsb"}
OLE_EXCEL_EXTENSIONS = {".xls"}
PDF_MIME = "application/pdf"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSB_MIME = "application/vnd.ms-excel.sheet.binary.macroenabled.12"
XLS_MIME = "application/vnd.ms-excel"
OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
AGGREGATE_ITEM_SUMMARY_RE = re.compile(
    r"\b\d+(?:[.,]\d+)?\s+(?:individual\s+)?(?:line\s+)?items?\b",
    re.IGNORECASE,
)


def max_upload_bytes():
    return int(getattr(settings, "QUOTATION_IMPORT_MAX_UPLOAD_BYTES", 5 * 1024 * 1024))


def max_excel_rows():
    return int(getattr(settings, "QUOTATION_IMPORT_MAX_EXCEL_ROWS", 500))


def max_pdf_pages():
    return int(getattr(settings, "QUOTATION_IMPORT_MAX_PDF_PAGES", 10))


def max_excel_sheets():
    return int(getattr(settings, "QUOTATION_IMPORT_MAX_EXCEL_SHEETS", 10))


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _column_ref(index):
    index += 1
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def read_upload_bytes(uploaded_file):
    if not uploaded_file:
        raise ValidationError("No file was uploaded.")
    limit = max_upload_bytes()
    data = bytearray()
    for chunk in uploaded_file.chunks():
        data.extend(chunk)
        if len(data) > limit:
            raise ValidationError(f"Uploaded file is too large. Maximum size is {limit // (1024 * 1024)} MB.")
    if not data:
        raise ValidationError("Uploaded file is empty.")
    return bytes(data)


def _sniff_mime(data):
    if magic:
        try:
            detected = magic.from_buffer(data[:4096], mime=True)
            if detected:
                return detected
        except Exception:
            pass
    guessed = filetype.guess(data)
    if guessed and guessed.mime:
        return guessed.mime
    if data.startswith(b"%PDF-"):
        return PDF_MIME
    if data.startswith(b"PK") and zipfile.is_zipfile(BytesIO(data)):
        return XLSX_MIME
    if data.startswith(OLE_SIGNATURE):
        return XLS_MIME
    return "application/octet-stream"


def _validate_upload_type(data, filename):
    extension = Path(filename or "").suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValidationError("Unsupported file type. Upload .xlsx, .xlsb, .xls, or .pdf files only.")

    sniffed_mime = _sniff_mime(data)
    if extension == ".pdf":
        if not data.startswith(b"%PDF-"):
            raise ValidationError("Invalid PDF file. The upload does not look like a PDF.")
        return extension, sniffed_mime or PDF_MIME

    if extension in ZIP_EXCEL_EXTENSIONS:
        if not data.startswith(b"PK") or not zipfile.is_zipfile(BytesIO(data)):
            raise ValidationError(f"Invalid Excel file. The upload does not look like a valid {extension} workbook.")
        return extension, sniffed_mime or (XLSB_MIME if extension == ".xlsb" else XLSX_MIME)

    if extension in OLE_EXCEL_EXTENSIONS:
        if not data.startswith(OLE_SIGNATURE):
            raise ValidationError("Invalid Excel file. The upload does not look like a valid .xls workbook.")
        return extension, sniffed_mime or XLS_MIME

    raise ValidationError("Unsupported file type.")


def _preview_response(
    *,
    source_type,
    source_filename="",
    source_mime_type="",
    source_sha256="",
    source_file_ref="",
    source_file_size=None,
    parse_method,
    original_text="",
    lines=None,
    warnings=None,
    skipped_count=0,
    meta=None,
):
    lines = lines or []
    warnings = list(warnings or [])
    aggregate_po_summary_detected = any(
        AGGREGATE_ITEM_SUMMARY_RE.search(
            " ".join(
                str(line.get(key) or "")
                for key in ("raw_line", "raw_source_line", "raw_name", "requested_item_name")
            )
        )
        for line in lines
    )
    if aggregate_po_summary_detected:
        warnings.append(
            "Aggregate PO item summary detected. No individual line outcome should be inferred from it; "
            "staff must review the selected PO document manually."
        )
    warnings = list(dict.fromkeys(warnings))
    summary = summarize_lines(lines, skipped_count=skipped_count)
    payload = {
        "source_type": source_type,
        "source_filename": source_filename,
        "source_mime_type": source_mime_type,
        "source_sha256": source_sha256,
        "source_file_ref": source_file_ref,
        "source_file_size": source_file_size,
        "parse_method": parse_method,
        "original_text": original_text or "",
        "lines": lines,
        "warnings": warnings,
        "summary": summary,
        "meta": {
            "line_count": len(lines),
            **(meta or {}),
            "aggregate_po_summary_detected": aggregate_po_summary_detected,
        },
    }
    return payload


def parse_text_preview(raw_text, raw_html=""):
    lines = []
    skipped = 0
    parse_method = "deterministic_text_v2"
    if raw_html:
        lines, skipped = parse_html_table_lines(raw_html)
        if lines:
            parse_method = "deterministic_clipboard_html_table_v1"
    if not lines:
        lines, skipped = parse_text_lines(raw_text)
    warnings = []
    if not lines:
        warnings.append("No item lines were detected. Review the pasted text or add rows manually.")
    return _preview_response(
        source_type="pasted_text",
        source_mime_type="text/plain",
        source_sha256=hashlib.sha256(str(raw_text or "").encode("utf-8")).hexdigest(),
        parse_method=parse_method,
        original_text=raw_text or "",
        lines=lines,
        warnings=warnings,
        skipped_count=skipped,
        meta={"skipped_noise_lines": skipped},
    )


def parse_file_preview(uploaded_file):
    data = read_upload_bytes(uploaded_file)
    filename = Path(uploaded_file.name or "").name
    extension, sniffed_mime = _validate_upload_type(data, filename)
    sha256 = hashlib.sha256(data).hexdigest()

    if extension in {".xlsx", ".xlsb", ".xls"}:
        preview = parse_excel_preview(
            data,
            filename,
            sniffed_mime,
            sha256,
            extension=extension,
        )
    else:
        preview = parse_pdf_preview(data, filename, sniffed_mime, sha256)

    source_file_ref = store_import_source(data, filename=filename, sha256=sha256)
    preview["source_file_ref"] = source_file_ref
    preview["meta"]["source_file_ref"] = source_file_ref
    return preview


def _mapped_columns(header):
    if not header:
        return {}
    return {
        role: {
            "index": index,
            "column": _column_ref(index),
            "label": header.labels.get(role, ""),
        }
        for role, index in header.columns.items()
    }


def _parse_sheet_rows(sheet_name, rows, *, parser_name):
    header = detect_header_row([row for _, row in rows], max_scan_rows=20)
    metadata = {
        "sheet_name": sheet_name,
        "selected": False,
        "parser": parser_name,
        "header_row": None,
        "mapped_columns": {},
        "score": 0,
        "data_score": 0,
        "rows_seen": len(rows),
        "parsed_rows": 0,
        "skipped_rows": 0,
    }
    if not header or header.score < 5 or header.data_score <= 0:
        return [], metadata

    metadata.update(
        {
            "selected": True,
            "header_row": rows[header.row_offset][0],
            "mapped_columns": _mapped_columns(header),
            "score": header.score,
            "data_score": header.data_score,
        }
    )
    lines = []
    skipped = 0
    for source_row, row in rows[header.row_offset + 1 :]:
        parsed, skipped_reason = parse_structured_row(
            row,
            header,
            source_sheet=sheet_name,
            source_row=source_row,
            base_confidence=0.85,
        )
        if parsed:
            lines.append(parsed)
        elif skipped_reason:
            skipped += 1
    metadata["parsed_rows"] = len(lines)
    metadata["skipped_rows"] = skipped
    return lines, metadata


def _fallback_parse_sheet_text(sheet_name, rows):
    lines = []
    skipped = 0
    for source_row, row in rows:
        raw_line = row_to_text(row)
        if not raw_line:
            continue
        parsed = parse_inquiry_line(raw_line, source_sheet=sheet_name, sheet_name=sheet_name, source_row=source_row, row_number=source_row)
        if parsed:
            lines.append(parsed)
        else:
            skipped += 1
    return lines, skipped


def _openpyxl_rows(data):
    workbook = load_openpyxl_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets[: max_excel_sheets()]:
            if getattr(sheet, "sheet_state", "visible") != "visible":
                continue
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                rows.append((row_index, tuple(row)))
                if len(rows) >= max_excel_rows():
                    break
            yield sheet.title, rows, "openpyxl_structured_v2"
    finally:
        workbook.close()


def _calamine_rows(data):
    workbook = load_calamine_workbook(BytesIO(data))
    try:
        for sheet_name in workbook.sheet_names[: max_excel_sheets()]:
            sheet = workbook.get_sheet_by_name(sheet_name)
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                rows.append((row_index, tuple(row)))
                if len(rows) >= max_excel_rows():
                    break
            yield sheet_name, rows, "calamine_structured_v2"
    finally:
        workbook.close()


def _iter_excel_rows(data, extension):
    if extension == ".xlsx":
        try:
            yield from _openpyxl_rows(data)
            return
        except Exception:
            yield from _calamine_rows(data)
            return
    yield from _calamine_rows(data)


def parse_excel_preview(data, filename, content_type, sha256, *, extension=".xlsx", source_file_ref=""):
    lines = []
    warnings = []
    sheet_metadata = []
    fallback_candidates = []
    skipped_count = 0
    parser_used = "excel_structured_v2"

    try:
        for sheet_name, rows, parser_name in _iter_excel_rows(data, extension):
            if not rows:
                sheet_metadata.append(
                    {
                        "sheet_name": sheet_name,
                        "selected": False,
                        "parser": parser_name,
                        "header_row": None,
                        "mapped_columns": {},
                        "score": 0,
                        "data_score": 0,
                        "rows_seen": 0,
                        "parsed_rows": 0,
                        "skipped_rows": 0,
                    }
                )
                continue
            parser_used = parser_name
            sheet_lines, metadata = _parse_sheet_rows(sheet_name, rows, parser_name=parser_name)
            sheet_metadata.append(metadata)
            if metadata["selected"]:
                lines.extend(sheet_lines)
                skipped_count += metadata["skipped_rows"]
            else:
                fallback_candidates.append((sheet_name, rows))
            if len(rows) >= max_excel_rows():
                warnings.append(f"Stopped reading sheet '{sheet_name}' after {max_excel_rows()} rows.")
    except Exception as exc:
        raise ValidationError(f"Could not read Excel workbook: {exc}") from exc

    if not lines and fallback_candidates:
        warnings.append("No clear header row detected. Parsed visible text rows instead; review all lines carefully.")
        for sheet_name, rows in fallback_candidates:
            fallback_lines, fallback_skipped = _fallback_parse_sheet_text(sheet_name, rows)
            lines.extend(fallback_lines)
            skipped_count += fallback_skipped

    if not lines:
        warnings.append("No item lines were detected in the Excel workbook.")

    selected_sheets = [sheet for sheet in sheet_metadata if sheet.get("selected")]
    return _preview_response(
        source_type="excel",
        source_filename=filename,
        source_mime_type=content_type or XLSX_MIME,
        source_sha256=sha256,
        source_file_ref=source_file_ref,
        source_file_size=len(data),
        parse_method=parser_used,
        lines=lines,
        warnings=warnings,
        skipped_count=skipped_count,
        meta={
            "sheet_metadata": sheet_metadata,
            "selected_sheets": selected_sheets,
            "inspected_sheets": [sheet["sheet_name"] for sheet in sheet_metadata],
            "source_file_ref": source_file_ref,
            "source_file_size": len(data),
        },
    )


def _preflight_pdf(data):
    if not data.startswith(b"%PDF-"):
        raise ValidationError("Invalid PDF file. The upload does not look like a PDF.")
    try:
        reader = PdfReader(BytesIO(data))
    except PdfReadError as exc:
        raise ValidationError(f"Could not read PDF: {exc}") from exc
    if reader.is_encrypted:
        raise ValidationError("Encrypted PDF files are not supported. Please upload an unlocked PDF.")
    page_count = len(reader.pages)
    if page_count > max_pdf_pages():
        raise ValidationError(f"PDF has {page_count} pages. Maximum supported pages: {max_pdf_pages()}.")
    return page_count


def _words_to_layout_text(words):
    """Recreate visually aligned PDF rows with explicit cell separators."""

    normalized_words = []
    for word in words or []:
        if len(word) < 5:
            continue
        text = str(word[4] or "").strip()
        if not text:
            continue
        x0, y0, x1, y1 = (float(word[index]) for index in range(4))
        normalized_words.append(
            {
                "x0": x0,
                "y0": y0,
                "x1": x1,
                "y1": y1,
                "height": max(1.0, y1 - y0),
                "center": (y0 + y1) / 2,
                "text": text,
            }
        )
    normalized_words.sort(key=lambda value: (value["center"], value["x0"]))

    visual_lines = []
    for word in normalized_words:
        if visual_lines:
            current = visual_lines[-1]
            tolerance = max(2.5, min(current["height"], word["height"]) * 0.45)
            if abs(word["center"] - current["center"]) <= tolerance:
                current["words"].append(word)
                count = len(current["words"])
                current["center"] = ((current["center"] * (count - 1)) + word["center"]) / count
                current["height"] = max(current["height"], word["height"])
                continue
        visual_lines.append(
            {
                "center": word["center"],
                "height": word["height"],
                "words": [word],
            }
        )

    rendered_lines = []
    for visual_line in visual_lines:
        line_words = sorted(visual_line["words"], key=lambda value: value["x0"])
        cells = []
        current_cell = [line_words[0]["text"]]
        previous = line_words[0]
        for word in line_words[1:]:
            gap = word["x0"] - previous["x1"]
            separator_gap = max(8.0, min(previous["height"], word["height"]) * 0.9)
            if gap >= separator_gap:
                cells.append(" ".join(current_cell))
                current_cell = [word["text"]]
            else:
                current_cell.append(word["text"])
            previous = word
        cells.append(" ".join(current_cell))
        rendered_lines.append(" | ".join(cells))
    return "\n".join(rendered_lines)


def _extract_pymupdf_text(data):
    text_chunks = []
    word_layout_chunks = []
    page_metadata = []
    if fitz is None:
        return text_chunks, word_layout_chunks, page_metadata

    with fitz.open(stream=data, filetype="pdf") as document:
        for page_number, page in enumerate(document, start=1):
            words = page.get_text("words") or []
            text = page.get_text("text") or ""
            word_layout_text = _words_to_layout_text(words)
            text_chunks.append(text)
            word_layout_chunks.append(word_layout_text)
            page_metadata.append(
                {
                    "page_number": page_number,
                    "text_based": len(words) >= 3 or len(text.strip()) >= 20,
                    "word_count": len(words),
                    "text_length": len(text.strip()),
                    "word_layout_text_length": len(word_layout_text.strip()),
                }
            )
    return text_chunks, word_layout_chunks, page_metadata


def _is_plausible_pdf_item_line(line):
    item_name = str(
        (line or {}).get("requested_item_name")
        or (line or {}).get("raw_name")
        or (line or {}).get("item_name")
        or ""
    ).strip()
    if not item_name or not re.search(r"[A-Za-z]", item_name):
        return False
    normalized = re.sub(r"[^a-z0-9]+", " ", item_name.lower()).strip()
    if classify_header_cell(item_name):
        return False
    if normalized in {
        "description",
        "item",
        "item description",
        "item number",
        "items",
        "ln",
        "line",
        "material description",
        "product",
        "req quote no",
        "request quote no",
        "quantity",
        "qty",
        "unit",
        "unit price",
        "price",
        "amount",
        "total",
    }:
        return False
    return any(
        (line or {}).get(key) not in (None, "")
        for key in ("quantity", "unit", "unit_price", "amount", "line_total")
    )


def _looks_like_pdf_metadata_table(rows):
    if not rows or len(rows) > 8:
        return False
    texts = [row_to_text(row) for _, row in rows]
    texts = [text for text in texts if text]
    if not texts:
        return True
    if any(re.search(r"\b(qty|quantity|req quantity|u\s*price|unit price|total|amount)\b", text, re.IGNORECASE) for text in texts):
        return False
    metadata_hits = 0
    for text in texts:
        if is_noise_line(text):
            metadata_hits += 1
        elif re.search(r"^[\w.+-]+@[\w.-]+\.\w+$", text, re.IGNORECASE):
            metadata_hits += 1
        elif re.search(r"^[\w.-]+\.(com|ae|net|org)$", text, re.IGNORECASE):
            metadata_hits += 1
        elif re.search(r"\b(procurement|officer|manager|buyer|seller|contact)\b", text, re.IGNORECASE):
            metadata_hits += 1
        elif len(text.split()) <= 4 and not re.search(r"\d", text):
            metadata_hits += 1
    return metadata_hits >= max(1, len(texts) - 1)


def _parse_pdfplumber_tables(data):
    lines = []
    page_metadata = []
    table_rows_seen = 0
    skipped_count = 0
    current_header = None
    with pdfplumber.open(BytesIO(data)) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            page_tables_seen = 0
            for table in page.extract_tables() or []:
                table_rows_seen += len(table)
                page_tables_seen += 1
                rows = [(index + 1, tuple(row or [])) for index, row in enumerate(table)]
                header = detect_header_row([row for _, row in rows], max_scan_rows=10)
                if header and header.data_score > 0:
                    current_header = header
                    for row_number, row in rows[header.row_offset + 1 :]:
                        parsed, skipped_reason = parse_structured_row(
                            row,
                            header,
                            source_page=page_index,
                            source_row=row_number,
                            base_confidence=0.80,
                        )
                        if parsed:
                            lines.append(parsed)
                        elif skipped_reason:
                            skipped_count += 1
                elif current_header and rows and len(rows[0][1]) >= max(current_header.columns.values(), default=0) + 1:
                    for row_number, row in rows:
                        parsed, skipped_reason = parse_structured_row(
                            row,
                            current_header,
                            source_page=page_index,
                            source_row=row_number,
                            base_confidence=0.78,
                        )
                        if parsed:
                            lines.append(parsed)
                        elif skipped_reason:
                            skipped_count += 1
                else:
                    if _looks_like_pdf_metadata_table(rows):
                        skipped_count += len(rows)
                        continue
                    for row_number, row in rows:
                        raw_line = row_to_text(row)
                        if not raw_line:
                            continue
                        parsed = parse_inquiry_line(raw_line, source_page=page_index, page_number=page_index, source_row=row_number)
                        if parsed:
                            lines.append(parsed)
            extracted_text = page.extract_text() or ""
            page_metadata.append(
                {
                    "page_number": page_index,
                    "pdfplumber_text_length": len(extracted_text.strip()),
                    "tables_seen": page_tables_seen,
                }
            )
    return lines, page_metadata, table_rows_seen, skipped_count


def _try_ocr_fallback(data, filename):
    provider_name = getattr(settings, "QUOTATION_IMPORT_OCR_PROVIDER", "")
    try:
        provider = get_ocr_provider(provider_name)
        return provider.extract_pdf(data=data, filename=filename)
    except OCRProviderUnavailable as exc:
        return "", str(exc)


def _parse_pdf_word_layout_item_rows(raw_text):
    """Recover compact pipe-delimited item rows from a PDF visual layout."""

    parsed_rows = []
    seen = set()
    for raw_line in str(raw_text or "").splitlines():
        normalized = re.sub(r"\s+", " ", raw_line or "").strip()
        cells = [cell.strip() for cell in normalized.split("|") if cell.strip()]
        if len(cells) < 4:
            continue
        for index, cell in enumerate(cells):
            quantity, unit = split_quantity_unit(cell)
            if quantity is None or not unit or index < 1 or index + 1 >= len(cells):
                continue
            description = cells[index - 1]
            if classify_header_cell(description) or len(re.findall(r"[A-Za-z]", description)) < 4:
                continue
            numeric_tail = [
                candidate
                for candidate in cells[index + 1 :]
                if re.fullmatch(r"\d+(?:[.,]\d+)?", candidate)
            ]
            if not numeric_tail:
                continue
            unit_price = numeric_tail[0]
            line_total = numeric_tail[-1] if len(numeric_tail) > 1 else ""
            reconstructed = " ".join(
                value for value in (description, cell, unit_price, line_total) if value
            )
            parsed = parse_inquiry_line(reconstructed)
            if not parsed:
                continue
            parsed["raw_line"] = normalized
            parsed["raw_source_line"] = normalized
            parsed["parse_confidence"] = max(float(parsed.get("parse_confidence") or 0), 0.92)
            key = (
                parsed.get("requested_item_name"),
                parsed.get("quantity"),
                parsed.get("unit"),
                parsed.get("unit_price"),
            )
            if key not in seen:
                seen.add(key)
                parsed_rows.append(parsed)
            break
    return parsed_rows


def parse_pdf_preview(data, filename, content_type, sha256, *, source_file_ref=""):
    page_count = _preflight_pdf(data)
    warnings = []
    skipped_count = 0
    pymupdf_text_chunks, pymupdf_word_layout_chunks, pymupdf_page_metadata = _extract_pymupdf_text(data)
    selectable_text = "\n".join(chunk for chunk in pymupdf_text_chunks if chunk and chunk.strip()).strip()
    selectable_word_layout = "\n".join(
        chunk for chunk in pymupdf_word_layout_chunks if chunk and chunk.strip()
    ).strip()

    lines = []
    pdfplumber_page_metadata = []
    table_rows_seen = 0
    table_layout_fallback_needed = False
    if selectable_text:
        try:
            lines, pdfplumber_page_metadata, table_rows_seen, table_skipped = _parse_pdfplumber_tables(data)
            skipped_count += table_skipped
            if table_rows_seen and not any(_is_plausible_pdf_item_line(line) for line in lines):
                lines = []
                table_layout_fallback_needed = True
                warnings.append(
                    "Extracted PDF tables contained headers or metadata but no plausible item rows; "
                    "used selectable text layout instead."
                )
        except Exception as exc:
            warnings.append(f"PDF table extraction failed; fell back to text lines. Detail: {exc}")

    parse_method = "pymupdf_pdfplumber_table_v2" if lines else "pymupdf_text_v2"
    if not lines and table_layout_fallback_needed and selectable_word_layout:
        layout_lines = _parse_pdf_word_layout_item_rows(selectable_word_layout)
        layout_skipped = 0
        if not layout_lines:
            layout_lines, layout_skipped = parse_text_lines(selectable_word_layout)
        plausible_layout_lines = [line for line in layout_lines if _is_plausible_pdf_item_line(line)]
        if plausible_layout_lines:
            lines = plausible_layout_lines
            skipped_count += layout_skipped
            parse_method = "pymupdf_word_layout_v1"
    if not lines and selectable_text:
        parsed_lines, skipped = parse_text_lines(selectable_text)
        lines = parsed_lines
        skipped_count += skipped
        if skipped:
            warnings.append(f"Skipped {skipped} likely heading/footer line(s) while parsing PDF text.")

    if not selectable_text:
        ocr_text, ocr_warning = _try_ocr_fallback(data, filename)
        if ocr_text:
            ocr_lines, ocr_skipped = parse_text_lines(ocr_text)
            lines = ocr_lines
            selectable_text = ocr_text
            skipped_count += ocr_skipped
            parse_method = "ocr_deterministic_v2"
        else:
            parse_method = "ocr_required_not_configured_v2"
            warnings.append(
                "No selectable text detected. OCR is not enabled in this environment."
                + (f" {ocr_warning}" if ocr_warning else "")
            )
    elif not lines:
        warnings.append("Selectable text was found, but no item lines were confidently detected. Add rows manually if needed.")

    return _preview_response(
        source_type="pdf",
        source_filename=filename,
        source_mime_type=content_type or PDF_MIME,
        source_sha256=sha256,
        source_file_ref=source_file_ref,
        source_file_size=len(data),
        parse_method=parse_method,
        original_text=selectable_text,
        lines=lines,
        warnings=warnings,
        skipped_count=skipped_count,
        meta={
            "page_count": page_count,
            "page_metadata": pymupdf_page_metadata,
            "pdfplumber_page_metadata": pdfplumber_page_metadata,
            "table_rows_seen": table_rows_seen,
            "source_file_ref": source_file_ref,
            "source_file_size": len(data),
        },
    )
