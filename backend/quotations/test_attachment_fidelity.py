import hashlib
import stat
from io import BytesIO
from unittest.mock import patch
import zipfile

from django.core.exceptions import ValidationError
from django.test import SimpleTestCase, override_settings
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from pypdf import PdfWriter

from .attachment_inspection import (
    inspect_pdf_attachment,
    inspect_spreadsheet_attachment,
)
from .import_parsers import parse_excel_preview, parse_file_preview


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook_bytes(*, item_rows=None, configure=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items"
    sheet.append(["Item", "Qty", "Unit"])
    for row in item_rows or [("Gloves", 5, "box")]:
        sheet.append(list(row))
    if configure:
        configure(workbook, sheet)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _preview(data, *, filename="items.xlsx"):
    return parse_excel_preview(
        data,
        filename,
        XLSX_MIME,
        hashlib.sha256(data).hexdigest(),
        extension=".xlsx",
    )


def _rebuild_xlsx(data, *, replacements=None, additions=None):
    replacements = replacements or {}
    additions = additions or []
    source_buffer = BytesIO(data)
    output_buffer = BytesIO()
    with zipfile.ZipFile(source_buffer, "r") as source, zipfile.ZipFile(
        output_buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as output:
        for info in source.infolist():
            payload = replacements.get(info.filename, source.read(info))
            output.writestr(info, payload)
        for name, payload in additions:
            output.writestr(name, payload)
    return output_buffer.getvalue()


class AttachmentContainerInspectionTests(SimpleTestCase):
    def test_valid_xlsx_with_generic_browser_mime_is_accepted(self):
        data = _workbook_bytes()
        upload = SimpleUploadedFile(
            "items.xlsx",
            data,
            content_type="application/octet-stream",
        )

        preview = parse_file_preview(upload, store_source=False)

        self.assertEqual(preview["lines"][0]["raw_name"], "Gloves")
        self.assertEqual(
            preview["meta"]["attachment_safety"]["validated_format"],
            "xlsx",
        )

    def test_non_excel_zip_renamed_xlsx_is_rejected(self):
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", "<Types />")
            archive.writestr("_rels/.rels", "<Relationships />")
            archive.writestr("word/document.xml", "<document />")

        with self.assertRaisesMessage(ValidationError, "not a valid .xlsx workbook package"):
            inspect_spreadsheet_attachment(buffer.getvalue(), extension=".xlsx")

    def test_duplicate_explicit_safe_directory_entries_are_accepted(self):
        buffer = BytesIO(_workbook_bytes())
        with zipfile.ZipFile(buffer, "a", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("customXml/", b"")
            with self.assertWarns(UserWarning):
                archive.writestr("customXml/", b"")

        report = inspect_spreadsheet_attachment(buffer.getvalue(), extension=".xlsx")

        self.assertEqual(report["safety"]["validated_format"], "xlsx")

    def test_duplicate_file_parts_remain_rejected(self):
        buffer = BytesIO(_workbook_bytes())
        with zipfile.ZipFile(buffer, "a", zipfile.ZIP_DEFLATED) as archive:
            with self.assertWarns(UserWarning):
                archive.writestr("xl/workbook.xml", b"<workbook />")

        with self.assertRaisesMessage(ValidationError, "duplicate archive parts"):
            inspect_spreadsheet_attachment(buffer.getvalue(), extension=".xlsx")

    def test_vba_like_custom_xml_filename_is_warning_only(self):
        data = _rebuild_xlsx(
            _workbook_bytes(),
            additions=[("customXml/vbaGlossary.xml", b"<glossary />")],
        )

        report = inspect_spreadsheet_attachment(data, extension=".xlsx")

        self.assertEqual(report["fidelity"]["macro_part_count"], 0)
        self.assertTrue(
            any("VBA-like filenames" in warning for warning in report["warnings"])
        )

    def test_real_vba_project_part_in_xlsx_is_rejected(self):
        data = _workbook_bytes()
        with zipfile.ZipFile(BytesIO(data), "r") as source:
            content_types = source.read("[Content_Types].xml").replace(
                b"</Types>",
                b'<Override PartName="/xl/vbaProject.bin" '
                b'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
            )
            workbook_relationships = source.read("xl/_rels/workbook.xml.rels").replace(
                b"</Relationships>",
                b'<Relationship Id="rIdVba" Target="vbaProject.bin" '
                b'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject"/>'
                b"</Relationships>",
            )
        data = _rebuild_xlsx(
            data,
            replacements={
                "[Content_Types].xml": content_types,
                "xl/_rels/workbook.xml.rels": workbook_relationships,
            },
            additions=[("xl/vbaProject.bin", b"not-executed")],
        )

        with self.assertRaisesMessage(ValidationError, "Macro-enabled workbook content"):
            inspect_spreadsheet_attachment(data, extension=".xlsx")

    def test_vba_content_type_declaration_is_rejected_even_for_nonstandard_name(self):
        data = _workbook_bytes()
        with zipfile.ZipFile(BytesIO(data), "r") as source:
            content_types = source.read("[Content_Types].xml").replace(
                b"</Types>",
                b'<Override PartName="/xl/customProject.bin" '
                b'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
            )
        data = _rebuild_xlsx(
            data,
            replacements={"[Content_Types].xml": content_types},
            additions=[("xl/customProject.bin", b"not-executed")],
        )

        with self.assertRaisesMessage(ValidationError, "Macro-enabled workbook content"):
            inspect_spreadsheet_attachment(data, extension=".xlsx")

    def test_vba_relationship_is_rejected_even_for_nonstandard_target(self):
        data = _workbook_bytes()
        with zipfile.ZipFile(BytesIO(data), "r") as source:
            workbook_relationships = source.read("xl/_rels/workbook.xml.rels").replace(
                b"</Relationships>",
                b'<Relationship Id="rIdVba" Target="customProject.bin" '
                b'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject"/>'
                b"</Relationships>",
            )
        data = _rebuild_xlsx(
            data,
            replacements={"xl/_rels/workbook.xml.rels": workbook_relationships},
            additions=[("xl/customProject.bin", b"not-executed")],
        )

        with self.assertRaisesMessage(ValidationError, "Macro-enabled workbook content"):
            inspect_spreadsheet_attachment(data, extension=".xlsx")

    def test_strict_ooxml_workbook_content_type_is_accepted(self):
        data = _workbook_bytes()
        with zipfile.ZipFile(BytesIO(data), "r") as source:
            content_types = source.read("[Content_Types].xml").replace(
                b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                b"application/vnd.ms-excel.sheet.main+xml",
            )
        data = _rebuild_xlsx(
            data,
            replacements={"[Content_Types].xml": content_types},
        )

        report = inspect_spreadsheet_attachment(data, extension=".xlsx")

        self.assertEqual(report["safety"]["validated_format"], "xlsx")

    def test_archive_traversal_member_is_rejected(self):
        buffer = BytesIO(_workbook_bytes())
        with zipfile.ZipFile(buffer, "a", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("../outside.xml", "<unsafe />")

        with self.assertRaisesMessage(ValidationError, "unsafe archive member"):
            inspect_spreadsheet_attachment(buffer.getvalue(), extension=".xlsx")

    def test_archive_symlink_member_is_rejected(self):
        buffer = BytesIO(_workbook_bytes())
        with zipfile.ZipFile(buffer, "a", zipfile.ZIP_DEFLATED) as archive:
            link = zipfile.ZipInfo("xl/linked-part.xml")
            link.create_system = 3
            link.external_attr = (stat.S_IFLNK | 0o777) << 16
            archive.writestr(link, "../outside.xml")

        with self.assertRaisesMessage(ValidationError, "symbolic-link archive member"):
            inspect_spreadsheet_attachment(buffer.getvalue(), extension=".xlsx")

    def test_encrypted_zip_flag_is_rejected_before_part_read(self):
        data = bytearray(_workbook_bytes())
        local_header = data.find(b"PK\x03\x04")
        central_header = data.find(b"PK\x01\x02")
        self.assertGreaterEqual(local_header, 0)
        self.assertGreaterEqual(central_header, 0)
        local_flags = int.from_bytes(data[local_header + 6 : local_header + 8], "little")
        central_flags = int.from_bytes(
            data[central_header + 8 : central_header + 10], "little"
        )
        data[local_header + 6 : local_header + 8] = (local_flags | 1).to_bytes(2, "little")
        data[central_header + 8 : central_header + 10] = (
            central_flags | 1
        ).to_bytes(2, "little")

        with self.assertRaisesMessage(ValidationError, "Password-encrypted Excel"):
            inspect_spreadsheet_attachment(bytes(data), extension=".xlsx")

    def test_xlsx_package_is_rejected_when_presented_as_xlsb(self):
        with self.assertRaisesMessage(ValidationError, "not a valid .xlsb workbook package"):
            inspect_spreadsheet_attachment(_workbook_bytes(), extension=".xlsb")

    @override_settings(QUOTATION_IMPORT_MAX_ARCHIVE_ENTRIES=2)
    def test_archive_entry_limit_is_enforced_before_workbook_parse(self):
        with self.assertRaisesMessage(ValidationError, "too many archive parts"):
            inspect_spreadsheet_attachment(_workbook_bytes(), extension=".xlsx")

    @override_settings(QUOTATION_IMPORT_MAX_ARCHIVE_MEMBER_BYTES=32)
    def test_archive_member_expansion_limit_is_enforced(self):
        with self.assertRaisesMessage(ValidationError, "expands beyond the safe limit"):
            inspect_spreadsheet_attachment(_workbook_bytes(), extension=".xlsx")

    @override_settings(
        QUOTATION_IMPORT_MAX_ARCHIVE_MEMBER_BYTES=1024 * 1024,
        QUOTATION_IMPORT_MAX_ARCHIVE_UNCOMPRESSED_BYTES=1024,
    )
    def test_aggregate_archive_expansion_limit_is_enforced(self):
        with self.assertRaisesMessage(ValidationError, "uncompressed-size limit"):
            inspect_spreadsheet_attachment(_workbook_bytes(), extension=".xlsx")

    def test_pdf_active_embedded_and_form_markers_are_warning_only(self):
        buffer = BytesIO()
        writer = PdfWriter()
        writer.add_blank_page(width=100, height=100)
        writer.write(buffer)
        data = buffer.getvalue() + b"\n% /JavaScript /EmbeddedFile /AcroForm\n"

        _, report = inspect_pdf_attachment(data, declared_mime_type="text/plain")

        warning_text = " ".join(report["warnings"])
        self.assertIn("declared content type", warning_text)
        self.assertIn("active-content", warning_text)
        self.assertIn("embedded-file", warning_text)
        self.assertIn("form or XFA", warning_text)

    @patch(
        "quotations.attachment_inspection.MAX_DETAILED_XML_INSPECTION_BYTES",
        2048,
    )
    def test_large_worksheet_detail_scan_is_bounded_and_warns(self):
        data = _workbook_bytes(
            item_rows=[(f"Item {index}", index, "PCS") for index in range(1, 101)]
        )

        report = inspect_spreadsheet_attachment(data, extension=".xlsx")

        self.assertEqual(report["fidelity"]["limited_worksheet_xml_count"], 1)
        self.assertTrue(
            any("could not receive detailed" in warning for warning in report["warnings"])
        )


class SpreadsheetFidelityTests(SimpleTestCase):
    def test_formula_hidden_merge_and_protection_are_visible_warnings(self):
        def configure(workbook, sheet):
            sheet["D1"] = "Budget"
            sheet["D2"] = "=1+1"
            sheet.row_dimensions[3].hidden = True
            sheet.column_dimensions["Z"].hidden = True
            sheet.merge_cells("A4:B4")
            sheet.protection.sheet = True
            hidden = workbook.create_sheet("Archived")
            hidden.sheet_state = "hidden"
            hidden.append(["Item", "Qty", "Unit"])
            hidden.append(["Obsolete item", 99, "box"])

        preview = _preview(_workbook_bytes(configure=configure))
        fidelity = preview["meta"]["spreadsheet_fidelity"]
        warning_text = " ".join(preview["warnings"])

        self.assertEqual([line["raw_name"] for line in preview["lines"]], ["Gloves"])
        self.assertGreaterEqual(fidelity["formula_cell_count"], 1)
        self.assertGreaterEqual(fidelity["formula_without_cached_value_count"], 1)
        self.assertEqual(fidelity["hidden_sheet_count"], 1)
        self.assertGreaterEqual(fidelity["hidden_row_count"], 1)
        self.assertGreaterEqual(fidelity["hidden_column_count"], 1)
        self.assertGreaterEqual(fidelity["merged_range_count"], 1)
        self.assertGreaterEqual(fidelity["protected_sheet_count"], 1)
        self.assertIn("formula cells", warning_text)
        self.assertIn("hidden sheets", warning_text)
        self.assertIn("merged cells", warning_text)
        self.assertIn("protection", warning_text)

    @override_settings(QUOTATION_IMPORT_MAX_EXCEL_SHEETS=1)
    def test_hidden_sheet_before_visible_sheet_does_not_consume_sheet_cap(self):
        def configure(workbook, sheet):
            sheet.title = "Archived"
            sheet.sheet_state = "hidden"
            live = workbook.create_sheet("Live request")
            live.append(["Item", "Qty", "Unit"])
            live.append(["Bandage", 3, "roll"])

        preview = _preview(_workbook_bytes(configure=configure))

        self.assertEqual([line["raw_name"] for line in preview["lines"]], ["Bandage"])
        self.assertEqual(preview["meta"]["inspected_sheets"], ["Live request"])

    @override_settings(QUOTATION_IMPORT_MAX_EXCEL_ROWS=2)
    def test_exact_row_limit_does_not_claim_truncation(self):
        preview = _preview(_workbook_bytes(item_rows=[("Gloves", 5, "box")]))

        self.assertFalse(any("Stopped reading sheet" in warning for warning in preview["warnings"]))

    @override_settings(QUOTATION_IMPORT_MAX_EXCEL_ROWS=2)
    def test_row_beyond_limit_produces_explicit_warning(self):
        preview = _preview(
            _workbook_bytes(
                item_rows=[("Gloves", 5, "box"), ("Bandage", 3, "roll")]
            )
        )

        self.assertEqual([line["raw_name"] for line in preview["lines"]], ["Gloves"])
        self.assertTrue(any("Stopped reading sheet" in warning for warning in preview["warnings"]))

    def test_duplicate_rows_across_selected_sheets_are_retained_and_warned(self):
        def configure(workbook, sheet):
            duplicate = workbook.create_sheet("Second request")
            duplicate.append(["Item", "Qty", "Unit"])
            duplicate.append(["Gloves", 5, "box"])

        preview = _preview(_workbook_bytes(configure=configure))
        warning_text = " ".join(preview["warnings"])

        self.assertEqual(len(preview["lines"]), 2)
        self.assertIn("multiple sheets", warning_text)
        self.assertIn("Potential duplicate", warning_text)

    def test_late_primary_reader_failure_cannot_duplicate_partial_rows(self):
        data = _workbook_bytes()

        def failing_primary(_data):
            yield (
                "Primary",
                [(1, ("Item", "Qty", "Unit")), (2, ("Wrong partial", 1, "box"))],
                "openpyxl_structured_v2",
                {},
            )
            raise OSError("late read failure")

        fallback = [
            (
                "Fallback",
                [(1, ("Item", "Qty", "Unit")), (2, ("Correct row", 2, "box"))],
                "calamine_structured_v2",
                {
                    "row_limit_reached": False,
                    "column_limit_reached": False,
                    "sheet_limit_reached": False,
                    "parser_fallback": False,
                },
            )
        ]
        with patch("quotations.import_parsers._openpyxl_rows", failing_primary), patch(
            "quotations.import_parsers._calamine_rows",
            return_value=iter(fallback),
        ):
            preview = _preview(data)

        self.assertEqual([line["raw_name"] for line in preview["lines"]], ["Correct Row"])
        self.assertTrue(any("fallback reader" in warning for warning in preview["warnings"]))
