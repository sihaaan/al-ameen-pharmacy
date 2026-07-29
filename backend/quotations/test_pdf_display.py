from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

from api.models import Brand, Product

from .excel import build_quotation_excel
from .models import Company, ProformaInvoice, ProformaInvoiceLine, Quotation, QuotationLine
from .pdf import (
    _number,
    _pdf_styles,
    _quotation_line_column_widths,
    _single_line_table_cell,
    build_proforma_invoice_pdf,
    build_quotation_pdf,
    build_standalone_proforma_invoice_pdf,
)


class PdfQuantityFormattingTests(SimpleTestCase):
    def test_number_hides_only_insignificant_decimal_places(self):
        cases = [
            (None, "-"),
            (Decimal("0.000"), "0"),
            (Decimal("50.000"), "50"),
            (Decimal("1000.000"), "1000"),
            (Decimal("1000.125"), "1000.125"),
            (Decimal("1.230"), "1.23"),
            (Decimal("0.125"), "0.125"),
        ]

        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(_number(value), expected)

    def test_single_line_cells_measure_and_shrink_to_their_actual_inner_width(self):
        styles = _pdf_styles(colors.HexColor("#0F766E"))
        cases = [
            ("999999999.999", styles["TableCellQuantity"], 16 * mm - 10, "RIGHT"),
            ("InternationalUnitsPerAmpouleContainerPackaging", styles["TableCellUnit"], 18 * mm - 10, "CENTER"),
        ]

        for value, style, available_width, alignment in cases:
            with self.subTest(value=value):
                cell = _single_line_table_cell(value, style, h_align=alignment)
                width, height = cell.wrap(available_width, 100)

                self.assertEqual(width, available_width)
                self.assertLessEqual(height, style.leading)
                self.assertLess(cell.draw_font_size, style.fontSize)
                self.assertLessEqual(cell.rendered_text_width, available_width + 0.001)


class QuotationPdfDisplayTests(TestCase):
    def pdf_text(self, pdf_bytes):
        reader = PdfReader(BytesIO(pdf_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    def assert_compact_values(self, text):
        self.assertIn("1000", text)
        self.assertIn("1000.125", text)
        self.assertIn("0.125", text)
        self.assertIn("Ampoules", text)
        self.assertIn("Suppositories", text)
        self.assertIn("Each/Ampoule", text)
        self.assertIn("999999999.999", text)
        self.assertIn("InternationalUnitsPerAmpouleContainerPackaging", text)
        self.assertNotIn("1000.000", text)
        self.assertNotIn("1000.00\n0", text)
        self.assertNotIn("1000.\n125", text)
        self.assertNotIn("Ampoule\ns", text)
        self.assertNotIn("Suppositorie\ns", text)
        self.assertNotIn("Each/\nAmpoule", text)
        self.assertNotIn("999999999.\n999", text)
        self.assertNotIn("InternationalUnitsPerAmpoule\nContainerPackaging", text)

    def test_pdf_keeps_compact_quantities_and_common_units_on_one_line(self):
        user = get_user_model().objects.create_user(username="pdf_display_staff", is_staff=True)
        company = Company.objects.create(name="PDF Display Company")
        quotation = Quotation.objects.create(company=company, created_by=user)
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Ammonia Inhalant - Bottle",
            quantity=Decimal("1000.000"),
            unit="Ampoules",
            unit_price=Decimal("1.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=1,
        )
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Measured Liquid",
            quantity=Decimal("0.125"),
            unit="Bottles",
            unit_price=Decimal("8.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=2,
        )
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Specialist Rectal Product",
            quantity=Decimal("1000.125"),
            unit="Suppositories",
            unit_price=Decimal("2.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=3,
        )
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Combined Dispensing Unit",
            quantity=Decimal("12.000"),
            unit="Each/Ampoule",
            unit_price=Decimal("3.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=4,
        )
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Maximum Quantity and Long Unit",
            quantity=Decimal("999999999.999"),
            unit="InternationalUnitsPerAmpouleContainerPackaging",
            unit_price=Decimal("0.01"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=5,
        )

        quotation_text = self.pdf_text(build_quotation_pdf(quotation))
        quotation_proforma_text = self.pdf_text(build_proforma_invoice_pdf(quotation))

        self.assert_compact_values(quotation_text)
        self.assert_compact_values(quotation_proforma_text)

        proforma = ProformaInvoice.objects.create(company=company, created_by=user)
        for index, line in enumerate(quotation.lines.order_by("sort_order", "id")):
            ProformaInvoiceLine.objects.create(
                proforma=proforma,
                item_name=line.item_name_snapshot,
                quantity=line.quantity,
                unit=line.unit,
                unit_price=line.unit_price,
                sort_order=index,
            )
        standalone_text = self.pdf_text(build_standalone_proforma_invoice_pdf(proforma))
        self.assert_compact_values(standalone_text)


class QuotationBrandColumnExportTests(TestCase):
    snapshot_brand = "Original Snapshot Brand"

    def setUp(self):
        self.user = get_user_model().objects.create_user(username="brand_export_staff", is_staff=True)
        self.company = Company.objects.create(name="Brand Export Company")
        self.catalog_brand = Brand.objects.create(name="Current Catalog Brand")
        self.product = Product.objects.create(
            name="Brand Export Product",
            brand=self.catalog_brand,
            price=Decimal("10.00"),
            status="draft",
        )

    def create_quotation(self, *, show_brand_column):
        quotation = Quotation.objects.create(
            company=self.company,
            created_by=self.user,
            show_brand_column=show_brand_column,
            subtotal=Decimal("30.00"),
            total=Decimal("30.00"),
        )
        QuotationLine.objects.create(
            quotation=quotation,
            product=self.product,
            item_name_snapshot="Customer wording retained",
            brand_name_snapshot=self.snapshot_brand,
            quantity=Decimal("2.000"),
            unit="box",
            unit_price=Decimal("10.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=1,
        )
        QuotationLine.objects.create(
            quotation=quotation,
            product=self.product,
            item_name_snapshot="Unbranded commodity",
            brand_name_snapshot="",
            quantity=Decimal("1.000"),
            unit="each",
            unit_price=Decimal("10.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
            sort_order=2,
        )
        return quotation

    @staticmethod
    def pdf_text(pdf_bytes):
        reader = PdfReader(BytesIO(pdf_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    @staticmethod
    def quotation_sheet(quotation):
        workbook = load_workbook(BytesIO(build_quotation_excel(quotation)), data_only=True)
        return workbook["Quotation"]

    @staticmethod
    def header_row(sheet):
        return next(
            row_index
            for row_index in range(1, sheet.max_row + 1)
            if sheet.cell(row=row_index, column=1).value == "S. No."
        )

    def test_brand_column_is_opt_in_and_uses_only_the_saved_snapshot(self):
        quotation = self.create_quotation(show_brand_column=False)

        hidden_quote_text = self.pdf_text(build_quotation_pdf(quotation))
        hidden_proforma_text = self.pdf_text(build_proforma_invoice_pdf(quotation))
        hidden_sheet = self.quotation_sheet(quotation)
        hidden_header_row = self.header_row(hidden_sheet)

        self.assertNotIn(self.snapshot_brand, hidden_quote_text)
        self.assertNotIn(self.snapshot_brand, hidden_proforma_text)
        self.assertEqual(
            [hidden_sheet.cell(hidden_header_row, column).value for column in range(1, 9)],
            ["S. No.", "Item Description", "Qty", "Unit", "Unit Price", "VAT %", "VAT Amount", "Line Total"],
        )
        self.assertEqual(hidden_sheet.max_column, 8)
        self.assertEqual(hidden_sheet.column_dimensions["B"].width, 48)

        quotation.show_brand_column = True
        quotation.save(update_fields=["show_brand_column", "updated_at"])
        self.catalog_brand.name = "Renamed Catalog Brand"
        self.catalog_brand.save(update_fields=["name"])

        shown_quote_text = self.pdf_text(build_quotation_pdf(quotation))
        shown_proforma_text = self.pdf_text(build_proforma_invoice_pdf(quotation))
        shown_sheet = self.quotation_sheet(quotation)
        shown_header_row = self.header_row(shown_sheet)

        for text in (shown_quote_text, shown_proforma_text):
            normalized_text = " ".join(text.split())
            self.assertIn("Brand", text)
            self.assertIn(self.snapshot_brand, normalized_text)
            self.assertNotIn("Renamed Catalog Brand", text)
        self.assertEqual(
            [shown_sheet.cell(shown_header_row, column).value for column in range(1, 10)],
            ["S. No.", "Item Description", "Brand", "Qty", "Unit", "Unit Price", "VAT %", "VAT Amount", "Line Total"],
        )
        self.assertEqual(shown_sheet.cell(shown_header_row + 1, column=3).value, self.snapshot_brand)
        self.assertEqual(shown_sheet.cell(shown_header_row + 2, column=3).value, "-")
        self.assertNotIn("Renamed Catalog Brand", [cell.value for row in shown_sheet.iter_rows() for cell in row])

    def test_brand_enabled_excel_keeps_numeric_formats_and_totals_aligned(self):
        quotation = self.create_quotation(show_brand_column=True)
        sheet = self.quotation_sheet(quotation)
        header_row = self.header_row(sheet)
        first_line_row = header_row + 1

        self.assertEqual(sheet.max_column, 9)
        self.assertEqual(sheet.cell(first_line_row, column=4).value, 2)
        self.assertEqual(sheet.cell(first_line_row, column=4).number_format, "#,##0.000")
        self.assertEqual(sheet.cell(first_line_row, column=6).value, 10)
        self.assertEqual(sheet.cell(first_line_row, column=6).number_format, "#,##0.00#")
        self.assertEqual(sheet.cell(first_line_row, column=7).number_format, "0.00")
        self.assertEqual(sheet.cell(first_line_row, column=8).number_format, "#,##0.00")
        self.assertEqual(sheet.cell(first_line_row, column=9).value, 20)
        self.assertEqual(sheet.cell(first_line_row, column=9).number_format, "#,##0.00")

        grand_total_row = next(
            row_index
            for row_index in range(header_row + 1, sheet.max_row + 1)
            if sheet.cell(row=row_index, column=8).value == "Grand Total"
        )
        self.assertEqual(sheet.cell(grand_total_row, column=9).value, 30)
        self.assertEqual(sheet.cell(grand_total_row, column=9).number_format, "#,##0.00")
        self.assertEqual(sheet.column_dimensions["C"].width, 24)

    def test_long_unicode_brands_fit_a4_and_repeat_on_multipage_documents(self):
        expected_content_width = A4[0] - (32 * mm)
        self.assertLessEqual(sum(_quotation_line_column_widths(True)), expected_content_width + 0.001)

        quotation = Quotation.objects.create(
            company=self.company,
            created_by=self.user,
            show_brand_column=True,
        )
        long_brand = "Müller Santé Ω International Medical Systems and Clinical Consumables"
        for index in range(1, 31):
            QuotationLine.objects.create(
                quotation=quotation,
                product=self.product,
                item_name_snapshot=f"Long customer item description number {index} with specification details",
                brand_name_snapshot=long_brand,
                quantity=Decimal("1000.125"),
                unit="Ampoules",
                unit_price=Decimal("12.375"),
                match_status=QuotationLine.MATCH_CONFIRMED,
                sort_order=index,
            )

        for pdf_bytes in (build_quotation_pdf(quotation), build_proforma_invoice_pdf(quotation)):
            reader = PdfReader(BytesIO(pdf_bytes))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)

            self.assertGreater(len(reader.pages), 1)
            self.assertGreaterEqual(text.count("Brand"), len(reader.pages))
            self.assertIn("Müller Santé", text)
            self.assertIn("Ω", text)
