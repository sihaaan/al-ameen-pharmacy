import base64
import hashlib
import json
import zipfile
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook

from .gmail_inquiry_import import (
    _build_source_analysis,
    _fetch_native_ai_attachment,
)
from .models import (
    GmailInquiryImport,
    GmailOAuthConnection,
    QuotationSettings,
)


MAILBOX_EMAIL = "quotes@example.com"
XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _gmail_data(content):
    return base64.urlsafe_b64encode(content).decode("ascii").rstrip("=")


def _message_with_attachment(message_id, attachment, content):
    private_attachment = {
        **attachment,
        "_inline_data": _gmail_data(content),
    }
    return {
        "gmail_message_id": message_id,
        "gmail_thread_id": "thread-fidelity",
        "label_ids": ["INBOX"],
        "subject": "Request for quotation",
        "sender": "Buyer <buyer@example.com>",
        "recipients": MAILBOX_EMAIL,
        "cc": "",
        "reply_to": "",
        "sent_at": timezone.now(),
        "snippet": "Please quote the attached request.",
        "newest_body_text": "Please quote the attached request.",
        "newest_body_html": "",
        "attachment_manifest": [attachment],
        "_attachment_refs": [private_attachment],
    }


class GmailAttachmentFidelityTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username="gmail-attachment-reviewer",
            is_staff=True,
        )
        self.connection = GmailOAuthConnection.objects.create(
            user=self.staff,
            is_shared=True,
            email=MAILBOX_EMAIL,
            status=GmailOAuthConnection.STATUS_CONNECTED,
        )
        quotation_settings = QuotationSettings.get_solo()
        quotation_settings.ai_pdf_vision_enabled = True
        quotation_settings.save(
            update_fields=["ai_pdf_vision_enabled", "updated_at"]
        )
        self.gmail_import = GmailInquiryImport.objects.create(
            gmail_connection=self.connection,
            mailbox_email=MAILBOX_EMAIL,
            gmail_thread_id="thread-fidelity",
            anchor_message_id="message-fidelity",
            selected_message_ids=["message-fidelity"],
            mode=GmailInquiryImport.MODE_CURRENT_MESSAGE,
            status=GmailInquiryImport.STATUS_CLAIMED,
            claimed_by=self.staff,
            claimed_at=timezone.now(),
        )

    def _fetch_native_xlsx(self, content, *, filename="bounded-request.xlsx"):
        return _fetch_native_ai_attachment(
            self.connection,
            "message-fidelity",
            {
                "filename": filename,
                "mime_type": XLSX_MIME,
                "size": len(content),
                "attachment_id": "bounded-xlsx",
                "part_id": "1",
                "_inline_data": _gmail_data(content),
            },
            max_bytes=5 * 1024 * 1024,
        )

    @override_settings(QUOTATION_MAILBOX_AI_VISION_ENABLED=True)
    @patch("quotations.gmail_inquiry_import._run_native_thread_analysis")
    def test_xlsx_inspection_warnings_propagate_without_changing_provider_bytes(
        self,
        run_native_analysis,
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "RFQ"
        sheet.append(["Item", "Quantity", "Unit", "Budget"])
        sheet.append(["Sterile Gauze", 2, "PCS", "=B2*25"])
        sheet.merge_cells("A4:B4")
        workbook_output = BytesIO()
        workbook.save(workbook_output)
        workbook.close()
        original_bytes = workbook_output.getvalue()
        attachment = {
            "filename": "customer-request.xlsx",
            "mime_type": XLSX_MIME,
            "size": len(original_bytes),
            "attachment_id": "xlsx-fidelity",
            "part_id": "1",
        }
        message = _message_with_attachment(
            "message-fidelity",
            attachment,
            original_bytes,
        )
        captured = {}

        def analysis_result(
            messages,
            evidence,
            file_inputs,
            _gmail_import,
            _actor,
            *,
            analysis_timings=None,
        ):
            captured["file_inputs"] = file_inputs
            attachment_source = next(
                source
                for source in evidence
                if source.get("kind") == "attachment"
            )
            attachment_source["line_count"] = 1
            return {
                "messages": {
                    "message-fidelity": {
                        "classification": "initial_inquiry",
                        "usage": "used",
                        "reason": "Customer request.",
                        "confidence": 0.99,
                    }
                },
                "rows": [
                    {
                        "raw_name": "Sterile Gauze",
                        "raw_line": "Sterile Gauze | 2 | PCS | 50",
                        "raw_source_line": "Sterile Gauze | 2 | PCS | 50",
                        "quantity": "2",
                        "unit": "PCS",
                        "customer_unit_price": "25.00",
                        "customer_line_total": "50.00",
                        "customer_vat": None,
                        "unit_price": "999.00",
                        "vat_rate": "5.00",
                        "vat_amount": "49.95",
                        "line_total": "1048.95",
                        "operation": "added",
                        "parse_status": "parsed",
                        "parse_confidence": 0.99,
                        "semantic_reason": "Read from RFQ sheet.",
                        "included": True,
                        "_source_keys": [attachment_source["source_key"]],
                        "_evidence_row_keys": ["ai:evidence-row"],
                        "evidence": [
                            {
                                "source_key": attachment_source["source_key"],
                                "gmail_message_id": "message-fidelity",
                                "kind": "attachment",
                                "filename": "customer-request.xlsx",
                                "source_sha256": attachment_source[
                                    "source_sha256"
                                ],
                                "evidence_row_key": "ai:evidence-row",
                                "page": "",
                                "sheet_name": "RFQ",
                                "cell_range": "A2:D2",
                                "raw_text": "Sterile Gauze | 2 | PCS | 50",
                            }
                        ],
                    }
                ],
                "warnings": [],
                "thread_summary": "Customer RFQ attachment.",
                "usage": {},
                "customer_identity": {
                    "company_name": "",
                    "contact_name": "",
                    "contact_email": "",
                    "source_keys": [],
                    "confidence": 0.0,
                    "reason": "",
                },
            }

        run_native_analysis.side_effect = analysis_result

        result = _build_source_analysis(
            [message],
            self.connection,
            self.gmail_import,
            self.staff,
            timeline_messages=[message],
        )

        provider_input = captured["file_inputs"][0]
        self.assertEqual(provider_input["content"], original_bytes)
        self.assertEqual(
            provider_input["source_sha256"],
            hashlib.sha256(original_bytes).hexdigest(),
        )
        manifest = result["attachment_manifest"][0]
        self.assertEqual(manifest["parse_status"], "parsed")
        self.assertEqual(manifest["attachment_safety"]["container"], "ooxml_zip")
        self.assertEqual(
            manifest["spreadsheet_fidelity"]["formula_cell_count"],
            1,
        )
        self.assertEqual(
            manifest["spreadsheet_fidelity"]["merged_range_count"],
            1,
        )
        self.assertNotIn(
            "hidden_sheet_names",
            manifest["spreadsheet_fidelity"],
        )
        self.assertTrue(
            any("formula cells" in warning for warning in manifest["warnings"])
        )
        attachment_evidence = next(
            source
            for source in result["evidence"]
            if source.get("kind") == "attachment"
        )
        self.assertEqual(
            attachment_evidence["attachment_safety"],
            manifest["attachment_safety"],
        )
        self.assertEqual(
            attachment_evidence["spreadsheet_fidelity"],
            manifest["spreadsheet_fidelity"],
        )
        self.assertTrue(
            any(
                "customer-request.xlsx" in warning
                and "formula cells" in warning
                for warning in result["preview"]["warnings"]
            )
        )
        quoted_line = result["preview"]["lines"][0]
        self.assertEqual(quoted_line["customer_unit_price"], "25.00")
        self.assertIsNone(quoted_line["unit_price"])
        self.assertIsNone(quoted_line["vat_amount"])
        self.assertIsNone(quoted_line["line_total"])
        json.dumps(result["attachment_manifest"])
        json.dumps(result["evidence"])

    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=1,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=10,
        QUOTATION_AI_NATIVE_MAX_SPREADSHEET_ROWS_PER_SHEET=10,
    )
    def test_native_xlsx_counts_only_visible_sheets_before_submission(self):
        workbook = Workbook()
        visible = workbook.active
        visible.title = "Current RFQ"
        visible.append(["Item", "Quantity"])
        hidden = workbook.create_sheet("Archive")
        hidden.sheet_state = "hidden"
        hidden.append(["Old item", 99])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        native_input, skipped_reason = self._fetch_native_xlsx(
            output.getvalue()
        )

        self.assertEqual(skipped_reason, "")
        self.assertEqual(native_input["content"], output.getvalue())
        self.assertEqual(native_input["spreadsheet_rows"], {"Current RFQ": 1})
        self.assertEqual(
            native_input["spreadsheet_fidelity"]["visible_sheet_count"],
            1,
        )
        self.assertEqual(
            native_input["spreadsheet_fidelity"]["hidden_sheet_count"],
            1,
        )

        workbook = Workbook()
        workbook.active.title = "RFQ 1"
        workbook.create_sheet("RFQ 2")
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        failed_input, failed_reason = self._fetch_native_xlsx(
            output.getvalue()
        )

        self.assertTrue(failed_input["hard_validation_failed"])
        self.assertIn("2 visible sheets", failed_reason)
        self.assertIn("limit is 1", failed_reason)
        self.assertNotIn("content", failed_input)

    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=10,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=2,
        QUOTATION_AI_NATIVE_MAX_SPREADSHEET_ROWS_PER_SHEET=10,
    )
    def test_native_xlsx_rejects_excess_columns_before_submission(self):
        workbook = Workbook()
        workbook.active.append(["Item", "Quantity", "Unit"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        failed_input, failed_reason = self._fetch_native_xlsx(
            output.getvalue()
        )

        self.assertTrue(failed_input["hard_validation_failed"])
        self.assertIn("has 3 columns", failed_reason)
        self.assertIn("limit is 2 columns", failed_reason)
        self.assertNotIn("content", failed_input)

    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=10,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=2,
        QUOTATION_AI_NATIVE_MAX_SPREADSHEET_ROWS_PER_SHEET=10,
    )
    @patch("quotations.gmail_inquiry_import.load_calamine_workbook")
    @patch("quotations.gmail_inquiry_import.inspect_spreadsheet_attachment")
    def test_native_legacy_xls_uses_the_same_column_preflight(
        self,
        inspect_spreadsheet,
        load_workbook,
    ):
        class Metadata:
            name = "Legacy RFQ"
            visible = "visible"

        class Sheet:
            height = 1
            width = 3

        class WorkbookStub:
            sheet_names = ["Legacy RFQ"]
            sheets_metadata = [Metadata()]

            @staticmethod
            def get_sheet_by_name(_sheet_name):
                return Sheet()

            @staticmethod
            def close():
                return None

        inspect_spreadsheet.return_value = {
            "warnings": [],
            "safety": {
                "container": "ole_compound_file",
                "validated_format": "xls",
                "hard_limits_applied": True,
            },
            "fidelity": {"inspection_level": "legacy_binary_limited"},
        }
        load_workbook.return_value = WorkbookStub()
        content = b"legacy-xls-test-bytes"

        failed_input, failed_reason = _fetch_native_ai_attachment(
            self.connection,
            "message-fidelity",
            {
                "filename": "legacy-request.xls",
                "mime_type": "application/vnd.ms-excel",
                "size": len(content),
                "attachment_id": "legacy-xls",
                "part_id": "1",
                "_inline_data": _gmail_data(content),
            },
            max_bytes=5 * 1024 * 1024,
        )

        self.assertTrue(failed_input["hard_validation_failed"])
        self.assertIn("has 3 columns", failed_reason)
        self.assertNotIn("content", failed_input)
        inspect_spreadsheet.assert_called_once()
        load_workbook.assert_called_once()

    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=10,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=10,
        QUOTATION_AI_NATIVE_MAX_SPREADSHEET_ROWS_PER_SHEET=10,
    )
    @patch(
        "quotations.gmail_inquiry_import."
        "HARD_MAX_NATIVE_AI_SPREADSHEET_TOTAL_ROWS",
        3,
    )
    def test_native_xlsx_rejects_excess_aggregate_rows_before_submission(self):
        workbook = Workbook()
        workbook.active.title = "RFQ 1"
        workbook.active.append(["Item 1"])
        workbook.active.append(["Item 2"])
        second = workbook.create_sheet("RFQ 2")
        second.append(["Item 3"])
        second.append(["Item 4"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        failed_input, failed_reason = self._fetch_native_xlsx(
            output.getvalue()
        )

        self.assertTrue(failed_input["hard_validation_failed"])
        self.assertIn("4 aggregate visible rows", failed_reason)
        self.assertIn("limit is 3", failed_reason)
        self.assertNotIn("content", failed_input)

    @override_settings(
        QUOTATION_IMPORT_MAX_EXCEL_SHEETS=10,
        QUOTATION_IMPORT_MAX_EXCEL_COLUMNS=10,
        QUOTATION_AI_NATIVE_MAX_SPREADSHEET_ROWS_PER_SHEET=10,
    )
    @patch(
        "quotations.gmail_inquiry_import."
        "HARD_MAX_NATIVE_AI_SPREADSHEET_TOTAL_CELLS",
        5,
    )
    def test_native_xlsx_rejects_excess_aggregate_cells_before_submission(self):
        workbook = Workbook()
        workbook.active.append(["A", "B", "C"])
        workbook.active.append(["D", "E", "F"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        failed_input, failed_reason = self._fetch_native_xlsx(
            output.getvalue()
        )

        self.assertTrue(failed_input["hard_validation_failed"])
        self.assertIn("6 aggregate visible cells", failed_reason)
        self.assertIn("limit is 5", failed_reason)
        self.assertNotIn("content", failed_input)

    @override_settings(QUOTATION_MAILBOX_AI_VISION_ENABLED=True)
    @patch("quotations.gmail_inquiry_import.get_ai_parse_provider")
    @patch("quotations.gmail_inquiry_import._run_native_thread_analysis")
    def test_unsafe_ooxml_is_failed_evidence_and_prevents_provider_invocation(
        self,
        run_native_analysis,
        get_provider,
    ):
        unsafe_output = BytesIO()
        with zipfile.ZipFile(unsafe_output, "w") as archive:
            archive.writestr("../unsafe.xml", "<unsafe/>")
            archive.writestr("[Content_Types].xml", "<Types/>")
            archive.writestr("_rels/.rels", "<Relationships/>")
            archive.writestr("xl/workbook.xml", "<workbook/>")
        unsafe_bytes = unsafe_output.getvalue()
        attachment = {
            "filename": "unsafe-request.xlsx",
            "mime_type": XLSX_MIME,
            "size": len(unsafe_bytes),
            "attachment_id": "unsafe-xlsx",
            "part_id": "1",
        }
        message = _message_with_attachment(
            "message-fidelity",
            attachment,
            unsafe_bytes,
        )

        result = _build_source_analysis(
            [message],
            self.connection,
            self.gmail_import,
            self.staff,
            timeline_messages=[message],
        )

        run_native_analysis.assert_not_called()
        get_provider.assert_not_called()
        manifest = result["attachment_manifest"][0]
        self.assertEqual(manifest["parse_status"], "failed")
        self.assertIn("unsafe archive member name", manifest["parse_reason"])
        self.assertTrue(manifest["attachment_safety"]["validation_failed"])
        failed_evidence = next(
            source
            for source in result["evidence"]
            if source.get("kind") == "attachment"
        )
        self.assertEqual(failed_evidence["parse_status"], "failed")
        self.assertEqual(
            failed_evidence["parse_reason"],
            manifest["parse_reason"],
        )
        self.assertNotIn("content", failed_evidence)
        self.assertEqual(result["preview"]["lines"], [])
        self.assertFalse(result["preview"]["meta"]["ai_used"])
        self.assertFalse(result["preview"]["meta"]["native_file_ai_used"])
        self.assertEqual(result["preview"]["meta"]["native_file_count"], 0)
        self.assertTrue(
            any(
                "AI analysis was not started" in warning
                for warning in result["preview"]["warnings"]
            )
        )
        json.dumps(result["attachment_manifest"])
        json.dumps(result["evidence"])

    @override_settings(QUOTATION_MAILBOX_AI_VISION_ENABLED=True)
    @patch("quotations.gmail_inquiry_import.get_ai_parse_provider")
    @patch("quotations.gmail_inquiry_import._run_native_thread_analysis")
    def test_one_unsafe_selected_attachment_blocks_analysis_of_valid_siblings(
        self,
        run_native_analysis,
        get_provider,
    ):
        valid_bytes = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Item", "Quantity", "Unit"])
        sheet.append(["Sterile Gauze", 2, "PCS"])
        workbook.save(valid_bytes)
        workbook.close()

        unsafe_output = BytesIO()
        with zipfile.ZipFile(unsafe_output, "w") as archive:
            archive.writestr("../unsafe.xml", "<unsafe/>")
            archive.writestr("[Content_Types].xml", "<Types/>")
            archive.writestr("_rels/.rels", "<Relationships/>")
            archive.writestr("xl/workbook.xml", "<workbook/>")

        attachments = [
            {
                "filename": "valid-request.xlsx",
                "mime_type": XLSX_MIME,
                "size": len(valid_bytes.getvalue()),
                "attachment_id": "valid-xlsx",
                "part_id": "1",
            },
            {
                "filename": "unsafe-request.xlsx",
                "mime_type": XLSX_MIME,
                "size": len(unsafe_output.getvalue()),
                "attachment_id": "unsafe-xlsx",
                "part_id": "2",
            },
        ]
        message = _message_with_attachment(
            "message-fidelity",
            attachments[0],
            valid_bytes.getvalue(),
        )
        message["attachment_manifest"] = attachments
        message["_attachment_refs"] = [
            {
                **attachments[0],
                "_inline_data": _gmail_data(valid_bytes.getvalue()),
            },
            {
                **attachments[1],
                "_inline_data": _gmail_data(unsafe_output.getvalue()),
            },
        ]

        result = _build_source_analysis(
            [message],
            self.connection,
            self.gmail_import,
            self.staff,
            timeline_messages=[message],
        )

        run_native_analysis.assert_not_called()
        get_provider.assert_not_called()
        manifests = {
            entry["filename"]: entry for entry in result["attachment_manifest"]
        }
        self.assertEqual(manifests["valid-request.xlsx"]["parse_status"], "skipped")
        self.assertEqual(manifests["unsafe-request.xlsx"]["parse_status"], "failed")
        self.assertIn(
            "selected attachment failed required safety inspection",
            manifests["valid-request.xlsx"]["parse_reason"],
        )
        self.assertEqual(result["preview"]["lines"], [])
        self.assertFalse(result["preview"]["meta"]["ai_used"])
        self.assertTrue(
            any(
                "AI analysis was not started" in warning
                for warning in result["preview"]["warnings"]
            )
        )
