from decimal import Decimal
from datetime import date
from io import BytesIO
import json
import tempfile
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError
from django.test import override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader, PdfWriter
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import status
from rest_framework.test import APITestCase

from api.models import Product, ProductImage

from .import_parsers import parse_text_preview
from .import_rules import detect_header_row, parse_inquiry_line, parse_text_lines, split_quantity_unit
from .models import (
    AIParseCache,
    AIParseLog,
    Company,
    CompanyContact,
    CompanyPriceHistory,
    HistoricalImportAISuggestion,
    HistoricalImportBatch,
    HistoricalPriceImport,
    HistoricalPriceImportLine,
    Inquiry,
    InquiryLine,
    ProformaInvoice,
    Quotation,
    QuotationAuditLog,
    QuotationLine,
    QuotationLPO,
    QuotationSettings,
    UserQuotationProfile,
    ProductAlias,
    QuoteItem,
)
from .pdf_config import get_quotation_pdf_config
from .matching import apply_match_to_preview_line, suggest_product_for_text
from .ocr import OCRProviderUnavailable, get_ocr_provider
from .services import recalculate_quotation_totals


def make_png_bytes(color=(15, 118, 110, 255)):
    buffer = BytesIO()
    PILImage.new("RGBA", (12, 12), color).save(buffer, format="PNG")
    return buffer.getvalue()


def make_png_upload(name="image.png", color=(15, 118, 110, 255)):
    return SimpleUploadedFile(name, make_png_bytes(color), content_type="image/png")


def extract_pdf_text(content):
    reader = PdfReader(BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


class QuotationPermissionTests(APITestCase):
    list_route_names = [
        "quotation-company-list",
        "quotation-contact-list",
        "quotation-item-list",
        "quotation-inquiry-list",
        "quotation-inquiry-line-list",
        "quotation-historical-import-batch-list",
        "quotation-historical-import-list",
        "quotation-historical-import-line-list",
        "quotation-historical-import-ai-suggestion-list",
        "quotation-list",
        "quotation-lpo-list",
        "quotation-standalone-proforma-list",
        "quotation-line-list",
        "quotation-price-history-list",
        "quotation-audit-log-list",
    ]

    def setUp(self):
        self.company = Company.objects.create(name="Blocked Test Company")
        self.staff = User.objects.create_user(username="staff", password="pass", is_staff=True)
        self.customer = User.objects.create_user(username="customer", password="pass")

    def test_anonymous_users_are_blocked_from_all_list_endpoints(self):
        for route_name in self.list_route_names:
            with self.subTest(route=route_name):
                response = self.client.get(reverse(route_name))
                self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_non_staff_users_are_blocked_from_all_list_endpoints(self):
        self.client.force_authenticate(self.customer)
        for route_name in self.list_route_names:
            with self.subTest(route=route_name):
                response = self.client.get(reverse(route_name))
                self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_staff_users_are_allowed_on_all_list_endpoints(self):
        self.client.force_authenticate(self.staff)
        for route_name in self.list_route_names:
            with self.subTest(route=route_name):
                response = self.client.get(reverse(route_name))
                self.assertEqual(response.status_code, status.HTTP_200_OK)


class QuotationWorkflowTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(username="staff", password="pass", is_staff=True)
        self.customer = User.objects.create_user(username="customer", password="pass")
        self.company = Company.objects.create(name="Workflow Company")
        self.product = Product.objects.create(name="Bandage Pack", price=Decimal("1.00"), pack_size="box", status="draft")
        self.client.force_authenticate(self.staff)

    def create_quote(self):
        return Quotation.objects.create(company=self.company, created_by=self.staff)

    def test_dashboard_endpoint_returns_lightweight_counts(self):
        quote = self.create_quote()
        quote.status = Quotation.STATUS_APPROVED
        quote.save(update_fields=["status", "updated_at"])
        Quotation.objects.create(company=self.company, created_by=self.staff, status=Quotation.STATUS_FINALIZED)
        Inquiry.objects.create(company=self.company, created_by=self.staff, subject="Count test")

        response = self.client.get(reverse("quotation-dashboard"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["companies"], 1)
        self.assertEqual(response.data["items"], 1)
        self.assertEqual(response.data["inquiries"], 1)
        self.assertEqual(response.data["quotes"], 2)
        self.assertEqual(response.data["pending"], 1)
        self.assertEqual(response.data["finalized"], 1)
        self.assertNotIn("lines", response.data)

    def test_quotation_list_is_lightweight_but_detail_includes_lines(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        list_response = self.client.get(reverse("quotation-list"))
        detail_response = self.client.get(reverse("quotation-detail", args=[quotation.id]))

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_response.status_code, status.HTTP_200_OK)
        self.assertNotIn("lines", list_response.data[0])
        self.assertIn("lines", detail_response.data)
        self.assertEqual(len(detail_response.data["lines"]), 1)
        self.assertEqual(list_response.data[0]["created_by_username"], self.staff.username)

    def test_delete_draft_quotation_keeps_audit_snapshot_and_reopens_single_inquiry(self):
        inquiry = Inquiry.objects.create(
            company=self.company,
            created_by=self.staff,
            subject="Disposable supplies",
            status=Inquiry.STATUS_QUOTED,
        )
        quotation = Quotation.objects.create(company=self.company, inquiry=inquiry, created_by=self.staff)
        line = self.create_valid_line(quotation)
        quotation_id = quotation.id
        line_id = line.id

        response = self.client.delete(reverse("quotation-detail", args=[quotation_id]))

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(Quotation.objects.filter(id=quotation_id).exists())
        self.assertFalse(QuotationLine.objects.filter(id=line_id).exists())
        inquiry.refresh_from_db()
        self.assertEqual(inquiry.status, Inquiry.STATUS_DRAFT)

        audit = QuotationAuditLog.objects.get(action=QuotationAuditLog.ACTION_DELETED, target_id=quotation_id)
        snapshot = audit.changes["snapshot"]
        self.assertEqual(snapshot["quotation"]["quotation_number"], quotation.quotation_number)
        self.assertEqual(snapshot["quotation"]["inquiry_id"], inquiry.id)
        self.assertEqual(snapshot["lines"][0]["item_name_snapshot"], "Bandage Pack")
        self.assertEqual(snapshot["lines"][0]["line_total"], "20.00")
        self.assertIsNone(audit.quotation_id)

    def test_delete_draft_quotation_does_not_reopen_inquiry_with_other_quotes(self):
        inquiry = Inquiry.objects.create(
            company=self.company,
            created_by=self.staff,
            subject="Disposable supplies",
            status=Inquiry.STATUS_QUOTED,
        )
        old_quote = Quotation.objects.create(company=self.company, inquiry=inquiry, created_by=self.staff)
        Quotation.objects.create(company=self.company, inquiry=inquiry, created_by=self.staff)

        response = self.client.delete(reverse("quotation-detail", args=[old_quote.id]))

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        inquiry.refresh_from_db()
        self.assertEqual(inquiry.status, Inquiry.STATUS_QUOTED)

    def test_audit_log_filters_can_hide_repetitive_line_activity(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        QuotationAuditLog.objects.create(
            actor=self.staff,
            action=QuotationAuditLog.ACTION_UPDATED,
            target_type="QuotationLine",
            target_id=line.id,
            company=self.company,
            quotation=quotation,
            message="Updated quotation line.",
        )
        QuotationAuditLog.objects.create(
            actor=self.staff,
            action=QuotationAuditLog.ACTION_FINALIZED,
            target_type="Quotation",
            target_id=quotation.id,
            company=self.company,
            quotation=quotation,
            message="Finalized quotation.",
        )

        compact_response = self.client.get(reverse("quotation-audit-log-list"), {"important": "true"})
        full_response = self.client.get(reverse("quotation-audit-log-list"))

        self.assertEqual(compact_response.status_code, status.HTTP_200_OK)
        self.assertEqual(full_response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(compact_response.data), 1)
        self.assertEqual(compact_response.data[0]["action"], QuotationAuditLog.ACTION_FINALIZED)
        self.assertEqual(compact_response.data[0]["action_display"], "Finalized")
        self.assertEqual(len(full_response.data), 2)

    def test_company_list_is_compact_but_detail_includes_contacts(self):
        CompanyContact.objects.create(company=self.company, name="Buyer", email="buyer@example.com")

        list_response = self.client.get(reverse("quotation-company-list"))
        include_response = self.client.get(reverse("quotation-company-list"), {"include_contacts": "true"})
        detail_response = self.client.get(reverse("quotation-company-detail", args=[self.company.id]))

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(include_response.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_response.status_code, status.HTTP_200_OK)
        self.assertNotIn("contacts", list_response.data[0])
        self.assertEqual(list_response.data[0]["contact_count"], 1)
        self.assertIn("contacts", include_response.data[0])
        self.assertIn("contacts", detail_response.data)
        self.assertEqual(detail_response.data["contacts"][0]["name"], "Buyer")

    def test_company_similar_endpoint_strips_legal_suffixes(self):
        intermass = Company.objects.create(name="Intermass")

        response = self.client.get(reverse("quotation-company-similar"), {"name": "Intermass Corp"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["suggestions"][0]["id"], intermass.id)
        self.assertGreaterEqual(response.data["suggestions"][0]["score"], 90)

    def test_creating_high_confidence_similar_company_requires_confirmation(self):
        Company.objects.create(name="Intermass")

        blocked_response = self.client.post(reverse("quotation-company-list"), {"name": "Intermass Corp"}, format="json")
        confirmed_response = self.client.post(
            reverse("quotation-company-list"),
            {"name": "Intermass Corp", "allow_similar": True},
            format="json",
        )

        self.assertEqual(blocked_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(blocked_response.data["requires_confirmation"])
        self.assertEqual(blocked_response.data["similar_companies"][0]["name"], "Intermass")
        self.assertEqual(confirmed_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(confirmed_response.data["name"], "Intermass Corp")

    def test_item_list_uses_compact_product_payload(self):
        self.product.short_description = "Short"
        self.product.detailed_description = "Detailed product copy that is not needed for the list page."
        self.product.save(update_fields=["short_description", "detailed_description"])

        response = self.client.get(reverse("quotation-item-list"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data[0]["id"], self.product.id)
        self.assertIn("slug", response.data[0])
        self.assertIn("price", response.data[0])
        self.assertNotIn("detailed_description", response.data[0])
        self.assertNotIn("stock_quantity", response.data[0])

    def create_valid_line(self, quotation):
        return QuotationLine.objects.create(
            quotation=quotation,
            product=self.product,
            item_name_snapshot="Bandage Pack",
            quantity=Decimal("2.000"),
            unit="box",
            unit_price=Decimal("10.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
        )

    def test_upload_lpo_records_reviewable_lpo_without_changing_outcome(self):
        quotation = self.create_quote()
        quotation.status = Quotation.STATUS_APPROVED
        quotation.save(update_fields=["status", "updated_at"])
        self.create_valid_line(quotation)

        response = self.client.post(
            reverse("quotation-upload-lpo", args=[quotation.id]),
            {
                "text": "Purchase Order No: LPO-77\nDate: 17/06/2026\nBandage Pack 2 box AED 10.00",
                "use_ai": "false",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        lpo = QuotationLPO.objects.get(quotation=quotation)
        self.assertEqual(lpo.lpo_number, "LPO-77")
        self.assertEqual(lpo.lpo_date.isoformat(), "2026-06-17")
        self.assertEqual(lpo.received_by, self.staff)
        quotation.refresh_from_db()
        self.assertEqual(quotation.outcome_status, Quotation.OUTCOME_PENDING)
        self.assertIn("outcome_suggestions", response.data)

    def test_upload_lpo_ignores_po_box_when_detecting_purchase_order_number(self):
        quotation = self.create_quote()
        quotation.status = Quotation.STATUS_APPROVED
        quotation.save(update_fields=["status", "updated_at"])
        self.create_valid_line(quotation)

        response = self.client.post(
            reverse("quotation-upload-lpo", args=[quotation.id]),
            {
                "text": (
                    "SOBHA CONSTRUCTIONS LLC\n"
                    "P.O.BOX - 25654, DUBAI - UNITED ARAB EMIRATES\n"
                    "MATERIAL PURCHASE ORDER\n"
                    "PO No\n"
                    "MPO-104F078-26-0142\n"
                    "PO Date\n"
                    "10-Jun-2026\n"
                    "Purchase Order # : MPO-104F078-26-0142\n"
                    "Sprays Deep Freeze Spray Bot 25.00 14.000 350.00\n"
                ),
                "use_ai": "false",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        lpo = QuotationLPO.objects.get(quotation=quotation)
        self.assertEqual(lpo.lpo_number, "MPO-104F078-26-0142")
        self.assertEqual(lpo.lpo_date.isoformat(), "2026-06-10")

    def test_proforma_pdf_requires_lpo_and_uses_standard_quote_lines(self):
        quotation = self.create_quote()
        quotation.status = Quotation.STATUS_APPROVED
        quotation.save(update_fields=["status", "updated_at"])
        self.create_valid_line(quotation)

        missing_response = self.client.get(reverse("quotation-proforma-pdf", args=[quotation.id]))
        self.assertEqual(missing_response.status_code, status.HTTP_400_BAD_REQUEST)

        QuotationLPO.objects.create(
            quotation=quotation,
            source_type=QuotationLPO.SOURCE_PASTED_TEXT,
            lpo_number="LPO-77",
            lpo_date=date(2026, 6, 17),
            received_by=self.staff,
        )
        response = self.client.get(reverse("quotation-proforma-pdf", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        text = extract_pdf_text(response.content)
        self.assertIn("PROFORMA", text)
        self.assertIn("INVOICE", text)
        self.assertIn("LPO-77", text)
        self.assertIn("Bandage Pack", text)
        self.assertNotIn("Payment Note", text)
        self.assertNotIn("Payment Terms", text)

    def test_standalone_proforma_can_be_created_without_quotation(self):
        create_response = self.client.post(
            reverse("quotation-standalone-proforma-list"),
            {"company": self.company.id, "currency": "AED", "notes": "Advance payment PI"},
            format="json",
        )

        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)
        proforma = ProformaInvoice.objects.get(id=create_response.data["id"])
        self.assertIsNone(proforma.quotation)
        self.assertEqual(proforma.created_by, self.staff)
        self.assertTrue(proforma.proforma_number.startswith("PI-"))

    def test_standalone_proforma_lines_and_pdf_work_without_quotation(self):
        proforma = ProformaInvoice.objects.create(company=self.company, created_by=self.staff)
        update_response = self.client.post(
            reverse("quotation-standalone-proforma-bulk-update-lines", args=[proforma.id]),
            {
                "lines": [
                    {
                        "item_name": "Bandage Pack",
                        "quantity": "2.000",
                        "unit": "box",
                        "unit_price": "10.00",
                        "vat_rate": "5.00",
                    },
                    {
                        "item_name": "Alcohol Swab",
                        "quantity": "3.000",
                        "unit": "pcs",
                        "unit_price": "1.50",
                        "vat_rate": "0.00",
                    },
                ]
            },
            format="json",
        )

        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        proforma.refresh_from_db()
        self.assertEqual(proforma.lines.count(), 2)
        self.assertEqual(proforma.total, Decimal("25.50"))

        pdf_response = self.client.get(reverse("quotation-standalone-proforma-pdf", args=[proforma.id]))

        self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        proforma.refresh_from_db()
        self.assertEqual(proforma.status, ProformaInvoice.STATUS_ISSUED)
        text = extract_pdf_text(pdf_response.content)
        self.assertIn("PROFORMA", text)
        self.assertIn("INVOICE", text)
        self.assertIn("Bandage Pack", text)
        self.assertIn(proforma.proforma_number, text)
        self.assertNotIn("Payment Note", text)
        self.assertNotIn("Payment Terms", text)

    def test_standalone_proforma_upload_lpo_parses_review_lines(self):
        proforma = ProformaInvoice.objects.create(company=self.company, created_by=self.staff)

        response = self.client.post(
            reverse("quotation-standalone-proforma-upload-lpo", args=[proforma.id]),
            {
                "text": (
                    "Purchase Order No: LPO-321\n"
                    "Date: 17/06/2026\n"
                    "Bandage Pack 2 box AED 10.00\n"
                ),
                "use_ai": "false",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        proforma.refresh_from_db()
        self.assertEqual(proforma.lpo_number, "LPO-321")
        self.assertEqual(proforma.lpo_date.isoformat(), "2026-06-17")
        self.assertGreaterEqual(proforma.lines.count(), 1)

    def test_cannot_finalize_invalid_quote(self):
        quotation = self.create_quote()
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Unknown Item",
            quantity=Decimal("1.000"),
            unit="box",
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )

        response = self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        quotation.refresh_from_db()
        self.assertEqual(quotation.status, Quotation.STATUS_DRAFT)
        self.assertEqual(CompanyPriceHistory.objects.count(), 0)

    def test_finalized_quote_appends_price_history_once(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        response = self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        quotation.refresh_from_db()
        self.assertEqual(quotation.status, Quotation.STATUS_FINALIZED)
        self.assertEqual(CompanyPriceHistory.objects.count(), 1)
        history = CompanyPriceHistory.objects.get()
        self.assertEqual(history.company, self.company)
        self.assertEqual(history.product, self.product)
        self.assertEqual(history.unit_price, Decimal("10.00"))

        second_response = self.client.post(reverse("quotation-finalize", args=[quotation.id]))
        self.assertEqual(second_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(CompanyPriceHistory.objects.count(), 1)

    def test_outcome_review_requires_finalized_or_sent_quote(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        response = self.client.get(reverse("quotation-outcome", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_mark_all_accepted_calculates_won_outcome(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        response = self.client.patch(
            reverse("quotation-outcome", args=[quotation.id]),
            {"bulk_action": "mark_all_accepted"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        quotation.refresh_from_db()
        self.assertEqual(line.outcome_status, QuotationLine.OUTCOME_ACCEPTED)
        self.assertEqual(line.accepted_total, line.line_total)
        self.assertEqual(line.lost_value, Decimal("0.00"))
        self.assertEqual(quotation.outcome_status, Quotation.OUTCOME_WON)

    def test_partial_accepted_quantity_calculates_lost_value_and_partial_status(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        response = self.client.patch(
            reverse("quotation-outcome", args=[quotation.id]),
            {
                "line_updates": [
                    {
                        "id": line.id,
                        "outcome_status": QuotationLine.OUTCOME_ACCEPTED,
                        "accepted_quantity": "1.000",
                        "accepted_unit_price": "10.00",
                    }
                ]
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        quotation.refresh_from_db()
        self.assertEqual(line.outcome_status, QuotationLine.OUTCOME_QUANTITY_CHANGED)
        self.assertEqual(line.accepted_total, Decimal("10.00"))
        self.assertEqual(line.lost_value, Decimal("10.00"))
        self.assertEqual(quotation.outcome_status, Quotation.OUTCOME_PARTIAL)

    def test_manual_outcome_override_requires_note_when_different_from_calculated_status(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        response = self.client.patch(
            reverse("quotation-outcome", args=[quotation.id]),
            {"manual_outcome": True, "outcome_status": Quotation.OUTCOME_CANCELLED, "outcome_notes": ""},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_po_outcome_parse_returns_review_suggestions_without_saving_outcomes(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))
        html = """
        <table>
          <tr><th>Item</th><th>Qty</th><th>Unit</th><th>Price</th></tr>
          <tr><td>Bandage Pack</td><td>2</td><td>box</td><td>10</td></tr>
          <tr><td>Extra PO Item</td><td>1</td><td>box</td><td>5</td></tr>
        </table>
        """

        response = self.client.post(
            reverse("quotation-parse-outcome-po", args=[quotation.id]),
            {"text": "Bandage Pack 2 box 10\nExtra PO Item 1 box 5", "html": html, "use_ai": False},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertGreaterEqual(len(response.data["suggestions"]), 1)
        self.assertEqual(response.data["suggestions"][0]["quotation_line_id"], line.id)
        line.refresh_from_db()
        self.assertEqual(line.outcome_status, QuotationLine.OUTCOME_PENDING)
        self.assertTrue(response.data["unmatched_po_rows"])

    def test_analysis_dashboard_and_followups_expose_outcome_metrics(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))
        quotation.status = Quotation.STATUS_SENT
        quotation.next_follow_up_date = date(2026, 1, 1)
        quotation.save(update_fields=["status", "next_follow_up_date", "updated_at"])
        self.client.patch(
            reverse("quotation-outcome", args=[quotation.id]),
            {
                "line_updates": [
                    {
                        "id": line.id,
                        "outcome_status": QuotationLine.OUTCOME_REJECTED,
                        "outcome_reason": QuotationLine.REASON_PRICE_TOO_HIGH,
                    }
                ]
            },
            format="json",
        )
        pending_quote = self.create_quote()
        self.create_valid_line(pending_quote)
        self.client.post(reverse("quotation-finalize", args=[pending_quote.id]))
        pending_quote.status = Quotation.STATUS_SENT
        pending_quote.next_follow_up_date = date(2026, 1, 1)
        pending_quote.save(update_fields=["status", "next_follow_up_date", "updated_at"])

        analysis = self.client.get(reverse("quotation-analysis-dashboard"), {"reason": QuotationLine.REASON_PRICE_TOO_HIGH})
        followups = self.client.get(reverse("quotation-followups"))

        self.assertEqual(analysis.status_code, status.HTTP_200_OK)
        self.assertEqual(analysis.data["cards"]["lost_value"], "20.00")
        self.assertEqual(analysis.data["tables"]["lost_value_by_reason"][0]["reason"], QuotationLine.REASON_PRICE_TOO_HIGH)
        self.assertEqual(followups.status_code, status.HTTP_200_OK)
        self.assertEqual(followups.data["overdue"][0]["id"], pending_quote.id)

    def test_quotation_number_generation_retries_after_collision(self):
        Quotation.objects.create(company=self.company, created_by=self.staff, quotation_number="QT-COLLIDE")

        with patch.object(Quotation, "_generate_quotation_number", side_effect=["QT-COLLIDE", "QT-COLLIDE-2"]):
            quotation = Quotation.objects.create(company=self.company, created_by=self.staff)

        self.assertEqual(quotation.quotation_number, "QT-COLLIDE-2")

    def test_quote_product_price_uses_latest_company_history(self):
        historical_quote = self.create_quote()
        self.create_valid_line(historical_quote)
        finalize_response = self.client.post(reverse("quotation-finalize", args=[historical_quote.id]))
        self.assertEqual(finalize_response.status_code, status.HTTP_200_OK)

        current_quote = self.create_quote()
        response = self.client.get(
            reverse("quotation-product-price", args=[current_quote.id]),
            {"product": self.product.id},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["source"], "company_price_history")
        self.assertEqual(response.data["unit_price"], "10.00")
        self.assertEqual(response.data["unit"], "box")

    def test_quote_product_price_does_not_fall_back_to_product_base_price(self):
        current_quote = self.create_quote()
        response = self.client.get(
            reverse("quotation-product-price", args=[current_quote.id]),
            {"product": self.product.id},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["source"], "no_company_price_history")
        self.assertEqual(response.data["unit_price"], "")

    def test_company_used_items_filter_only_returns_company_products(self):
        historical_quote = self.create_quote()
        self.create_valid_line(historical_quote)
        finalize_response = self.client.post(reverse("quotation-finalize", args=[historical_quote.id]))
        self.assertEqual(finalize_response.status_code, status.HTTP_200_OK)
        alias_product = Product.objects.create(name="Alias Product", price=Decimal("2.00"), status="draft")
        ProductAlias.objects.create(company=self.company, product=alias_product, alias="Company alias", created_by=self.staff)
        other_company = Company.objects.create(name="Other Company")
        other_product = Product.objects.create(name="Other Product", price=Decimal("3.00"), status="draft")
        other_quote = Quotation.objects.create(company=other_company, created_by=self.staff)
        other_line = QuotationLine.objects.create(
            quotation=other_quote,
            product=other_product,
            item_name_snapshot="Other Product",
            quantity=Decimal("1.000"),
            unit="box",
            unit_price=Decimal("3.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
        )
        finalize_other = self.client.post(reverse("quotation-finalize", args=[other_quote.id]))
        self.assertEqual(finalize_other.status_code, status.HTTP_200_OK)

        response = self.client.get(reverse("quotation-item-list"), {"active": "true", "company_used": self.company.id})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        product_ids = {row["id"] for row in response.data}
        self.assertIn(self.product.id, product_ids)
        self.assertIn(alias_product.id, product_ids)
        self.assertNotIn(other_product.id, product_ids)

    def test_download_filenames_start_with_company_name(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        pdf_response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        excel_response = self.client.get(reverse("quotation-excel", args=[quotation.id]))

        self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)
        self.assertIn(f'filename="WORKFLOW_COMPANY-{quotation.quotation_number}.pdf"', pdf_response["Content-Disposition"])
        self.assertEqual(excel_response.status_code, status.HTTP_200_OK)
        self.assertIn(f'filename="WORKFLOW_COMPANY-{quotation.quotation_number}.xlsx"', excel_response["Content-Disposition"])

    def test_quotation_detail_includes_purchaser_contact_fields(self):
        contact = CompanyContact.objects.create(
            company=self.company,
            name="Ahmed Khan",
            role="Purchase Officer",
            department="Clinic Supplies",
            phone="+971501234567",
            email="ahmed@example.com",
        )
        quotation = Quotation.objects.create(company=self.company, contact=contact, created_by=self.staff)

        response = self.client.get(reverse("quotation-detail", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["contact_name"], "Ahmed Khan")
        self.assertEqual(response.data["contact_role"], "Purchase Officer")
        self.assertEqual(response.data["contact_department"], "Clinic Supplies")
        self.assertEqual(response.data["contact_phone"], "+971501234567")
        self.assertEqual(response.data["contact_email"], "ahmed@example.com")

    def test_finalized_quote_cannot_be_edited(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        quote_response = self.client.patch(
            reverse("quotation-detail", args=[quotation.id]),
            {"notes": "Should fail"},
            format="json",
        )
        line_response = self.client.patch(
            reverse("quotation-line-detail", args=[line.id]),
            {"unit_price": "12.00"},
            format="json",
        )

        self.assertEqual(quote_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(line_response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_revision_creates_new_draft(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)
        self.client.post(reverse("quotation-finalize", args=[quotation.id]))

        response = self.client.post(reverse("quotation-revise", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        quotation.refresh_from_db()
        self.assertEqual(quotation.status, Quotation.STATUS_REVISED)
        revision = Quotation.objects.get(id=response.data["id"])
        self.assertEqual(revision.status, Quotation.STATUS_DRAFT)
        self.assertEqual(revision.version, 2)
        self.assertEqual(revision.parent, quotation)
        self.assertEqual(revision.lines.count(), 1)

    def test_create_quote_from_inquiry_is_idempotent(self):
        inquiry = Inquiry.objects.create(company=self.company, subject="Repeat inquiry", created_by=self.staff)
        InquiryLine.objects.create(
            inquiry=inquiry,
            raw_name="Bandage Pack",
            matched_product=self.product,
            quantity=Decimal("2.000"),
            unit="box",
            match_status=InquiryLine.MATCH_CONFIRMED,
        )

        first_response = self.client.post(reverse("quotation-inquiry-create-quote", args=[inquiry.id]))
        second_response = self.client.post(reverse("quotation-inquiry-create-quote", args=[inquiry.id]))

        self.assertEqual(first_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(second_response.status_code, status.HTTP_200_OK)
        self.assertEqual(first_response.data["id"], second_response.data["id"])
        self.assertEqual(Quotation.objects.filter(inquiry=inquiry).count(), 1)

    def test_price_reference_workbook_fills_inquiry_preview_prices(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Sl No", "Items", "unit", "qty", "uprice", "TOTAL", "Vat", "g total"])
        sheet.append([1, "Adol Infant suspension", "pcs", 3, 12, 36, 0, 36])
        sheet.append([2, "Fenestil Gel", "pcs", 3, 12.5, 37.5, 0, 37.5])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "paddington.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        preview = {
            "source_type": "excel",
            "lines": [
                {"raw_name": "Adol Infant suspension", "quantity": "3", "unit": "pcs", "parse_status": "parsed"},
                {"raw_name": "Unknown Clinic Item", "quantity": "1", "unit": "pcs", "parse_status": "needs_review"},
            ],
        }

        response = self.client.post(
            reverse("quotation-inquiry-apply-price-reference"),
            {"file": upload, "preview": json.dumps(preview)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["price_reference_summary"]["matched_count"], 1)
        self.assertEqual(response.data["lines"][0]["unit_price"], "12.00")
        self.assertEqual(response.data["lines"][0]["vat_rate"], "0.00")
        self.assertEqual(response.data["lines"][0]["price_reference_status"], "matched")
        self.assertEqual(response.data["lines"][1]["price_reference_status"], "unmatched")

    def test_price_reference_pasted_text_fills_prices_and_preserves_quantity(self):
        preview = {
            "source_type": "pasted_text",
            "lines": [
                {"raw_name": "Pulse Oximeter", "quantity": "2", "unit": "", "parse_status": "parsed"},
                {"raw_name": "Unknown Clinic Item", "quantity": "5", "unit": "PCS", "parse_status": "parsed"},
            ],
        }
        raw_text = "\n".join(
            [
                "S. No.\tItem Description\tQty\tUOM\tPrice",
                "1\tPulse Oximeter\t1\tNUM\t55",
                "2\tGlucometer\t1\tNUM\t145",
            ]
        )

        response = self.client.post(
            reverse("quotation-inquiry-apply-price-reference"),
            {"raw_text": raw_text, "preview": json.dumps(preview), "use_ai": "false"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["price_reference_summary"]["matched_count"], 1)
        self.assertEqual(response.data["lines"][0]["quantity"], "2")
        self.assertEqual(response.data["lines"][0]["unit"], "")
        self.assertEqual(response.data["lines"][0]["unit_price"], "55.00")
        self.assertEqual(response.data["lines"][1]["price_reference_status"], "unmatched")

    def test_imported_inquiry_prices_carry_into_created_quote_lines(self):
        response = self.client.post(
            reverse("quotation-inquiry-create-imported"),
            {
                "company": self.company.id,
                "subject": "Clinic supplies with workbook prices",
                "source_type": Inquiry.SOURCE_TYPE_EXCEL,
                "lines": [
                    {
                        "raw_name": "Fenestil Gel",
                        "quantity": "3",
                        "unit": "pcs",
                        "unit_price": "12.50",
                        "vat_rate": "0",
                        "parse_status": InquiryLine.PARSE_PARSED,
                        "parse_confidence": 0.95,
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        inquiry = Inquiry.objects.get(pk=response.data["id"])

        quote_response = self.client.post(reverse("quotation-inquiry-create-quote", args=[inquiry.id]))

        self.assertEqual(quote_response.status_code, status.HTTP_201_CREATED)
        line = QuotationLine.objects.get(quotation_id=quote_response.data["id"])
        self.assertEqual(line.unit_price, Decimal("12.50"))
        self.assertEqual(line.vat_rate, Decimal("0.00"))
        self.assertEqual(line.line_total, Decimal("37.50"))

    def test_pdf_endpoint_is_staff_only(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        self.client.force_authenticate(self.customer)
        blocked = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        self.assertEqual(blocked.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        allowed = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        self.assertEqual(allowed.status_code, status.HTTP_200_OK)
        self.assertEqual(allowed["Content-Type"], "application/pdf")

    def test_excel_endpoint_generates_workbook_for_staff(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        self.client.force_authenticate(self.customer)
        blocked = self.client.get(reverse("quotation-excel", args=[quotation.id]))
        self.assertEqual(blocked.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        response = self.client.get(reverse("quotation-excel", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook["Quotation"]
        self.assertEqual(sheet["A1"].value, "Quotation Export")
        header_row = next(
            row_index
            for row_index in range(1, sheet.max_row + 1)
            if sheet.cell(row=row_index, column=1).value == "S. No."
        )
        self.assertEqual(sheet.cell(row=header_row, column=2).value, "Item Description")
        self.assertEqual(sheet.cell(row=header_row + 1, column=2).value, "Bandage Pack")
        self.assertEqual(sheet.cell(row=header_row + 1, column=8).value, 20)
        self.assertEqual(sheet.cell(row=header_row + 5, column=7).value, "Grand Total")
        self.assertEqual(len(sheet.merged_cells.ranges), 0)
        self.assertIsNone(sheet.freeze_panes)
        self.assertIsNone(sheet.auto_filter.ref)

    def test_ignored_lines_are_excluded_from_totals_pdf_and_excel(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)
        QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="Internal skipped item",
            quantity=Decimal("3.000"),
            unit="box",
            unit_price=Decimal("99.00"),
            vat_rate=Decimal("5.00"),
            match_status=QuotationLine.MATCH_IGNORED,
        )

        recalculate_quotation_totals(quotation)
        quotation.refresh_from_db()

        self.assertEqual(quotation.subtotal, Decimal("20.00"))
        self.assertEqual(quotation.total, Decimal("20.00"))

        pdf_response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)
        pdf_text = extract_pdf_text(pdf_response.content)
        self.assertIn("Bandage Pack", pdf_text)
        self.assertNotIn("Internal skipped item", pdf_text)

        excel_response = self.client.get(reverse("quotation-excel", args=[quotation.id]))
        self.assertEqual(excel_response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
        values = [cell.value for row in workbook["Quotation"].iter_rows() for cell in row]
        self.assertIn("Bandage Pack", values)
        self.assertNotIn("Internal skipped item", values)

    def test_quote_payment_terms_can_be_selected_and_exported(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        self.assertEqual(quotation.payment_terms, Quotation.PAYMENT_AS_PER_AGREEMENT)

        update_response = self.client.patch(
            reverse("quotation-detail", args=[quotation.id]),
            {"payment_terms": Quotation.PAYMENT_PDC_60},
            format="json",
        )
        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        self.assertEqual(update_response.data["payment_terms"], Quotation.PAYMENT_PDC_60)
        self.assertEqual(update_response.data["payment_terms_display"], "PDC 60 days")

        pdf_response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        self.assertIn("PDC 60 days", extract_pdf_text(pdf_response.content))

        excel_response = self.client.get(reverse("quotation-excel", args=[quotation.id]))
        workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
        values = [str(cell.value) for row in workbook["Quotation"].iter_rows() for cell in row if cell.value is not None]
        self.assertTrue(any("PDC 60 days" in value for value in values))

    def test_as_per_agreement_payment_term_has_professional_pdf_wording(self):
        quotation = self.create_quote()
        self.create_valid_line(quotation)

        update_response = self.client.patch(
            reverse("quotation-detail", args=[quotation.id]),
            {"payment_terms": Quotation.PAYMENT_AS_PER_AGREEMENT},
            format="json",
        )
        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        self.assertEqual(update_response.data["payment_terms_display"], "As per agreement")

        pdf_response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        pdf_text = extract_pdf_text(pdf_response.content)
        self.assertIn("As per mutually agreed terms.", pdf_text)
        self.assertNotIn("Payment Terms: As per agreement", pdf_text)

    def test_unmatched_quotation_line_can_create_internal_product(self):
        quotation = self.create_quote()
        line = QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="PULSE OXMETER",
            quantity=Decimal("2.000"),
            unit="NUM",
            unit_price=Decimal("35.00"),
            vat_rate=Decimal("5.00"),
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )

        response = self.client.post(reverse("quotation-line-create-product", args=[line.id]), {"product_name": "Pulse Oxmeter"}, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        line.refresh_from_db()
        product = Product.objects.get(name="Pulse Oxmeter")
        self.assertEqual(line.product, product)
        self.assertEqual(line.match_status, QuotationLine.MATCH_CONFIRMED)
        self.assertEqual(product.status, "draft")
        self.assertFalse(product.show_price)

    def test_quotation_line_can_upload_and_include_product_image(self):
        quotation = self.create_quote()
        line = self.create_valid_line(quotation)
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                response = self.client.post(
                    reverse("quotation-line-upload-product-image", args=[line.id]),
                    {"image": make_png_upload("bandage.png")},
                    format="multipart",
                )

                self.assertEqual(response.status_code, status.HTTP_201_CREATED)
                line.refresh_from_db()
                self.assertTrue(line.include_product_image)
                self.assertIsNotNone(line.product_image)
                self.assertEqual(line.product_image.product, self.product)
                self.assertEqual(ProductImage.objects.filter(product=self.product).count(), 1)

                pdf_response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
                self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)

    def test_bulk_create_products_dedupes_same_normalized_names(self):
        quotation = self.create_quote()
        first = QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="ENO",
            quantity=Decimal("1.000"),
            unit="PKT",
            unit_price=Decimal("3.00"),
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )
        second = QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot=" eno ",
            quantity=Decimal("5.000"),
            unit="PKT",
            unit_price=Decimal("3.00"),
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )

        response = self.client.post(
            reverse("quotation-bulk-create-products-for-lines", args=[quotation.id]),
            {"line_ids": [first.id, second.id], "names": {str(first.id): "ENO", str(second.id): "ENO"}},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(Product.objects.filter(name__iexact="ENO").count(), 1)
        product = Product.objects.get(name__iexact="ENO")
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.product, product)
        self.assertEqual(second.product, product)
        self.assertEqual(response.data["unique_products"], 1)

    def test_create_product_reuses_same_slug_product_instead_of_duplicate(self):
        quotation = self.create_quote()
        existing = Product.objects.create(name="Alcohol Detector Mouth-Piece", price=Decimal("1.00"), status="draft")
        line = QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="ALCOHOL DETECTOR MOUTH PIECE",
            quantity=Decimal("50.000"),
            unit="NOS",
            unit_price=Decimal("2.75"),
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )

        response = self.client.post(reverse("quotation-line-create-product", args=[line.id]), format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        self.assertEqual(line.product, existing)
        self.assertEqual(Product.objects.filter(name__icontains="Alcohol Detector").count(), 1)

    def test_bulk_create_products_returns_json_error_for_invalid_selection(self):
        quotation = self.create_quote()

        response = self.client.post(
            reverse("quotation-bulk-create-products-for-lines", args=[quotation.id]),
            {"line_ids": [999999]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)

    def test_save_all_lines_and_finalize_with_created_product_and_skipped_row(self):
        quotation = self.create_quote()
        active = QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="GLUCOMETER",
            quantity=Decimal("1.000"),
            unit="NUM",
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )
        skipped = QuotationLine.objects.create(
            quotation=quotation,
            item_name_snapshot="IGNORE ME",
            quantity=Decimal("1.000"),
            unit="NUM",
            match_status=QuotationLine.MATCH_UNRESOLVED,
        )
        create_response = self.client.post(reverse("quotation-line-create-product", args=[active.id]), {"product_name": "Glucometer"}, format="json")
        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)

        save_response = self.client.post(
            reverse("quotation-bulk-update-lines", args=[quotation.id]),
            {
                "lines": [
                    {"id": active.id, "unit_price": "25.00", "vat_rate": "5", "quantity": "2", "unit": "NUM"},
                    {"id": skipped.id, "match_status": QuotationLine.MATCH_IGNORED, "unit_price": "", "vat_rate": "0", "quantity": "1", "unit": "NUM"},
                ]
            },
            format="json",
        )

        self.assertEqual(save_response.status_code, status.HTTP_200_OK)
        active.refresh_from_db()
        skipped.refresh_from_db()
        self.assertEqual(active.vat_rate, Decimal("5"))
        self.assertEqual(active.line_total, Decimal("52.50"))
        self.assertEqual(skipped.match_status, QuotationLine.MATCH_IGNORED)

        finalize_response = self.client.post(reverse("quotation-finalize", args=[quotation.id]))
        self.assertEqual(finalize_response.status_code, status.HTTP_200_OK)
        self.assertEqual(CompanyPriceHistory.objects.count(), 1)


class ProductCatalogMatchingTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(username="product_staff", password="pass", is_staff=True)
        self.company_a = Company.objects.create(name="Company A")
        self.company_b = Company.objects.create(name="Company B")
        self.product_a = Product.objects.create(name="Plastic Band-Aids", price=Decimal("1.00"), status="draft")
        self.product_b = Product.objects.create(name="Thin Band-Aids", price=Decimal("1.00"), status="draft")

    def test_company_alias_overrides_same_global_or_other_company_alias(self):
        ProductAlias.objects.create(company=self.company_a, product=self.product_a, alias="band aids", created_by=self.staff)
        ProductAlias.objects.create(company=self.company_b, product=self.product_b, alias="band aids", created_by=self.staff)

        match_a = suggest_product_for_text("band aids", self.company_a)
        match_b = suggest_product_for_text("band aids", self.company_b)

        self.assertEqual(match_a.product, self.product_a)
        self.assertEqual(match_a.method, "company_alias")
        self.assertEqual(match_b.product, self.product_b)
        self.assertEqual(match_b.method, "company_alias")

    def test_global_alias_is_used_when_company_alias_is_missing(self):
        ProductAlias.objects.create(product=self.product_a, alias="plasters", created_by=self.staff)

        match = suggest_product_for_text("plasters", self.company_b)

        self.assertEqual(match.product, self.product_a)
        self.assertEqual(match.method, "global_alias")

    def test_company_safe_preview_matching_requires_alias_or_company_history(self):
        exact_catalog_product = Product.objects.create(name="Pulse Oxmeter", price=Decimal("1.00"), status="draft")
        line = {"raw_name": "Pulse Oxmeter"}

        apply_match_to_preview_line(line, self.company_a)

        self.assertIsNone(line["matched_product"])
        self.assertEqual(line["match_status"], "unresolved")

        quotation = Quotation.objects.create(company=self.company_a, created_by=self.staff)
        quotation_line = QuotationLine.objects.create(
            quotation=quotation,
            product=exact_catalog_product,
            item_name_snapshot="Pulse Oxmeter",
            quantity=Decimal("1.000"),
            unit="No",
            unit_price=Decimal("10.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
        )
        CompanyPriceHistory.objects.create(
            company=self.company_a,
            product=exact_catalog_product,
            quotation=quotation,
            quotation_line=quotation_line,
            unit_price=Decimal("10.00"),
            unit="No",
            created_by=self.staff,
        )

        matched_line = {"raw_name": "Pulse Oxmeter"}
        apply_match_to_preview_line(matched_line, self.company_a)

        self.assertEqual(matched_line["matched_product"], exact_catalog_product.id)
        self.assertEqual(matched_line["match_status"], "confirmed")
        self.assertEqual(matched_line["match_method"], "company_price_history")

    def test_public_product_list_hides_internal_draft_products_from_customers(self):
        Product.objects.create(name="Public Item", price=Decimal("2.00"), status="active")

        public_response = self.client.get(reverse("product-list"))

        self.assertEqual(public_response.status_code, status.HTTP_200_OK)
        names = [row["name"] for row in public_response.data]
        self.assertIn("Public Item", names)
        self.assertNotIn("Plastic Band-Aids", names)

    def test_staff_can_create_internal_product_through_quotation_items_api(self):
        self.client.force_authenticate(self.staff)

        response = self.client.post(
            reverse("quotation-item-list"),
            {"name": "Internal Gauze Roll", "pack_size": "roll"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        product = Product.objects.get(name="Internal Gauze Roll")
        self.assertEqual(product.status, "draft")
        self.assertFalse(product.show_price)


class InquiryParserRuleTests(APITestCase):
    def test_header_detection_skips_title_and_maps_first_aid_columns(self):
        rows = [
            ("REQUEST FOR FIRST AID ITEMS", None, None, None),
            ("SL NO", "ITEMS", "UNIT", "QUANTITY"),
            (1, "ANTI SEPTIC SOLUTION", None, "1 bottle"),
            (2, "GAUZE PIECES", None, "1 BOX"),
        ]

        header = detect_header_row(rows)

        self.assertIsNotNone(header)
        self.assertEqual(header.row_number, 2)
        self.assertEqual(header.columns["serial_no"], 0)
        self.assertEqual(header.columns["requested_item_name"], 1)
        self.assertEqual(header.columns["unit"], 2)
        self.assertEqual(header.columns["quantity"], 3)

    def test_quantity_unit_splitting(self):
        quantity, unit = split_quantity_unit("5 Nos", "")

        self.assertEqual(str(quantity), "5")
        self.assertEqual(unit.lower(), "nos")

    def test_serial_prefix_is_not_part_of_item_name(self):
        parsed = parse_inquiry_line("01 - TRIANGULAR BANDAGES 5 Nos")

        self.assertEqual(parsed["raw_name"], "TRIANGULAR BANDAGES")
        self.assertEqual(parsed["quantity"], "5")
        self.assertEqual(parsed["unit"].lower(), "nos")
        self.assertEqual(parsed["serial_no"], "01")

    def test_repeated_header_row_skipped_in_text_parser(self):
        lines, skipped = parse_text_lines("SL NO | ITEMS | UNIT | QUANTITY\n1 | GAUZE PIECES | | 1 box")

        self.assertEqual(skipped, 1)
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]["raw_name"], "GAUZE PIECES")

    def test_gmail_cell_per_line_table_parses_item_rows(self):
        pasted = "\n".join(
            [
                "S. No.",
                "Item Description",
                "Qty",
                "UOM",
                "Price",
                "1",
                "PULSE OXMETER",
                "2",
                "NUM",
                "2",
                "GLUCOMETER",
                "1",
                "NUM",
                "3",
                "SURGICAL SCISSOR",
                "2",
                "NUM",
                "4",
                "BP MACHINE",
                "1",
                "NUM",
                "5",
                "STERILE MOUND DRESSING",
                "50",
                "NUM",
            ]
        )

        lines, skipped = parse_text_lines(pasted)

        self.assertEqual([line["raw_name"] for line in lines[:5]], [
            "PULSE OXMETER",
            "GLUCOMETER",
            "SURGICAL SCISSOR",
            "BP MACHINE",
            "STERILE MOUND DRESSING",
        ])
        self.assertEqual(lines[0]["quantity"], "2")
        self.assertEqual(lines[0]["unit"], "NUM")
        self.assertNotIn("Item Description", [line["raw_name"] for line in lines])

    def test_gmail_html_table_parses_rows_without_headers(self):
        html = """
        <table>
          <tr><th>S. No.</th><th>Item Description</th><th>Qty</th><th>UOM</th><th>Price</th></tr>
          <tr><td>1</td><td>PULSE OXMETER</td><td>2</td><td>NUM</td><td></td></tr>
          <tr><td>2</td><td>GLUCOMETER</td><td>1</td><td>NUM</td><td></td></tr>
        </table>
        """

        preview = parse_text_preview("PULSE OXMETER\nGLUCOMETER", raw_html=html)

        self.assertEqual(preview["parse_method"], "deterministic_clipboard_html_table_v1")
        self.assertEqual(preview["lines"][0]["raw_name"], "PULSE OXMETER")
        self.assertEqual(preview["lines"][0]["quantity"], "2")
        self.assertEqual(preview["lines"][0]["unit"], "NUM")
        self.assertEqual(preview["lines"][1]["raw_name"], "GLUCOMETER")

    def test_ocr_provider_interface_is_explicitly_unavailable_by_default(self):
        with self.assertRaises(OCRProviderUnavailable):
            get_ocr_provider("")


class InquiryImportTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(username="import_staff", password="pass", is_staff=True)
        self.customer = User.objects.create_user(username="import_customer", password="pass")
        self.company = Company.objects.create(name="Import Company")
        self.contact = CompanyContact.objects.create(company=self.company, name="Buyer")
        self.client.force_authenticate(self.staff)

    def make_excel_upload(self, name="inquiry.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LPO"
        sheet.append(["Item", "Qty", "Unit"])
        sheet.append(["Panadol 500mg", 10, "box"])
        sheet.append(["Gloves medium", 5, "packs"])
        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            name,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def make_first_aid_excel_upload(self, name="FIRST AID MATERIAL LOG.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "First Aid"
        sheet.append(["REQUEST FOR FIRST AID ITEMS", None, None, None])
        sheet.append(["SL NO", "ITEMS", "UNIT", "QUANTITY"])
        sheet.append([1, "ANTI SEPTIC SOLUTION", None, "1 bottle"])
        sheet.append([2, "GAUZE PIECES", None, "1 BOX"])
        sheet.append([3, "TRIANGULAR BANDAGES", None, "5 Nos"])
        sheet.append([4, "SPLINTS", None, "2 packs"])
        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            name,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def make_multi_sheet_excel_upload(self, name="multi.xlsx"):
        workbook = Workbook()
        title_sheet = workbook.active
        title_sheet.title = "Cover"
        title_sheet.append(["REQUEST FOR FIRST AID ITEMS"])
        data_sheet = workbook.create_sheet("Items")
        data_sheet.append(["SL NO", "ITEMS", "UNIT", "QUANTITY"])
        data_sheet.append([1, "GLOVES MEDIUM", "", "5 packs"])
        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            name,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def make_pdf_upload(self, text=True, encrypted=False, name="inquiry.pdf"):
        buffer = BytesIO()
        if text:
            pdf = canvas.Canvas(buffer)
            pdf.drawString(72, 740, "Panadol 500mg - 10 boxes")
            pdf.drawString(72, 720, "Gloves medium 5 packs")
            pdf.save()
            data = buffer.getvalue()
        else:
            writer = PdfWriter()
            writer.add_blank_page(width=300, height=300)
            if encrypted:
                writer.encrypt("secret")
            writer.write(buffer)
            data = buffer.getvalue()
        return SimpleUploadedFile(name, data, content_type="application/pdf")

    def make_material_description_pdf_upload(self, name="244047.pdf"):
        buffer = BytesIO()
        styles = getSampleStyleSheet()
        document = SimpleDocTemplate(buffer, pagesize=A4)
        metadata_table = Table(
            [
                ["DATE: 023/05/2026", ""],
                [
                    "From(Seller):\nAl Ameen Pharmacy LLC\nE-Mail: pharmacydxb@gmail.com",
                    "To (The Buyer):\nKind Attn:\nTROJAN 244047",
                ],
            ],
            colWidths=[250, 250],
        )
        rows = [
            ["", "Tender No. : 244047", "", "", "", ""],
            ["", "Material Description", "Req Quantity", "unit", "u price", "total"],
            ["", "Deep Heat Spray 150ml", "5", "No", "12", "60"],
            ["", "Band Aid waterproof -\nBrand : Broplast (1x100)", "5", "Boxes", "8", "40"],
            ["", "Triangular Bandage", "30", "No", "3", "90"],
            ["", "Hand wash Liquid Dettol\n200ml", "10", "No", "8", "80"],
            ["", "Panadol Extra (1x48)", "5", "No", "30", "150"],
        ]
        continuation_rows = [
            ["", "Face mask Earloop\n(1x50) (Surgical mask)\nBrand :Biogreen", "5", "Pkts", "6", "30"],
            ["", "Alcohol swab (200/box)", "10", "No", "8", "80"],
        ]
        table_style = TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 1), (-1, 1), colors.lightgrey),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
        first_table = Table(rows, colWidths=[22, 245, 76, 55, 55, 55])
        first_table.setStyle(table_style)
        continuation_table = Table(continuation_rows, colWidths=[22, 245, 76, 55, 55, 55])
        continuation_table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        document.build(
            [
                Paragraph("QUOTATION", styles["Title"]),
                metadata_table,
                Spacer(1, 14),
                first_table,
                PageBreak(),
                continuation_table,
            ]
        )
        return SimpleUploadedFile(name, buffer.getvalue(), content_type="application/pdf")

    def test_import_actions_are_staff_only(self):
        actions = [
            ("post", reverse("quotation-inquiry-parse-text"), {"raw_text": "Panadol 500mg - 10 boxes"}, "json"),
            ("post", reverse("quotation-inquiry-parse-file"), {"file": self.make_excel_upload()}, "multipart"),
            (
                "post",
                reverse("quotation-inquiry-create-imported"),
                {
                    "company": self.company.id,
                    "source_type": Inquiry.SOURCE_TYPE_PASTED_TEXT,
                    "lines": [{"raw_name": "Panadol 500mg", "raw_line": "Panadol 500mg - 10 boxes"}],
                },
                "json",
            ),
        ]

        self.client.force_authenticate(None)
        for method, url, payload, request_format in actions:
            with self.subTest(url=url, user="anonymous"):
                response = getattr(self.client, method)(url, payload, format=request_format)
                self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

        self.client.force_authenticate(self.customer)
        for method, url, payload, request_format in actions:
            with self.subTest(url=url, user="customer"):
                if url.endswith("parse_file/"):
                    payload = {"file": self.make_excel_upload()}
                response = getattr(self.client, method)(url, payload, format=request_format)
                self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        staff_actions = [
            ("post", reverse("quotation-inquiry-parse-text"), {"raw_text": "Panadol 500mg - 10 boxes"}, "json", status.HTTP_200_OK),
            ("post", reverse("quotation-inquiry-parse-file"), {"file": self.make_excel_upload()}, "multipart", status.HTTP_200_OK),
            (
                "post",
                reverse("quotation-inquiry-create-imported"),
                {
                    "company": self.company.id,
                    "source_type": Inquiry.SOURCE_TYPE_PASTED_TEXT,
                    "lines": [{"raw_name": "Panadol 500mg", "raw_line": "Panadol 500mg - 10 boxes"}],
                },
                "json",
                status.HTTP_201_CREATED,
            ),
        ]
        for method, url, payload, request_format, expected_status in staff_actions:
            with self.subTest(url=url, user="staff"):
                response = getattr(self.client, method)(url, payload, format=request_format)
                self.assertEqual(response.status_code, expected_status)

    def test_parse_text_examples(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-text"),
            {
                "raw_text": "\n".join(
                    [
                        "Panadol 500mg - 10 boxes",
                        "Panadol 500mg x 10",
                        "10 boxes Panadol 500mg",
                        "Gloves medium 5 packs",
                        "1. Panadol 500mg - 10 box",
                    ]
                )
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["lines"]), 5)
        self.assertEqual(response.data["lines"][0]["raw_name"], "Panadol 500mg")
        self.assertEqual(response.data["lines"][0]["quantity"], "10")
        self.assertEqual(response.data["lines"][0]["unit"].lower(), "boxes")

    def test_invalid_extension_rejected(self):
        upload = SimpleUploadedFile("inquiry.txt", b"Panadol 500mg - 10 boxes", content_type="text/plain")
        response = self.client.post(reverse("quotation-inquiry-parse-file"), {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Unsupported file type", str(response.data))

    def test_invalid_excel_signature_rejected(self):
        upload = SimpleUploadedFile("inquiry.xlsx", b"not-a-zip-workbook", content_type="application/octet-stream")
        response = self.client.post(reverse("quotation-inquiry-parse-file"), {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("does not look like a valid .xlsx workbook", str(response.data))

    @override_settings(QUOTATION_IMPORT_MAX_UPLOAD_BYTES=12)
    def test_file_size_limit_rejected(self):
        upload = SimpleUploadedFile("big.xlsx", b"PK" + b"x" * 100, content_type="application/octet-stream")
        response = self.client.post(reverse("quotation-inquiry-parse-file"), {"file": upload}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("too large", str(response.data))

    def test_excel_parse_happy_path(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_excel_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["source_type"], Inquiry.SOURCE_TYPE_EXCEL)
        self.assertEqual(response.data["parse_method"], "openpyxl_structured_v2")
        self.assertEqual(len(response.data["lines"]), 2)
        self.assertEqual(response.data["lines"][0]["source_sheet"], "LPO")
        self.assertEqual(response.data["lines"][0]["raw_name"], "Panadol 500mg")
        self.assertIn("summary", response.data)
        self.assertIn("sheet_metadata", response.data["meta"])

    def test_first_aid_material_log_excel_pattern(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_first_aid_excel_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["parse_method"], "openpyxl_structured_v2")
        self.assertEqual(response.data["warnings"], [])
        self.assertEqual(len(response.data["lines"]), 4)
        self.assertEqual(response.data["meta"]["selected_sheets"][0]["header_row"], 2)

        expected = [
            ("ANTI SEPTIC SOLUTION", "1", "bottle", "1"),
            ("GAUZE PIECES", "1", "BOX", "2"),
            ("TRIANGULAR BANDAGES", "5", "Nos", "3"),
            ("SPLINTS", "2", "packs", "4"),
        ]
        for line, (name, quantity, unit, serial) in zip(response.data["lines"], expected):
            with self.subTest(name=name):
                self.assertEqual(line["raw_name"], name)
                self.assertEqual(line["quantity"], quantity)
                self.assertEqual(line["unit"], unit)
                self.assertEqual(line["serial_no"], serial)
                self.assertEqual(line["parse_status"], InquiryLine.PARSE_PARSED)
                self.assertGreaterEqual(line["parse_confidence"], 0.85)

    def test_multi_sheet_excel_selects_data_sheet_only(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_multi_sheet_excel_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        selected = [sheet["sheet_name"] for sheet in response.data["meta"]["selected_sheets"]]
        self.assertEqual(selected, ["Items"])
        self.assertEqual(response.data["lines"][0]["raw_name"], "GLOVES MEDIUM")

    def test_pdf_parse_happy_path(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_pdf_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["source_type"], Inquiry.SOURCE_TYPE_PDF)
        self.assertIn(response.data["parse_method"], ["pymupdf_pdfplumber_table_v2", "pymupdf_text_v2"])
        self.assertGreaterEqual(len(response.data["lines"]), 2)
        self.assertIn("Panadol 500mg", response.data["lines"][0]["raw_name"])

    def test_pdf_material_description_table_splits_price_columns_and_skips_metadata(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_material_description_pdf_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["warnings"], [])
        names = [line["raw_name"] for line in response.data["lines"]]
        self.assertNotIn("DATE: 023/05/2026", names)
        self.assertNotIn("Tender No.", names)
        self.assertFalse(any("Material Description" in name for name in names))

        first_line = response.data["lines"][0]
        self.assertEqual(first_line["raw_name"], "Deep Heat Spray 150ml")
        self.assertEqual(first_line["quantity"], "5")
        self.assertEqual(first_line["unit"], "No")
        self.assertEqual(first_line["unit_price"], "12")
        self.assertEqual(first_line["line_total"], "60")
        self.assertEqual(first_line["parse_status"], InquiryLine.PARSE_PARSED)

        band_aid = response.data["lines"][1]
        self.assertEqual(band_aid["raw_name"], "Band Aid waterproof - Brand : Broplast (1x100)")
        self.assertEqual(band_aid["quantity"], "5")
        self.assertEqual(band_aid["unit"], "Boxes")
        self.assertEqual(band_aid["unit_price"], "8")
        self.assertEqual(band_aid["line_total"], "40")

        continuation = next(line for line in response.data["lines"] if line["raw_name"].startswith("Face mask Earloop"))
        self.assertEqual(continuation["quantity"], "5")
        self.assertEqual(continuation["unit_price"], "6")

    def test_pasted_email_price_text_extracts_item_price_and_ambiguous_quantity(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-text"),
            {
                "raw_text": "\n".join(
                    [
                        "Electrorush 21gm sache for 1 Ltr solution",
                        "1 box 10 sachets 50 box cartoon price : 375 per cartoon",
                        "",
                        "Zest Ors 21 gm sachet per for 1 Ltr solution, 25 sachet per box price 18 per box",
                    ]
                )
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["lines"]), 2)
        electrorush = response.data["lines"][0]
        self.assertEqual(electrorush["raw_name"], "Electrorush 21gm sache for 1 Ltr solution")
        self.assertEqual(electrorush["quantity"], "50")
        self.assertEqual(electrorush["unit"], "carton")
        self.assertEqual(electrorush["unit_price"], "375")
        self.assertIn("1 box 10 sachets", electrorush["notes"])

        zest = response.data["lines"][1]
        self.assertEqual(zest["raw_name"], "Zest Ors 21 gm sachet per for 1 Ltr solution")
        self.assertIsNone(zest["quantity"])
        self.assertEqual(zest["unit"], "box")
        self.assertEqual(zest["unit_price"], "18")
        self.assertEqual(zest["parse_status"], InquiryLine.PARSE_NEEDS_REVIEW)

    def test_pdf_no_selectable_text_returns_warning(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_pdf_upload(text=False)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["lines"], [])
        self.assertIn("No selectable text detected", response.data["warnings"][0])

    def test_encrypted_pdf_rejected(self):
        response = self.client.post(
            reverse("quotation-inquiry-parse-file"),
            {"file": self.make_pdf_upload(text=False, encrypted=True)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Encrypted PDF files are not supported", str(response.data))

    def test_create_imported_creates_inquiry_and_lines_atomically(self):
        payload = {
            "company": self.company.id,
            "contact": self.contact.id,
            "subject": "Imported LPO",
            "original_text": "Panadol 500mg - 10 boxes",
            "source_type": Inquiry.SOURCE_TYPE_PASTED_TEXT,
            "source_filename": "",
            "source_mime_type": "text/plain",
            "source_sha256": "a" * 64,
            "source_file_ref": "inquiry_sources/2026/05/22/a_import.xlsx",
            "source_file_size": 1234,
            "parse_method": "deterministic_text_v1",
            "parse_meta": {"warnings": []},
            "lines": [
                {
                    "raw_name": "Panadol 500mg",
                    "raw_line": "Panadol 500mg - 10 boxes",
                    "quantity": "10.000",
                    "unit": "boxes",
                    "parse_status": InquiryLine.PARSE_PARSED,
                    "parse_confidence": 0.9,
                }
            ],
        }

        response = self.client.post(reverse("quotation-inquiry-create-imported"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        inquiry = Inquiry.objects.get(id=response.data["id"])
        self.assertEqual(inquiry.source, Inquiry.SOURCE_IMPORTED)
        self.assertEqual(inquiry.source_type, Inquiry.SOURCE_TYPE_PASTED_TEXT)
        self.assertEqual(inquiry.source_file_ref, "inquiry_sources/2026/05/22/a_import.xlsx")
        self.assertEqual(inquiry.source_file_size, 1234)
        self.assertEqual(inquiry.lines.count(), 1)

    def test_create_imported_rejects_unsafe_source_file_ref(self):
        payload = {
            "company": self.company.id,
            "source_type": Inquiry.SOURCE_TYPE_EXCEL,
            "source_file_ref": "../secret.xlsx",
            "lines": [{"raw_name": "Panadol 500mg", "raw_line": "Panadol 500mg - 10 boxes"}],
        }

        response = self.client.post(reverse("quotation-inquiry-create-imported"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("source_file_ref", response.data)


class HistoricalPriceImportTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(username="historical_staff", password="pass", is_staff=True)
        self.customer = User.objects.create_user(username="historical_customer", password="pass")
        self.company = Company.objects.create(name="Ancient Builders Constructions LLC")
        self.item_one = Product.objects.create(name="Savlon Antiseptic Solution 1000ml", price=Decimal("1.00"), pack_size="bottle", status="draft")
        self.item_two = Product.objects.create(name="Gauze Pieces", price=Decimal("1.00"), pack_size="BOX", status="draft")
        self.client.force_authenticate(self.staff)

    def make_historical_pdf_upload(
        self,
        name="ANCIENT BUILDERS CONSTN 21052026.pdf",
        encrypted=False,
        document_number="QUOTATION-26052101",
        document_date="21/05/2026",
        extra_text="",
    ):
        buffer = BytesIO()
        if encrypted:
            writer = PdfWriter()
            writer.add_blank_page(width=300, height=300)
            writer.encrypt("secret")
            writer.write(buffer)
            data = buffer.getvalue()
        else:
            styles = getSampleStyleSheet()
            document = SimpleDocTemplate(buffer, pagesize=A4)
            rows = [
                ["SN", "ITEM DESCRIPTION", "UOM", "QTY", "U/P", "AMOUNT", "VAT", "TOTAL"],
                ["1", "SAVLON ANTISEPTIC SOLUTION", "1000ml\nbottle", "1", "5.00", "5.00", "0.25", "5.25"],
                ["2", "GAUZE PIECES", "BOX", "3", "2.50", "7.50", "0.38", "7.88"],
                ["TOTAL", "", "", "", "", "12.50", "0.63", "13.13"],
            ]
            table = Table(rows, colWidths=[28, 170, 70, 42, 48, 58, 48, 58])
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            document.build(
                [
                    Paragraph(document_number, styles["Title"]),
                    Paragraph(f"DATE :{document_date}", styles["Normal"]),
                    Paragraph(extra_text, styles["Normal"]) if extra_text else Spacer(1, 0),
                    Spacer(1, 16),
                    table,
                ]
            )
            data = buffer.getvalue()
        return SimpleUploadedFile(name, data, content_type="application/pdf")

    def make_material_description_historical_pdf_upload(self, name="244047.pdf"):
        buffer = BytesIO()
        styles = getSampleStyleSheet()
        document = SimpleDocTemplate(buffer, pagesize=A4)
        rows = [
            ["", "Tender No. : 244047", "", "", "", ""],
            ["", "Material Description", "Req Quantity", "unit", "u price", "total"],
            ["", "Deep Heat Spray 150ml", "5", "No", "12", "60"],
            ["", "Band Aid waterproof -\nBrand : Broplast (1x100)", "5", "Boxes", "8", "40"],
            ["", "Triangular Bandage", "30", "No", "3", "90"],
            ["", "Hand wash Liquid Dettol\n200ml", "10", "No", "8", "80"],
            ["", "Panadol Extra (1x48)", "5", "No", "30", "150"],
        ]
        continuation_rows = [
            ["", "Face mask Earloop\n(1x50) (Surgical mask)\nBrand :Biogreen", "5", "Pkts", "6", "30"],
            ["", "Alcohol swab (200/box)", "10", "No", "8", "80"],
        ]
        table_style = TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 1), (-1, 1), colors.lightgrey),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
        first_table = Table(rows, colWidths=[22, 245, 76, 55, 55, 55])
        first_table.setStyle(table_style)
        continuation_table = Table(continuation_rows, colWidths=[22, 245, 76, 55, 55, 55])
        continuation_table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        document.build(
            [
                Paragraph("QUOTATION", styles["Title"]),
                Paragraph("DATE: 023/05/2026", styles["Normal"]),
                Spacer(1, 12),
                first_table,
                PageBreak(),
                continuation_table,
            ]
        )
        return SimpleUploadedFile(name, buffer.getvalue(), content_type="application/pdf")

    def make_qty_uom_shifted_historical_pdf_upload(self, name="qty-uom-shift.pdf"):
        buffer = BytesIO()
        styles = getSampleStyleSheet()
        document = SimpleDocTemplate(buffer, pagesize=A4)
        rows = [
            ["S. No.", "Item Description", "Qty", "UOM", "U PRICE", "TOTAL", "VAT", "G TOTAL"],
            ["1", "Fastum Gel 50gm", "Tubes.", "3", "30", "90", "0", "90"],
            ["2", "Silvadiazin Ointment 30gm", "Tubes.", "2", "8", "16", "0", "16"],
        ]
        table = Table(rows, colWidths=[40, 185, 60, 48, 58, 58, 48, 58])
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        document.build(
            [
                Paragraph("QUOTATION-FASTUM", styles["Title"]),
                Paragraph("DATE :21/05/2026", styles["Normal"]),
                Spacer(1, 16),
                table,
            ]
        )
        return SimpleUploadedFile(name, buffer.getvalue(), content_type="application/pdf")

    def parse_historical_import(self):
        return self.client.post(
            reverse("quotation-historical-import-parse-file"),
            {"file": self.make_historical_pdf_upload()},
            format="multipart",
        )

    def create_parsed_historical_import(self):
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                response = self.parse_historical_import()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        return response

    def test_historical_import_parse_action_is_staff_only(self):
        self.client.force_authenticate(None)
        anonymous = self.client.post(
            reverse("quotation-historical-import-parse-file"),
            {"file": self.make_historical_pdf_upload()},
            format="multipart",
        )
        self.assertIn(anonymous.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

        self.client.force_authenticate(self.customer)
        non_staff = self.client.post(
            reverse("quotation-historical-import-parse-file"),
            {"file": self.make_historical_pdf_upload()},
            format="multipart",
        )
        self.assertEqual(non_staff.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                staff = self.parse_historical_import()
        self.assertEqual(staff.status_code, status.HTTP_201_CREATED)

    def test_historical_pdf_parse_stages_review_rows_and_private_source_ref(self):
        response = self.create_parsed_historical_import()

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["source_type"], HistoricalPriceImport.SOURCE_TYPE_PDF)
        self.assertEqual(response.data["parse_method"], "al_ameen_pdf_price_table_v1")
        self.assertEqual(response.data["document_number"], "QUOTATION-26052101")
        self.assertEqual(response.data["document_date"], "2026-05-21")
        self.assertIn("Ancient Builders", response.data["suggested_company_name"])
        self.assertTrue(response.data["source_file_ref"])
        self.assertEqual(len(response.data["lines"]), 2)
        first_line = response.data["lines"][0]
        self.assertEqual(first_line["item_name"], "SAVLON ANTISEPTIC SOLUTION 1000ml")
        self.assertEqual(first_line["unit"].lower(), "bottle")
        self.assertEqual(first_line["quantity"], "1.000")
        self.assertEqual(first_line["unit_price"], "5.00")
        self.assertEqual(first_line["status"], HistoricalPriceImportLine.STATUS_NEEDS_REVIEW)

    def test_historical_pdf_material_description_table_parses_price_rows(self):
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                response = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {"file": self.make_material_description_historical_pdf_upload()},
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["document_number"], "244047")
        self.assertEqual(response.data["document_date"], "2026-05-23")
        self.assertEqual(len(response.data["lines"]), 7)

        first_line = response.data["lines"][0]
        self.assertEqual(first_line["item_name"], "Deep Heat Spray 150ml")
        self.assertEqual(first_line["quantity"], "5.000")
        self.assertEqual(first_line["unit"], "No")
        self.assertEqual(first_line["unit_price"], "12.00")
        self.assertEqual(first_line["line_total"], "60.00")

        band_aid = response.data["lines"][1]
        self.assertEqual(band_aid["item_name"], "Band Aid waterproof - Brand : Broplast (1x100)")
        self.assertEqual(band_aid["quantity"], "5.000")
        self.assertEqual(band_aid["unit"], "Boxes")
        self.assertEqual(band_aid["unit_price"], "8.00")
        self.assertEqual(band_aid["line_total"], "40.00")

    def test_historical_import_commit_appends_price_history_and_hides_backfill_quote(self):
        response = self.create_parsed_historical_import()
        import_id = response.data["id"]
        lines = response.data["lines"]

        self.client.patch(
            reverse("quotation-historical-import-detail", args=[import_id]),
            {"company": self.company.id},
            format="json",
        )
        self.client.patch(
            reverse("quotation-historical-import-line-detail", args=[lines[0]["id"]]),
            {"product": self.item_one.id, "status": HistoricalPriceImportLine.STATUS_READY},
            format="json",
        )
        self.client.patch(
            reverse("quotation-historical-import-line-detail", args=[lines[1]["id"]]),
            {"product": self.item_two.id, "status": HistoricalPriceImportLine.STATUS_READY},
            format="json",
        )

        commit = self.client.post(reverse("quotation-historical-import-commit", args=[import_id]))

        self.assertEqual(commit.status_code, status.HTTP_200_OK)
        historical_import = HistoricalPriceImport.objects.get(pk=import_id)
        self.assertEqual(historical_import.status, HistoricalPriceImport.STATUS_COMMITTED)
        self.assertEqual(CompanyPriceHistory.objects.filter(company=self.company).count(), 2)
        self.assertTrue(historical_import.created_quotation.is_historical_import)
        self.assertEqual(historical_import.created_quotation.status, Quotation.STATUS_FINALIZED)

        default_quotes = self.client.get(reverse("quotation-list"))
        self.assertEqual(default_quotes.status_code, status.HTTP_200_OK)
        self.assertNotIn(historical_import.created_quotation.id, [quote["id"] for quote in default_quotes.data])

        all_quotes = self.client.get(reverse("quotation-list"), {"include_historical": "true"})
        self.assertIn(historical_import.created_quotation.id, [quote["id"] for quote in all_quotes.data])

    def test_historical_import_commit_allows_missing_document_number_with_fallback_reference(self):
        historical_import = HistoricalPriceImport.objects.create(
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="Intermass 30032026.pdf",
            source_sha256="a" * 64,
            document_date=date(2026, 3, 30),
            created_by=self.staff,
        )
        HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="SAVLON ANTISEPTIC SOLUTION 1000ml",
            quantity=Decimal("1.000"),
            unit="bottle",
            unit_price=Decimal("5.00"),
            product=self.item_one,
            status=HistoricalPriceImportLine.STATUS_READY,
        )

        response = self.client.post(reverse("quotation-historical-import-commit", args=[historical_import.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        historical_import.refresh_from_db()
        self.assertEqual(historical_import.status, HistoricalPriceImport.STATUS_COMMITTED)
        self.assertEqual(historical_import.created_quotation.quotation_number, f"HIST-{historical_import.id:06d}")
        self.assertEqual(CompanyPriceHistory.objects.filter(company=self.company, product=self.item_one).count(), 1)

    def test_batch_commit_allows_missing_document_number_when_rows_are_ready(self):
        batch = HistoricalImportBatch.objects.create(name="Initial learning batch", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="Intermass 30032026.pdf",
            source_sha256="b" * 64,
            document_date=date(2026, 3, 30),
            created_by=self.staff,
        )
        HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="SAVLON ANTISEPTIC SOLUTION 1000ml",
            quantity=Decimal("1.000"),
            unit="bottle",
            unit_price=Decimal("5.00"),
            product=self.item_one,
            status=HistoricalPriceImportLine.STATUS_READY,
        )

        detail = self.client.get(reverse("quotation-historical-import-batch-detail", args=[batch.id]))
        blocker = detail.data["wizard_summary"]["commit_blockers"][0]
        self.assertEqual(blocker["ready_row_count"], 1)
        self.assertNotIn("missing document number", blocker["blockers"])
        self.assertTrue(blocker["can_commit"])

        response = self.client.post(
            reverse("quotation-historical-import-batch-commit-ready-imports", args=[batch.id]),
            {"import_ids": [historical_import.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["summary"]["committed"], 1)
        historical_import.refresh_from_db()
        self.assertEqual(historical_import.status, HistoricalPriceImport.STATUS_COMMITTED)

    def test_historical_import_duplicate_commit_does_not_append_price_history_twice(self):
        upload = self.make_historical_pdf_upload()
        upload_bytes = upload.read()

        def duplicate_upload():
            return SimpleUploadedFile(upload.name, upload_bytes, content_type="application/pdf")

        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                first = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {"file": duplicate_upload()},
                    format="multipart",
                )
                second = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {"file": duplicate_upload()},
                    format="multipart",
                )

        self.assertEqual(first.status_code, status.HTTP_201_CREATED)
        self.assertEqual(second.status_code, status.HTTP_200_OK)
        self.assertEqual(second.data["id"], first.data["id"])
        self.assertTrue(second.data["duplicate_check"]["blocked_new_import"])
        self.assertEqual(second.data["duplicate_check"]["message"], "This PDF has already been added before.")
        self.assertEqual(HistoricalPriceImport.objects.filter(source_sha256=first.data["source_sha256"]).count(), 1)

        import_id = first.data["id"]
        self.client.patch(
            reverse("quotation-historical-import-detail", args=[import_id]),
            {"company": self.company.id},
            format="json",
        )
        self.client.patch(
            reverse("quotation-historical-import-line-detail", args=[first.data["lines"][0]["id"]]),
            {"product": self.item_one.id, "status": HistoricalPriceImportLine.STATUS_READY},
            format="json",
        )

        first_commit = self.client.post(reverse("quotation-historical-import-commit", args=[import_id]))
        third = self.client.post(
            reverse("quotation-historical-import-parse-file"),
            {"file": duplicate_upload()},
            format="multipart",
        )

        self.assertEqual(first_commit.status_code, status.HTTP_200_OK)
        self.assertEqual(third.status_code, status.HTTP_200_OK)
        self.assertEqual(third.data["id"], import_id)
        self.assertEqual(third.data["status"], HistoricalPriceImport.STATUS_COMMITTED)
        self.assertEqual(CompanyPriceHistory.objects.filter(company=self.company, product=self.item_one).count(), 1)

    def test_batch_upload_creates_staged_imports_and_detects_duplicate_per_file(self):
        batch = self.client.post(reverse("quotation-historical-import-batch-list"), {"name": "May history"}, format="json")
        self.assertEqual(batch.status_code, status.HTTP_201_CREATED)
        batch_id = batch.data["id"]
        upload = self.make_historical_pdf_upload()
        upload_bytes = upload.read()

        def file_upload():
            return SimpleUploadedFile(upload.name, upload_bytes, content_type="application/pdf")

        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                first = self.client.post(
                    reverse("quotation-historical-import-batch-upload-file", args=[batch_id]),
                    {"file": file_upload()},
                    format="multipart",
                )
                second = self.client.post(
                    reverse("quotation-historical-import-batch-upload-file", args=[batch_id]),
                    {"file": file_upload()},
                    format="multipart",
                )

        self.assertEqual(first.status_code, status.HTTP_201_CREATED)
        self.assertEqual(first.data["status"], "parsed")
        self.assertEqual(first.data["import"]["batch"], batch_id)
        self.assertEqual(second.status_code, status.HTTP_200_OK)
        self.assertEqual(second.data["status"], "duplicate")
        self.assertTrue(second.data["duplicate_check"]["blocked_new_import"])
        self.assertEqual(HistoricalPriceImport.objects.filter(batch_id=batch_id).count(), 1)
        refreshed_batch = HistoricalImportBatch.objects.get(pk=batch_id)
        self.assertEqual(refreshed_batch.summary["duplicate_file_count"], 1)
        self.assertEqual(second.data["duplicate_check"]["primary_match"]["id"], first.data["import"]["id"])
        self.assertEqual(second.data["batch"]["summary"]["files"][-1]["duplicate_match"]["id"], first.data["import"]["id"])

    def test_historical_import_can_be_removed_from_batch_before_line_decisions_are_applied(self):
        batch = self.client.post(reverse("quotation-historical-import-batch-list"), {"name": "Duplicate review"}, format="json")
        batch_id = batch.data["id"]
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                upload = self.client.post(
                    reverse("quotation-historical-import-batch-upload-file", args=[batch_id]),
                    {"file": self.make_historical_pdf_upload()},
                    format="multipart",
                )
        self.assertEqual(upload.status_code, status.HTTP_201_CREATED)
        historical_import = HistoricalPriceImport.objects.get(pk=upload.data["import"]["id"])
        line = historical_import.lines.first()
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch_id=batch_id,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_NEEDS_MANUAL_REVIEW,
            status=HistoricalImportAISuggestion.STATUS_PENDING,
            reason="Pending review.",
            created_by=self.staff,
        )

        response = self.client.post(reverse("quotation-historical-import-remove-from-batch", args=[historical_import.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        historical_import.refresh_from_db()
        suggestion.refresh_from_db()
        self.assertIsNone(historical_import.batch_id)
        self.assertEqual(historical_import.status, HistoricalPriceImport.STATUS_CANCELLED)
        self.assertIsNone(suggestion.batch_id)
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_REJECTED)
        self.assertEqual(response.data["status"], "removed")
        self.assertEqual(response.data["batch"]["imports"], [])
        self.assertEqual(response.data["batch"]["summary"]["files"][0]["status"], "removed")

    def test_historical_filename_company_hint_strips_date_suffix_and_matches_existing_company(self):
        intermass = Company.objects.create(name="Intermass")
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                response = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {"file": self.make_historical_pdf_upload(name="Intermass 27032026A.pdf")},
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["suggested_company_name"], "Intermass")
        self.assertEqual(response.data["company"], intermass.id)
        historical_import = HistoricalPriceImport.objects.get(pk=response.data["id"])
        self.assertEqual(historical_import.parse_meta["company_match"]["company_name"], "Intermass")

    def test_historical_parser_keeps_uom_text_and_numeric_quantity_when_columns_shift(self):
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                response = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {"file": self.make_qty_uom_shifted_historical_pdf_upload()},
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        lines = {line["item_name"]: line for line in response.data["lines"]}
        fastum = lines["Fastum Gel 50gm"]
        silvadiazin = lines["Silvadiazin Ointment 30gm"]
        self.assertEqual(fastum["quantity"], "3.000")
        self.assertEqual(fastum["unit"], "Tubes.")
        self.assertEqual(fastum["unit_price"], "30.00")
        self.assertEqual(fastum["line_total"], "90.00")
        self.assertEqual(silvadiazin["quantity"], "2.000")
        self.assertEqual(silvadiazin["unit"], "Tubes.")
        self.assertEqual(silvadiazin["unit_price"], "8.00")
        self.assertEqual(silvadiazin["line_total"], "16.00")

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
        QUOTATION_AI_PARSE_VISION_MODEL="test-vision-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_ai_learning_suggestions_are_review_only_until_staff_applies(self):
        batch = self.client.post(reverse("quotation-historical-import-batch-list"), {"name": "AI learning"}, format="json")
        batch_id = batch.data["id"]
        settings_obj = QuotationSettings.get_solo()
        settings_obj.ai_parsing_enabled = True
        settings_obj.save()
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                upload = self.client.post(
                    reverse("quotation-historical-import-batch-upload-file", args=[batch_id]),
                    {"file": self.make_historical_pdf_upload()},
                    format="multipart",
                )
        self.assertEqual(upload.status_code, status.HTTP_201_CREATED)
        before_products = Product.objects.count()

        provider = MockLearningProvider()
        with patch("quotations.ai_learning.get_ai_parse_provider", return_value=provider):
            response = self.client.post(
                reverse("quotation-historical-import-batch-run-ai-suggestions", args=[batch_id]),
                {"import_ids": [upload.data["import"]["id"]], "mode": "text"},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(HistoricalImportAISuggestion.objects.filter(batch_id=batch_id, status=HistoricalImportAISuggestion.STATUS_PENDING).count(), 3)
        self.assertEqual(Product.objects.count(), before_products)
        self.assertEqual(ProductAlias.objects.count(), 0)
        self.assertEqual(CompanyPriceHistory.objects.count(), 0)
        self.assertEqual(provider.calls[0]["schema_name"], "quotation_historical_learning")

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_staff_approval_creates_alias_and_new_draft_product_then_commit_ready_rows(self):
        batch = self.client.post(reverse("quotation-historical-import-batch-list"), {"name": "AI approve"}, format="json")
        batch_id = batch.data["id"]
        settings_obj = QuotationSettings.get_solo()
        settings_obj.ai_parsing_enabled = True
        settings_obj.save()
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                upload = self.client.post(
                    reverse("quotation-historical-import-batch-upload-file", args=[batch_id]),
                    {"file": self.make_historical_pdf_upload()},
                    format="multipart",
                )
        import_id = upload.data["import"]["id"]
        with patch("quotations.ai_learning.get_ai_parse_provider", return_value=MockLearningProvider()):
            self.client.post(
                reverse("quotation-historical-import-batch-run-ai-suggestions", args=[batch_id]),
                {"import_ids": [import_id], "mode": "text"},
                format="json",
            )

        company_suggestion = HistoricalImportAISuggestion.objects.get(
            batch_id=batch_id,
            suggestion_type=HistoricalImportAISuggestion.TYPE_COMPANY,
        )
        line_suggestions = list(
            HistoricalImportAISuggestion.objects.filter(
                batch_id=batch_id,
                suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            ).order_by("line__sort_order")
        )
        apply_company = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch_id]),
            {"suggestion_ids": [company_suggestion.id]},
            format="json",
        )
        apply_lines = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch_id]),
            {"suggestion_ids": [line_suggestions[0].id, line_suggestions[1].id]},
            format="json",
        )

        self.assertEqual(apply_company.status_code, status.HTTP_200_OK)
        self.assertEqual(apply_lines.status_code, status.HTTP_200_OK)
        self.assertEqual(apply_lines.data["summary"]["applied"], 2)
        self.assertEqual(ProductAlias.objects.filter(company=self.company, product=self.item_one).count(), 1)
        created_product = Product.objects.get(name="Custom Historical Refill")
        self.assertEqual(created_product.status, "draft")
        self.assertFalse(created_product.show_price)
        self.assertEqual(CompanyPriceHistory.objects.count(), 0)

        commit = self.client.post(
            reverse("quotation-historical-import-batch-commit-ready-imports", args=[batch_id]),
            {"import_ids": [import_id]},
            format="json",
        )

        self.assertEqual(commit.status_code, status.HTTP_200_OK)
        self.assertEqual(commit.data["summary"]["committed"], 1)
        self.assertEqual(CompanyPriceHistory.objects.filter(company=self.company).count(), 2)

    def test_alias_conflict_blocks_ai_suggestion_approval_without_overwriting(self):
        response = self.create_parsed_historical_import()
        import_id = response.data["id"]
        historical_import = HistoricalPriceImport.objects.get(pk=import_id)
        historical_import.company = self.company
        historical_import.save(update_fields=["company", "updated_at"])
        line = historical_import.lines.order_by("sort_order").first()
        ProductAlias.objects.create(company=self.company, product=self.item_two, alias=line.item_name, created_by=self.staff)
        suggestion = HistoricalImportAISuggestion.objects.create(
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_CREATE_COMPANY_ALIAS,
            suggested_product=self.item_one,
            alias_text=line.item_name,
            confidence=0.95,
            reason="AI thinks this maps to item one.",
            created_by=self.staff,
        )

        result = self.client.post(
            reverse("quotation-historical-import-ai-suggestion-apply"),
            {"suggestion_ids": [suggestion.id]},
            format="json",
        )

        self.assertEqual(result.status_code, status.HTTP_200_OK)
        suggestion.refresh_from_db()
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_CONFLICT)
        self.assertEqual(ProductAlias.objects.get(company=self.company, normalized_alias=line.normalized_item_name).product, self.item_two)

    def test_repeated_new_product_suggestions_reuse_one_draft_product_and_mark_rows_ready(self):
        batch = HistoricalImportBatch.objects.create(name="Repeated ENO", created_by=self.staff)
        first_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="eno-latest.pdf",
            source_sha256="a" * 64,
            document_number="Q-ENO-2",
            document_date=date(2026, 5, 28),
            created_by=self.staff,
        )
        second_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="eno-old.pdf",
            source_sha256="b" * 64,
            document_number="Q-ENO-1",
            document_date=date(2026, 5, 20),
            created_by=self.staff,
        )
        line_one = HistoricalPriceImportLine.objects.create(
            historical_import=first_import,
            item_name="ENO Sachet",
            quantity=Decimal("2.000"),
            unit="box",
            unit_price=Decimal("12.00"),
            sort_order=1,
        )
        line_two = HistoricalPriceImportLine.objects.create(
            historical_import=second_import,
            item_name="ENO Sachet",
            quantity=Decimal("1.000"),
            unit="box",
            unit_price=Decimal("11.50"),
            sort_order=1,
        )
        first_suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=first_import,
            line=line_one,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_CREATE_NEW_PRODUCT,
            proposed_product_name="ENO Sachet",
            proposed_pack_size="box",
            confidence=0.91,
            reason="Likely new Product.",
            created_by=self.staff,
        )
        second_suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=second_import,
            line=line_two,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_CREATE_NEW_PRODUCT,
            proposed_product_name="ENO Sachet",
            proposed_pack_size="box",
            confidence=0.9,
            reason="Repeated Product.",
            created_by=self.staff,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
            {"suggestion_ids": [first_suggestion.id, second_suggestion.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["summary"]["applied"], 1)
        self.assertEqual(response.data["summary"]["applied_similar"], 1)
        self.assertEqual(Product.objects.filter(name__iexact="ENO Sachet").count(), 1)
        product = Product.objects.get(name__iexact="ENO Sachet")
        line_one.refresh_from_db()
        line_two.refresh_from_db()
        self.assertEqual(line_one.product, product)
        self.assertEqual(line_two.product, product)
        self.assertEqual(line_one.status, HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(line_two.status, HistoricalPriceImportLine.STATUS_READY)

    def test_apply_ai_suggestion_returns_updated_card_state(self):
        batch = HistoricalImportBatch.objects.create(name="Updated cards", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="deep-heat.pdf",
            source_sha256="d" * 64,
            document_number="Q-DH-1",
            document_date=date(2026, 6, 1),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="DEEP HEAT SPRAY",
            quantity=Decimal("10.000"),
            unit="BOTTLES",
            unit_price=Decimal("18.00"),
            line_total=Decimal("180.00"),
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            confidence=0.95,
            reason="Strong product match.",
            created_by=self.staff,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
            {"suggestion_ids": [suggestion.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["summary"]["applied"], 1)
        updated = response.data["updated_suggestions"][0]
        self.assertEqual(updated["id"], suggestion.id)
        self.assertEqual(updated["status"], HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(updated["line_status"], HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(updated["suggested_product"], self.item_one.id)

    def test_apply_alias_suggestion_returns_ready_card_state(self):
        batch = HistoricalImportBatch.objects.create(name="Updated alias cards", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="tongue-depressors.pdf",
            source_sha256="c" * 64,
            document_date=date(2026, 6, 1),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="TONGUE DEPRESSORS",
            quantity=Decimal("1.000"),
            unit="PKT",
            unit_price=Decimal("15.00"),
            line_total=Decimal("15.00"),
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_CREATE_COMPANY_ALIAS,
            suggested_product=self.item_two,
            alias_text="TONGUE DEPRESSORS",
            confidence=0.90,
            reason="Customer wording for an existing Product.",
            created_by=self.staff,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
            {"suggestion_ids": [suggestion.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        suggestion.refresh_from_db()
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(line.product, self.item_two)
        self.assertEqual(line.status, HistoricalPriceImportLine.STATUS_READY)
        self.assertTrue(ProductAlias.objects.filter(company=self.company, alias="TONGUE DEPRESSORS", product=self.item_two).exists())
        updated = response.data["updated_suggestions"][0]
        self.assertEqual(updated["id"], suggestion.id)
        self.assertEqual(updated["status"], HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(updated["line_status"], HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(updated["line_ready_blockers"], [])

    def test_apply_alias_suggestion_reuses_existing_same_product_alias(self):
        ProductAlias.objects.create(
            company=self.company,
            product=self.item_two,
            alias="TONGUE DEPRESSORS",
            created_by=self.staff,
        )
        batch = HistoricalImportBatch.objects.create(name="Existing alias approval", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="tongue-depressors-repeat.pdf",
            source_sha256="8" * 64,
            document_date=date(2026, 6, 1),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="TONGUE DEPRESSORS",
            quantity=Decimal("1.000"),
            unit="PKT",
            unit_price=Decimal("15.00"),
            line_total=Decimal("15.00"),
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_CREATE_COMPANY_ALIAS,
            suggested_product=self.item_two,
            alias_text="TONGUE DEPRESSORS",
            confidence=0.90,
            reason="Customer wording for an existing Product.",
            created_by=self.staff,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
            {"suggestion_ids": [suggestion.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        suggestion.refresh_from_db()
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(line.product, self.item_two)
        self.assertEqual(line.status, HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(
            ProductAlias.objects.filter(company=self.company, normalized_alias=line.normalized_item_name).count(),
            1,
        )
        updated = response.data["updated_suggestions"][0]
        self.assertEqual(updated["line_status"], HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(updated["line_ready_blockers"], [])

    def test_apply_new_product_suggestion_returns_ready_card_state(self):
        batch = HistoricalImportBatch.objects.create(name="Updated new product cards", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="fastum.pdf",
            source_sha256="9" * 64,
            document_date=date(2026, 6, 1),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="FASTUM GEL 50GM",
            quantity=Decimal("3.000"),
            unit="Tubes",
            unit_price=Decimal("30.00"),
            line_total=Decimal("90.00"),
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_CREATE_NEW_PRODUCT,
            proposed_product_name="Fastum Gel 50gm",
            proposed_pack_size="Tubes",
            confidence=0.91,
            reason="Genuine new Product.",
            created_by=self.staff,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
            {"suggestion_ids": [suggestion.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        suggestion.refresh_from_db()
        product = Product.objects.get(name="Fastum Gel 50gm")
        self.assertEqual(product.status, "draft")
        self.assertFalse(product.show_price)
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(line.product, product)
        self.assertEqual(line.status, HistoricalPriceImportLine.STATUS_READY)
        updated = response.data["updated_suggestions"][0]
        self.assertEqual(updated["id"], suggestion.id)
        self.assertEqual(updated["status"], HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(updated["line_status"], HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(updated["line_ready_blockers"], [])

    def test_batch_apply_ai_suggestions_returns_json_for_database_conflict(self):
        batch = HistoricalImportBatch.objects.create(name="Apply conflict", created_by=self.staff)

        with self.assertLogs("quotations.views", level="ERROR"), patch(
            "quotations.views.apply_historical_ai_suggestions",
            side_effect=IntegrityError("duplicate alias"),
        ):
            response = self.client.post(
                reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
                {"suggestion_ids": [123]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIsInstance(response.data, dict)
        self.assertIn("detail", response.data)
        self.assertIn("duplicate or conflicting", response.data["detail"])
        self.assertNotIn("<html", str(response.data).lower())

    def test_batch_apply_ai_suggestions_returns_json_for_unexpected_old_batch_error(self):
        batch = HistoricalImportBatch.objects.create(name="Apply old batch error", created_by=self.staff)

        with self.assertLogs("quotations.views", level="ERROR"), patch(
            "quotations.views.apply_historical_ai_suggestions",
            side_effect=RuntimeError("legacy stale row"),
        ):
            response = self.client.post(
                reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
                {"suggestion_ids": [123]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIsInstance(response.data, dict)
        self.assertIn("Apply AI suggestions failed", response.data["detail"])
        self.assertNotIn("<html", str(response.data).lower())

    def test_batch_apply_ai_suggestions_returns_json_when_response_refresh_fails(self):
        batch = HistoricalImportBatch.objects.create(name="Apply refresh error", created_by=self.staff)

        with self.assertLogs("quotations.views", level="ERROR"), patch(
            "quotations.views.apply_historical_ai_suggestions",
            return_value=({"applied": 1}, [{"suggestion_id": 123, "status": "applied"}]),
        ), patch(
            "quotations.views.refresh_historical_import_batch_summary",
            side_effect=RuntimeError("legacy summary refresh failed"),
        ):
            response = self.client.post(
                reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
                {"suggestion_ids": [123]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIsInstance(response.data, dict)
        self.assertIn("Apply AI suggestions failed", response.data["detail"])
        self.assertNotIn("<html", str(response.data).lower())

    def test_batch_apply_ai_suggestions_returns_json_when_updated_suggestion_serializing_fails(self):
        batch = HistoricalImportBatch.objects.create(name="Apply serializer error", created_by=self.staff)

        with self.assertLogs("quotations.views", level="ERROR"), patch(
            "quotations.views.apply_historical_ai_suggestions",
            return_value=({"applied": 1}, [{"suggestion_id": 123, "status": "applied"}]),
        ), patch(
            "quotations.views._serialized_ai_suggestions_for_results",
            side_effect=RuntimeError("legacy suggestion serialize failed"),
        ):
            response = self.client.post(
                reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
                {"suggestion_ids": [123]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIsInstance(response.data, dict)
        self.assertIn("Apply AI suggestions failed", response.data["detail"])
        self.assertNotIn("<html", str(response.data).lower())

    def test_apply_ai_suggestion_marks_product_row_ready_even_when_document_details_missing(self):
        batch = HistoricalImportBatch.objects.create(name="Missing document details", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            suggested_company_name="Intermass 17032026A",
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="Intermass 17032026A.pdf",
            source_sha256="f" * 64,
            document_date=date(2026, 3, 17),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="DEEP HEAT SPRAY",
            quantity=Decimal("10.000"),
            unit="BOTTLES",
            unit_price=Decimal("18.00"),
            line_total=Decimal("180.00"),
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            confidence=0.95,
            reason="Strong product match.",
            created_by=self.staff,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-apply-ai-suggestions", args=[batch.id]),
            {"suggestion_ids": [suggestion.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        line.refresh_from_db()
        suggestion.refresh_from_db()
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(line.status, HistoricalPriceImportLine.STATUS_READY)
        updated = response.data["updated_suggestions"][0]
        self.assertEqual(updated["line_status"], HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(updated["line_ready_blockers"], [])
        self.assertIn("missing company", updated["import_commit_blockers"])
        self.assertNotIn("missing document number", updated["import_commit_blockers"])
        self.assertEqual(response.data["batch"]["wizard_summary"]["line_counts"]["ready"], 1)

    def test_batch_summary_closes_pending_suggestions_for_committed_rows(self):
        batch = HistoricalImportBatch.objects.create(name="Committed stale cards", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="committed.pdf",
            source_sha256="e" * 64,
            document_number="Q-COM-1",
            document_date=date(2026, 6, 2),
            status=HistoricalPriceImport.STATUS_COMMITTED,
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="DEEP HEAT SPRAY",
            quantity=Decimal("10.000"),
            unit="BOTTLES",
            unit_price=Decimal("18.00"),
            line_total=Decimal("180.00"),
            product=self.item_one,
            status=HistoricalPriceImportLine.STATUS_COMMITTED,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            status=HistoricalImportAISuggestion.STATUS_PENDING,
            confidence=0.95,
            reason="Old pending card.",
            created_by=self.staff,
        )

        response = self.client.get(reverse("quotation-historical-import-batch-detail", args=[batch.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        suggestion.refresh_from_db()
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(response.data["pending_suggestion_count"], 0)
        self.assertEqual(
            response.data["wizard_summary"]["suggestion_status_counts"][HistoricalImportAISuggestion.STATUS_APPLIED],
            1,
        )
        self.assertEqual(response.data["wizard_summary"]["line_counts"]["committed"], 1)

    def test_batch_summary_closes_pending_product_link_suggestion_and_repairs_ready_line(self):
        batch = HistoricalImportBatch.objects.create(name="Old product-linked pending card", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            suggested_company_name="Intermass 17032026A",
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="Intermass 17032026A.pdf",
            source_sha256="1" * 64,
            document_date=date(2026, 3, 17),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="DEEP HEAT SPRAY",
            quantity=Decimal("10.000"),
            unit="BOTTLES",
            unit_price=Decimal("18.00"),
            line_total=Decimal("180.00"),
            product=self.item_one,
            match_reason="Approved AI Product match: old response",
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            status=HistoricalImportAISuggestion.STATUS_PENDING,
            confidence=0.95,
            reason="Old pending card with already linked Product.",
            created_by=self.staff,
        )

        response = self.client.get(reverse("quotation-historical-import-batch-detail", args=[batch.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        suggestion.refresh_from_db()
        line.refresh_from_db()
        self.assertEqual(suggestion.status, HistoricalImportAISuggestion.STATUS_APPLIED)
        self.assertEqual(line.status, HistoricalPriceImportLine.STATUS_READY)
        self.assertEqual(response.data["pending_suggestion_count"], 0)
        self.assertEqual(response.data["wizard_summary"]["line_counts"]["ready"], 1)

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {}, clear=True)
    def test_ai_learning_missing_key_fails_cleanly(self):
        batch = self.client.post(reverse("quotation-historical-import-batch-list"), {"name": "No key"}, format="json")
        batch_id = batch.data["id"]
        settings_obj = QuotationSettings.get_solo()
        settings_obj.ai_parsing_enabled = True
        settings_obj.save()
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                upload = self.client.post(
                    reverse("quotation-historical-import-batch-upload-file", args=[batch_id]),
                    {"file": self.make_historical_pdf_upload()},
                    format="multipart",
                )

        response = self.client.post(
            reverse("quotation-historical-import-batch-run-ai-suggestions", args=[batch_id]),
            {"import_ids": [upload.data["import"]["id"]], "mode": "text"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["summary"]["failed"], 1)
        self.assertIn("missing API key", response.data["summary"]["results"][0]["message"])
        self.assertEqual(HistoricalImportAISuggestion.objects.count(), 0)

    def test_batch_commit_blocks_selected_imports_with_zero_ready_rows(self):
        batch = HistoricalImportBatch.objects.create(name="No ready rows", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="blocked.pdf",
            source_sha256="c" * 64,
            document_number="Q-BLOCKED",
            document_date=date(2026, 5, 21),
            created_by=self.staff,
        )
        HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="Pending Item",
            quantity=Decimal("1.000"),
            unit_price=Decimal("5.00"),
            status=HistoricalPriceImportLine.STATUS_NEEDS_REVIEW,
        )

        response = self.client.post(
            reverse("quotation-historical-import-batch-commit-ready-imports", args=[batch.id]),
            {"import_ids": [historical_import.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["summary"]["blocked"], 1)
        self.assertEqual(response.data["summary"]["committed"], 0)
        self.assertIn("no ready rows", response.data["results"][0]["message"])
        self.assertEqual(CompanyPriceHistory.objects.count(), 0)

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {}, clear=True)
    def test_ai_learning_failure_reports_previous_pending_suggestions(self):
        batch = HistoricalImportBatch.objects.create(name="Stale suggestions", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="stale.pdf",
            source_sha256="1" * 64,
            document_number="Q-STale",
            document_date=date(2026, 5, 21),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="SAVLON ANTISEPTIC SOLUTION",
            quantity=Decimal("1.000"),
            unit_price=Decimal("5.00"),
        )
        HistoricalImportAISuggestion.objects.create(
            batch=batch,
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            confidence=0.9,
            reason="Previous run.",
            created_by=self.staff,
        )
        settings_obj = QuotationSettings.get_solo()
        settings_obj.ai_parsing_enabled = True
        settings_obj.save()

        response = self.client.post(
            reverse("quotation-historical-import-batch-run-ai-suggestions", args=[batch.id]),
            {"import_ids": [historical_import.id], "mode": "text"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        result = response.data["summary"]["results"][0]
        self.assertEqual(result["status"], "failed")
        self.assertTrue(result["showing_previous_suggestions"])
        self.assertEqual(result["previous_suggestion_count"], 1)
        self.assertIn("missing API key", result["message"])

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_ai_learning_skips_obvious_historical_noise_before_product_review(self):
        batch = HistoricalImportBatch.objects.create(name="Noise gate", created_by=self.staff)
        historical_import = HistoricalPriceImport.objects.create(
            batch=batch,
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="noise.pdf",
            source_sha256="0" * 64,
            document_number="Q-TEST",
            document_date=date(2026, 5, 21),
            created_by=self.staff,
        )
        HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="Item Description",
            raw_line="Item Description | Qty | Unit | Unit Price | Total",
            sort_order=0,
        )
        HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="SAVLON ANTISEPTIC SOLUTION",
            raw_line="1 | SAVLON ANTISEPTIC SOLUTION | 1 | bottle | 5.00 | 5.25",
            quantity=Decimal("1.000"),
            unit="bottle",
            unit_price=Decimal("5.00"),
            line_total=Decimal("5.25"),
            sort_order=1,
        )
        settings_obj = QuotationSettings.get_solo()
        settings_obj.ai_parsing_enabled = True
        settings_obj.save()

        provider = MockLearningProvider()
        with patch("quotations.ai_learning.get_ai_parse_provider", return_value=provider):
            response = self.client.post(
                reverse("quotation-historical-import-batch-run-ai-suggestions", args=[batch.id]),
                {"import_ids": [historical_import.id], "mode": "text"},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        context = json.loads(provider.calls[0]["text_context"])
        self.assertEqual([row["item_name"] for row in context["rows"]], ["SAVLON ANTISEPTIC SOLUTION"])
        skip_suggestion = HistoricalImportAISuggestion.objects.get(action=HistoricalImportAISuggestion.ACTION_SKIP)
        self.assertEqual(skip_suggestion.line.item_name, "Item Description")
        self.assertIn("table header", skip_suggestion.reason.lower())

        batch_response = self.client.get(reverse("quotation-historical-import-batch-detail", args=[batch.id]))
        self.assertEqual(batch_response.status_code, status.HTTP_200_OK)
        self.assertEqual(batch_response.data["wizard_summary"]["pending_suggestion_action_counts"]["skip"], 1)
        self.assertEqual(batch_response.data["wizard_summary"]["line_counts"]["total"], 2)

    def test_ai_suggestion_source_context_is_staff_only(self):
        response = self.create_parsed_historical_import()
        historical_import = HistoricalPriceImport.objects.get(pk=response.data["id"])
        line = historical_import.lines.first()
        suggestion = HistoricalImportAISuggestion.objects.create(
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            confidence=0.91,
            reason="Review source.",
            created_by=self.staff,
        )
        url = reverse("quotation-historical-import-ai-suggestion-source-context", args=[suggestion.id])

        self.client.force_authenticate(None)
        anonymous = self.client.get(url)
        self.assertIn(anonymous.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

        self.client.force_authenticate(self.customer)
        blocked = self.client.get(url)
        self.assertEqual(blocked.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        allowed = self.client.get(url)
        self.assertEqual(allowed.status_code, status.HTTP_200_OK)
        self.assertTrue(allowed.data["available"])
        self.assertIn("preview_page", allowed.data["preview_url"])

    def test_ai_suggestion_serializer_includes_price_history_context_and_variance_warning(self):
        quotation = Quotation.objects.create(company=self.company, quotation_number="Q-PRICE", status=Quotation.STATUS_FINALIZED)
        quotation_line = QuotationLine.objects.create(
            quotation=quotation,
            product=self.item_one,
            item_name_snapshot=self.item_one.name,
            quantity=Decimal("1.000"),
            unit_price=Decimal("10.00"),
            vat_rate=Decimal("0.00"),
        )
        CompanyPriceHistory.objects.create(
            company=self.company,
            product=self.item_one,
            quotation=quotation,
            quotation_line=quotation_line,
            unit_price=Decimal("10.00"),
            quantity=Decimal("1.000"),
            unit="bottle",
        )
        historical_import = HistoricalPriceImport.objects.create(
            company=self.company,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="price.pdf",
            source_sha256="2" * 64,
            document_number="Q-PRICE-NEW",
            document_date=date(2026, 5, 22),
            created_by=self.staff,
        )
        line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="SAVLON ANTISEPTIC SOLUTION",
            quantity=Decimal("1.000"),
            unit_price=Decimal("18.00"),
        )
        suggestion = HistoricalImportAISuggestion.objects.create(
            historical_import=historical_import,
            line=line,
            suggestion_type=HistoricalImportAISuggestion.TYPE_LINE,
            action=HistoricalImportAISuggestion.ACTION_MATCH_EXISTING_PRODUCT,
            suggested_product=self.item_one,
            confidence=0.92,
            reason="Existing Product.",
            created_by=self.staff,
        )

        response = self.client.get(reverse("quotation-historical-import-ai-suggestion-detail", args=[suggestion.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        summary = response.data["price_history_summary"]
        self.assertEqual(summary["last_company_price"], "10.00")
        self.assertEqual(summary["imported_unit_price"], "18.00")
        self.assertEqual(summary["price_difference"], "8.00")
        self.assertIn("Large variance", summary["variance_warning"])

    def test_historical_import_same_company_document_number_opens_existing_import(self):
        first = self.create_parsed_historical_import()
        import_id = first.data["id"]
        self.client.patch(
            reverse("quotation-historical-import-detail", args=[import_id]),
            {"company": self.company.id},
            format="json",
        )

        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                second = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {
                        "file": self.make_historical_pdf_upload(
                            name="ANCIENT BUILDERS CONSTN 21052026 copy.pdf",
                            extra_text="Resaved copy",
                        )
                    },
                    format="multipart",
                )

        self.assertEqual(second.status_code, status.HTTP_200_OK)
        self.assertEqual(second.data["id"], import_id)
        self.assertTrue(second.data["duplicate_check"]["blocked_new_import"])
        self.assertEqual(second.data["duplicate_check"]["message"], "This quotation already exists for this company.")

    def test_historical_import_similar_rows_warns_but_allows_review(self):
        first = self.create_parsed_historical_import()
        self.client.patch(
            reverse("quotation-historical-import-detail", args=[first.data["id"]]),
            {"company": self.company.id},
            format="json",
        )

        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                second = self.client.post(
                    reverse("quotation-historical-import-parse-file"),
                    {
                        "file": self.make_historical_pdf_upload(
                            name="ANCIENT BUILDERS CONSTN alternate.pdf",
                            document_number="QUOTATION-26052101-ALT",
                            extra_text="Alternate exported file with same rows",
                        )
                    },
                    format="multipart",
                )

        self.assertEqual(second.status_code, status.HTTP_201_CREATED)
        self.assertNotEqual(second.data["id"], first.data["id"])
        duplicate_check = second.data["duplicate_check"]
        self.assertTrue(duplicate_check["is_duplicate"])
        self.assertFalse(duplicate_check["blocking"])
        self.assertEqual(duplicate_check["message"], "This looks similar to a previous import.")

    def test_encrypted_historical_pdf_is_rejected(self):
        response = self.client.post(
            reverse("quotation-historical-import-parse-file"),
            {"file": self.make_historical_pdf_upload(encrypted=True)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Encrypted PDF files are not supported", str(response.data))

    def test_historical_bulk_actions_are_staff_only(self):
        response = self.create_parsed_historical_import()
        import_id = response.data["id"]
        row_id = response.data["lines"][0]["id"]
        url = reverse("quotation-historical-import-bulk-create-quote-items", args=[import_id])

        self.client.force_authenticate(None)
        anonymous = self.client.post(url, {"row_ids": [row_id]}, format="json")
        self.assertIn(anonymous.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

        self.client.force_authenticate(self.customer)
        non_staff = self.client.post(url, {"row_ids": [row_id]}, format="json")
        self.assertEqual(non_staff.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        staff = self.client.post(url, {"row_ids": [row_id]}, format="json")
        self.assertEqual(staff.status_code, status.HTTP_200_OK)

    def test_bulk_create_quote_items_links_existing_and_creates_missing_without_duplicates(self):
        response = self.create_parsed_historical_import()
        import_id = response.data["id"]
        historical_import = HistoricalPriceImport.objects.get(pk=import_id)
        new_line = HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="CUSTOM FIRST AID REFILL",
            quantity=Decimal("2.000"),
            unit="pack",
            unit_price=Decimal("4.00"),
            sort_order=99,
        )
        before_count = Product.objects.count()
        row_ids = [response.data["lines"][0]["id"], new_line.id]

        result = self.client.post(
            reverse("quotation-historical-import-bulk-create-quote-items", args=[import_id]),
            {"row_ids": row_ids},
            format="json",
        )

        self.assertEqual(result.status_code, status.HTTP_200_OK)
        self.assertEqual(result.data["summary"]["linked_existing"], 1)
        self.assertEqual(result.data["summary"]["created"], 1)
        self.assertEqual(Product.objects.count(), before_count + 1)
        existing_line = HistoricalPriceImportLine.objects.get(pk=row_ids[0])
        created_line = HistoricalPriceImportLine.objects.get(pk=new_line.id)
        self.assertEqual(existing_line.product, self.item_one)
        self.assertEqual(created_line.product.name, "CUSTOM FIRST AID REFILL")
        self.assertEqual(created_line.status, HistoricalPriceImportLine.STATUS_NEEDS_REVIEW)

        repeat = self.client.post(
            reverse("quotation-historical-import-bulk-create-quote-items", args=[import_id]),
            {"row_ids": [new_line.id]},
            format="json",
        )

        self.assertEqual(repeat.status_code, status.HTTP_200_OK)
        self.assertEqual(repeat.data["summary"]["linked_existing"], 1)
        self.assertEqual(Product.objects.count(), before_count + 1)

    def test_bulk_update_ready_validation_blocks_missing_required_fields(self):
        response = self.create_parsed_historical_import()
        import_id = response.data["id"]
        row_id = response.data["lines"][0]["id"]

        result = self.client.post(
            reverse("quotation-historical-import-bulk-update-rows", args=[import_id]),
            {"row_ids": [row_id], "status": HistoricalPriceImportLine.STATUS_READY},
            format="json",
        )

        self.assertEqual(result.status_code, status.HTTP_200_OK)
        self.assertEqual(result.data["summary"]["failed"], 1)
        self.assertIn("Select the company", result.data["summary"]["results"][0]["message"])
        line = HistoricalPriceImportLine.objects.get(pk=row_id)
        self.assertEqual(line.status, HistoricalPriceImportLine.STATUS_NEEDS_REVIEW)

    def test_bulk_skip_excludes_rows_from_commit(self):
        response = self.create_parsed_historical_import()
        import_id = response.data["id"]
        lines = response.data["lines"]
        self.client.patch(
            reverse("quotation-historical-import-detail", args=[import_id]),
            {"company": self.company.id},
            format="json",
        )
        self.client.post(
            reverse("quotation-historical-import-bulk-create-quote-items", args=[import_id]),
            {"row_ids": [line["id"] for line in lines]},
            format="json",
        )
        skipped = self.client.post(
            reverse("quotation-historical-import-bulk-skip-rows", args=[import_id]),
            {"row_ids": [lines[1]["id"]]},
            format="json",
        )
        ready = self.client.post(
            reverse("quotation-historical-import-bulk-update-rows", args=[import_id]),
            {"row_ids": [lines[0]["id"]], "status": HistoricalPriceImportLine.STATUS_READY},
            format="json",
        )

        self.assertEqual(skipped.status_code, status.HTTP_200_OK)
        self.assertEqual(ready.status_code, status.HTTP_200_OK)

        commit = self.client.post(reverse("quotation-historical-import-commit", args=[import_id]))

        self.assertEqual(commit.status_code, status.HTTP_200_OK)
        self.assertEqual(CompanyPriceHistory.objects.filter(company=self.company).count(), 1)
        skipped_line = HistoricalPriceImportLine.objects.get(pk=lines[1]["id"])
        self.assertEqual(skipped_line.status, HistoricalPriceImportLine.STATUS_SKIPPED)


class MockAIProvider:
    def __init__(self, result=None):
        self.result = result or {
            "rows": [
                {
                    "item_name": "Electrorush 21gm sache for 1 Ltr solution",
                    "quantity": "50",
                    "unit": "carton",
                    "unit_price": "375",
                    "line_total": "",
                    "pack_info": "1 box = 10 sachets",
                    "notes": "",
                    "raw_source_text": "1 box 10 sachets 50 box cartoon price : 375 per cartoon",
                    "page_number": "",
                    "confidence": 85,
                    "parse_status": "parsed",
                    "reason": "Detected order quantity and price per carton.",
                },
                {
                    "item_name": "Zest Ors 21 gm sachet for 1 Ltr solution",
                    "quantity": "",
                    "unit": "box",
                    "unit_price": "18",
                    "line_total": "",
                    "pack_info": "25 sachet per box",
                    "notes": "",
                    "raw_source_text": "25 sachet per box price 18 per box",
                    "page_number": "",
                    "confidence": 65,
                    "parse_status": "needs_review",
                    "reason": "Price per box detected, but order quantity is unclear.",
                },
                {
                    "item_name": "",
                    "quantity": "",
                    "unit": "",
                    "unit_price": "",
                    "line_total": "",
                    "pack_info": "",
                    "notes": "",
                    "raw_source_text": "DATE: 21/05/2026",
                    "page_number": "",
                    "confidence": 95,
                    "parse_status": "ignored",
                    "reason": "Document metadata.",
                },
            ],
            "warnings": [],
            "document_notes": "Cleaned rows for staff review.",
        }
        self.calls = []

    def clean_rows(self, **kwargs):
        self.calls.append(kwargs)
        return self.result, {"input_tokens": 10, "output_tokens": 20}


class MockLearningProvider:
    def __init__(self, result=None):
        self.result = result
        self.calls = []

    def clean_rows(self, **kwargs):
        self.calls.append(kwargs)
        if self.result is not None:
            return self.result, {"input_tokens": 20, "output_tokens": 30}
        context = __import__("json").loads(kwargs["text_context"])
        company_candidates = context.get("candidate_companies") or []
        rows = []
        for index, row in enumerate(context.get("rows") or []):
            candidate_products = row.get("candidate_products") or []
            if index == 0 and candidate_products:
                rows.append(
                    {
                        "line_id": str(row["line_id"]),
                        "action": "create_company_alias",
                        "product_id": str(candidate_products[0]["id"]),
                        "alias_text": row["item_name"],
                        "new_product_name": "",
                        "new_product_unit": "",
                        "new_product_pack_size": "",
                        "new_product_dosage": "",
                        "confidence": 0.93,
                        "reason": "Customer wording is a clear alias for the candidate Product.",
                        "candidate_product_ids": [str(candidate["id"]) for candidate in candidate_products],
                    }
                )
            elif index == 1:
                rows.append(
                    {
                        "line_id": str(row["line_id"]),
                        "action": "create_new_product",
                        "product_id": "",
                        "alias_text": "",
                        "new_product_name": "Custom Historical Refill",
                        "new_product_unit": row.get("unit") or "box",
                        "new_product_pack_size": row.get("unit") or "box",
                        "new_product_dosage": "",
                        "confidence": 0.86,
                        "reason": "No candidate Product is specific enough.",
                        "candidate_product_ids": [str(candidate["id"]) for candidate in candidate_products],
                    }
                )
            else:
                rows.append(
                    {
                        "line_id": str(row["line_id"]),
                        "action": "needs_manual_review",
                        "product_id": "",
                        "alias_text": "",
                        "new_product_name": "",
                        "new_product_unit": "",
                        "new_product_pack_size": "",
                        "new_product_dosage": "",
                        "confidence": 0.45,
                        "reason": "Ambiguous row.",
                        "candidate_product_ids": [str(candidate["id"]) for candidate in candidate_products],
                    }
                )
        return {
            "company": {
                "action": "match_existing_company" if company_candidates else "needs_manual_review",
                "company_id": str(company_candidates[0]["id"]) if company_candidates else "",
                "proposed_company_name": context["document"].get("suggested_company_name") or "",
                "confidence": 0.9 if company_candidates else 0.4,
                "reason": "Company candidate matches the document name." if company_candidates else "No clear company candidate.",
                "candidate_company_ids": [str(candidate["id"]) for candidate in company_candidates],
            },
            "rows": rows,
            "warnings": [],
            "document_notes": "Learning suggestions for staff review.",
        }, {"input_tokens": 20, "output_tokens": 30}


class AIImportParsingTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(username="ai_staff", password="pass", is_staff=True)
        self.customer = User.objects.create_user(username="ai_customer", password="pass")
        self.company = Company.objects.create(name="AI Company")
        self.client.force_authenticate(self.staff)

    def enable_ai(self, *, auto=False, vision=False):
        settings_obj = QuotationSettings.get_solo()
        settings_obj.ai_parsing_enabled = True
        settings_obj.ai_auto_cleanup_enabled = auto
        settings_obj.ai_pdf_vision_enabled = vision
        settings_obj.save()
        return settings_obj

    def preview_payload(self):
        return {
            "source_type": Inquiry.SOURCE_TYPE_PASTED_TEXT,
            "source_filename": "",
            "source_mime_type": "text/plain",
            "source_sha256": "b" * 64,
            "source_file_ref": "",
            "parse_method": "deterministic_text_v2",
            "original_text": "messy text",
            "warnings": [],
            "meta": {},
            "lines": [
                {
                    "raw_name": "Electrorush 21gm sache for 1 Ltr solution 50 carton price 375",
                    "raw_line": "Electrorush 21gm sache for 1 Ltr solution 50 carton price 375",
                    "parse_status": InquiryLine.PARSE_NEEDS_REVIEW,
                    "parse_confidence": 0.4,
                }
            ],
        }

    def make_pdf_upload(self):
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        pdf.drawString(72, 740, "Material Description Req Quantity unit u price total")
        pdf.drawString(72, 720, "Deep Heat Spray 150ml 5 No 12 60")
        pdf.save()
        return SimpleUploadedFile("ai-test.pdf", buffer.getvalue(), content_type="application/pdf")

    def test_ai_settings_defaults_and_update(self):
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-settings"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["ai_parsing_enabled"])
        self.assertIn("ai_available", response.data)

        update = self.client.patch(
            reverse("quotation-settings"),
            {
                "ai_parsing_enabled": True,
                "ai_auto_cleanup_enabled": True,
                "ai_pdf_vision_enabled": True,
            },
            format="json",
        )
        self.assertEqual(update.status_code, status.HTTP_200_OK)
        settings_obj = QuotationSettings.get_solo()
        self.assertTrue(settings_obj.ai_parsing_enabled)
        self.assertTrue(settings_obj.ai_auto_cleanup_enabled)
        self.assertTrue(settings_obj.ai_pdf_vision_enabled)

    def test_ai_clean_parse_staff_only(self):
        url = reverse("quotation-inquiry-ai-clean-parse")
        payload = {"preview": self.preview_payload()}

        self.client.force_authenticate(None)
        anonymous = self.client.post(url, payload, format="json")
        self.assertIn(anonymous.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

        self.client.force_authenticate(self.customer)
        non_staff = self.client.post(url, payload, format="json")
        self.assertEqual(non_staff.status_code, status.HTTP_403_FORBIDDEN)

    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_ai_disabled_returns_clear_response_without_provider_call(self):
        provider = MockAIProvider()
        with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
            response = self.client.post(
                reverse("quotation-inquiry-ai-clean-parse"),
                {"preview": self.preview_payload()},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("AI disabled in settings", response.data["detail"])
        self.assertEqual(provider.calls, [])

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
        QUOTATION_AI_PARSE_VISION_MODEL="test-vision-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_text_ai_valid_json_returns_candidate_rows_without_side_effects(self):
        self.enable_ai()
        provider = MockAIProvider()

        with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
            response = self.client.post(
                reverse("quotation-inquiry-ai-clean-parse"),
                {"preview": self.preview_payload(), "company": self.company.id},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result_source"], "ai_text_cleanup")
        self.assertEqual(response.data["lines"][0]["raw_name"], "Electrorush 21gm sache for 1 Ltr solution")
        self.assertEqual(response.data["lines"][0]["quantity"], "50")
        self.assertEqual(response.data["lines"][0]["unit_price"], "375")
        self.assertEqual(response.data["lines"][1]["parse_status"], InquiryLine.PARSE_NEEDS_REVIEW)
        self.assertFalse(response.data["lines"][0]["matched_product"])
        self.assertGreaterEqual(response.data["lines"][0]["parse_confidence"], 0.85)
        self.assertEqual(Product.objects.count(), 0)
        self.assertEqual(ProductAlias.objects.count(), 0)
        self.assertEqual(CompanyPriceHistory.objects.count(), 0)
        self.assertEqual(Quotation.objects.count(), 0)
        self.assertEqual(AIParseLog.objects.filter(success=True).count(), 1)

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_invalid_ai_json_is_rejected_safely(self):
        self.enable_ai()
        provider = MockAIProvider(result={"not_rows": []})

        with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
            response = self.client.post(
                reverse("quotation-inquiry-ai-clean-parse"),
                {"preview": self.preview_payload()},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["ai_status"], "ai_failed_using_original_parse")
        self.assertEqual(Inquiry.objects.count(), 0)
        self.assertEqual(AIParseLog.objects.filter(success=False).count(), 1)

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
        QUOTATION_AI_PARSE_VISION_MODEL="test-vision-model",
        QUOTATION_PRIVATE_STORAGE_ROOT=None,
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_vision_ai_path_for_pdf_when_enabled(self):
        self.enable_ai(vision=True)
        provider = MockAIProvider()
        with tempfile.TemporaryDirectory() as private_root:
            with override_settings(QUOTATION_PRIVATE_STORAGE_ROOT=private_root):
                parsed = self.client.post(
                    reverse("quotation-inquiry-parse-file"),
                    {"file": self.make_pdf_upload()},
                    format="multipart",
                )
                self.assertEqual(parsed.status_code, status.HTTP_200_OK)
                with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
                    response = self.client.post(
                        reverse("quotation-inquiry-ai-clean-parse"),
                        {"preview": parsed.data, "mode": "vision"},
                        format="json",
                    )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result_source"], "ai_vision_cleanup")
        self.assertEqual(provider.calls[0]["mode"], "vision")
        self.assertGreaterEqual(len(provider.calls[0]["image_data_urls"]), 1)

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
        QUOTATION_AI_PARSE_VISION_MODEL="test-vision-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_vision_ai_blocked_when_setting_disabled(self):
        self.enable_ai(vision=False)
        response = self.client.post(
            reverse("quotation-inquiry-ai-clean-parse"),
            {"preview": {**self.preview_payload(), "source_type": Inquiry.SOURCE_TYPE_PDF}, "mode": "vision"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("vision cleanup is disabled", response.data["detail"])

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_auto_ai_triggers_only_for_poor_parse_not_missing_product_match(self):
        self.enable_ai(auto=True)
        provider = MockAIProvider()
        with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
            poor = self.client.post(
                reverse("quotation-inquiry-parse-text"),
                {"raw_text": "DATE: 21/05/2026\nFrom(Seller): Al Ameen Pharmacy"},
                format="json",
            )
        self.assertEqual(poor.status_code, status.HTTP_200_OK)
        self.assertIn("ai_candidate", poor.data)
        self.assertEqual(len(provider.calls), 1)

        provider.calls.clear()
        with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
            good_unmatched = self.client.post(
                reverse("quotation-inquiry-parse-text"),
                {"raw_text": "Unknown Private Medicine - 10 boxes", "company": self.company.id},
                format="json",
            )
        self.assertEqual(good_unmatched.status_code, status.HTTP_200_OK)
        self.assertNotIn("ai_candidate", good_unmatched.data)
        self.assertEqual(provider.calls, [])
        self.assertFalse(good_unmatched.data["lines"][0].get("matched_product"))
        self.assertGreaterEqual(good_unmatched.data["lines"][0]["parse_confidence"], 0.8)

    @override_settings(
        QUOTATION_AI_PARSE_GLOBAL_ENABLED=True,
        QUOTATION_AI_PARSE_PROVIDER="openai",
        QUOTATION_AI_PARSE_TEXT_MODEL="test-text-model",
    )
    @patch.dict("os.environ", {"OPENAI_API_KEY": "test-key"}, clear=False)
    def test_historical_ai_candidate_requires_apply_before_replacing_rows(self):
        self.enable_ai()
        historical_import = HistoricalPriceImport.objects.create(
            company=self.company,
            suggested_company_name=self.company.name,
            source_type=HistoricalPriceImport.SOURCE_TYPE_PDF,
            source_filename="old-quotation.pdf",
            source_sha256="c" * 64,
            parse_method="deterministic_test",
            document_number="QT-AI-1",
            document_date="2026-05-21",
            created_by=self.staff,
        )
        HistoricalPriceImportLine.objects.create(
            historical_import=historical_import,
            item_name="Messy AI Row 10 box price 5",
            raw_line="Messy AI Row 10 box price 5",
            quantity=Decimal("10.000"),
            unit="box",
            unit_price=Decimal("5.00"),
            parse_confidence=0.4,
        )
        original_line_count = historical_import.lines.count()
        provider = MockAIProvider()

        with patch("quotations.ai_parsing.get_ai_parse_provider", return_value=provider):
            clean_response = self.client.post(
                reverse("quotation-historical-import-ai-clean-rows", args=[historical_import.id]),
                {"mode": "text"},
                format="json",
            )

        self.assertEqual(clean_response.status_code, status.HTTP_200_OK)
        historical_import.refresh_from_db()
        self.assertEqual(historical_import.lines.count(), original_line_count)
        self.assertEqual(Product.objects.count(), 0)

        apply_response = self.client.post(
            reverse("quotation-historical-import-apply-ai-clean-rows", args=[historical_import.id]),
            {
                "lines": clean_response.data["lines"],
                "result_source": clean_response.data["result_source"],
                "provider": clean_response.data["provider"],
                "model": clean_response.data["model"],
            },
            format="json",
        )

        self.assertEqual(apply_response.status_code, status.HTTP_200_OK)
        historical_import.refresh_from_db()
        self.assertEqual(historical_import.lines.count(), 2)
        self.assertTrue(all(line.status == HistoricalPriceImportLine.STATUS_NEEDS_REVIEW for line in historical_import.lines.all()))
        self.assertEqual(Product.objects.count(), 0)
        self.assertEqual(ProductAlias.objects.count(), 0)
        self.assertEqual(CompanyPriceHistory.objects.count(), 0)


class QuotationSettingsTests(APITestCase):
    def setUp(self):
        self.staff = User.objects.create_user(username="settings_staff", password="pass", is_staff=True)
        self.customer = User.objects.create_user(username="settings_customer", password="pass")
        self.company = Company.objects.create(name="Settings Company")
        self.product = Product.objects.create(name="Settings Item", price=Decimal("1.00"), pack_size="box", status="draft")

    def create_valid_quote(self):
        quotation = Quotation.objects.create(company=self.company, created_by=self.staff)
        QuotationLine.objects.create(
            quotation=quotation,
            product=self.product,
            item_name_snapshot="Settings Item",
            quantity=Decimal("1.000"),
            unit="box",
            unit_price=Decimal("25.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
        )
        return quotation

    def test_settings_permissions(self):
        url = reverse("quotation-settings")

        anonymous = self.client.get(url)
        self.assertIn(anonymous.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

        self.client.force_authenticate(self.customer)
        non_staff = self.client.get(url)
        self.assertEqual(non_staff.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.staff)
        staff = self.client.get(url)
        self.assertEqual(staff.status_code, status.HTTP_200_OK)

    def test_settings_defaults_returned_if_missing(self):
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-settings"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["company_name"], "Al Ameen Pharmacy")
        self.assertEqual(response.data["payment_terms"], "As per mutually agreed terms.")
        self.assertEqual(response.data["validity_days"], 30)
        self.assertEqual(QuotationSettings.objects.count(), 1)

    def test_settings_update_works(self):
        self.client.force_authenticate(self.staff)

        response = self.client.patch(
            reverse("quotation-settings"),
            {
                "company_name": "Custom Pharmacy",
                "trn": "123456789",
                "license_number": "LIC-42",
                "validity_days": 30,
                "primary_color": "#123456",
                "logo_layout": QuotationSettings.LOGO_LAYOUT_LOGO_TEXT,
                "show_stamp_area": False,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        settings_obj = QuotationSettings.get_solo()
        self.assertEqual(settings_obj.company_name, "Custom Pharmacy")
        self.assertEqual(settings_obj.validity_days, 30)
        self.assertEqual(settings_obj.logo_layout, QuotationSettings.LOGO_LAYOUT_LOGO_TEXT)
        self.assertEqual(settings_obj.updated_by, self.staff)

    def test_invalid_color_values_are_rejected(self):
        self.client.force_authenticate(self.staff)

        response = self.client.patch(reverse("quotation-settings"), {"primary_color": "teal"}, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("primary_color", response.data)

    def test_logo_upload_rejects_invalid_file_type(self):
        self.client.force_authenticate(self.staff)
        upload = SimpleUploadedFile("logo.txt", b"not-an-image", content_type="text/plain")

        response = self.client.patch(reverse("quotation-settings"), {"logo": upload}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("valid image", str(response.data))

    def test_signature_and_stamp_uploads_work(self):
        self.client.force_authenticate(self.staff)
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                response = self.client.patch(
                    reverse("quotation-settings"),
                    {
                        "signature_image": make_png_upload("signature.png"),
                        "stamp_image": make_png_upload("stamp.png", color=(212, 160, 65, 255)),
                    },
                    format="multipart",
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertTrue(response.data["signature_image_url"])
                self.assertTrue(response.data["stamp_image_url"])
                settings_obj = QuotationSettings.get_solo()
                self.assertTrue(settings_obj.signature_image.name.startswith("quotations/signatures/"))
                self.assertTrue(settings_obj.stamp_image.name.startswith("quotations/stamps/"))

    def test_user_signature_uploads_work(self):
        self.client.force_authenticate(self.staff)
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                response = self.client.patch(
                    reverse("quotation-my-signature"),
                    {"signature_image": make_png_upload("my-signature.png")},
                    format="multipart",
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertTrue(response.data["signature_image_url"])
                profile = UserQuotationProfile.objects.get(user=self.staff)
                self.assertTrue(profile.signature_image.name.startswith("quotations/user-signatures/"))

    def test_pdf_config_prefers_creator_signature_over_shared_signature(self):
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                settings_obj = QuotationSettings.get_solo()
                settings_obj.signature_image.save("shared-signature.png", ContentFile(make_png_bytes()), save=True)
                profile = UserQuotationProfile.objects.create(user=self.staff)
                profile.signature_image.save("user-signature.png", ContentFile(make_png_bytes(color=(212, 160, 65, 255))), save=True)
                quotation = self.create_valid_quote()

                config = get_quotation_pdf_config(quotation=quotation)

                self.assertIn("quotations/user-signatures/", config.signature_image_path.replace("\\", "/"))

    def test_pdf_config_ignores_shared_signature_when_user_has_none(self):
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                settings_obj = QuotationSettings.get_solo()
                settings_obj.signature_image.save("shared-signature.png", ContentFile(make_png_bytes()), save=True)
                quotation = self.create_valid_quote()

                config = get_quotation_pdf_config(quotation=quotation)

                self.assertEqual(config.signature_image_path, "")

    def test_staff_can_clear_logo_signature_and_stamp(self):
        self.client.force_authenticate(self.staff)
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                self.client.patch(
                    reverse("quotation-settings"),
                    {
                        "logo": make_png_upload("logo.png"),
                        "signature_image": make_png_upload("signature.png"),
                        "stamp_image": make_png_upload("stamp.png"),
                    },
                    format="multipart",
                )
                settings_obj = QuotationSettings.get_solo()
                self.assertTrue(settings_obj.logo)
                self.assertTrue(settings_obj.signature_image)
                self.assertTrue(settings_obj.stamp_image)

                response = self.client.patch(
                    reverse("quotation-settings"),
                    {"clear_logo": True, "clear_signature_image": True, "clear_stamp_image": True},
                    format="json",
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                settings_obj.refresh_from_db()
                self.assertFalse(settings_obj.logo)
                self.assertFalse(settings_obj.signature_image)
                self.assertFalse(settings_obj.stamp_image)

    def test_non_staff_and_anonymous_cannot_clear_images(self):
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                settings_obj = QuotationSettings.get_solo()
                settings_obj.logo.save("logo.png", ContentFile(make_png_bytes()), save=True)

                anonymous = self.client.patch(reverse("quotation-settings"), {"clear_logo": True}, format="json")
                self.assertIn(anonymous.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])
                settings_obj.refresh_from_db()
                self.assertTrue(settings_obj.logo)

                self.client.force_authenticate(self.customer)
                non_staff = self.client.patch(reverse("quotation-settings"), {"clear_logo": True}, format="json")
                self.assertEqual(non_staff.status_code, status.HTTP_403_FORBIDDEN)
                settings_obj.refresh_from_db()
                self.assertTrue(settings_obj.logo)

    def test_stamp_upload_rejects_invalid_file_type(self):
        self.client.force_authenticate(self.staff)
        upload = SimpleUploadedFile("stamp.txt", b"not-an-image", content_type="text/plain")

        response = self.client.patch(reverse("quotation-settings"), {"stamp_image": upload}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("valid image", str(response.data))

    def test_pdf_generation_uses_settings_and_still_works(self):
        QuotationSettings.objects.create(
            company_name="PDF Settings Pharmacy",
            default_terms="Settings terms apply.",
            payment_terms="Net 15.",
            prepared_by_default="Quotation Team",
            footer_note="Thank you for your business.",
        )
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_draft_pdf_includes_watermark(self):
        quotation = self.create_valid_quote()
        quotation.status = Quotation.STATUS_DRAFT
        quotation.save(update_fields=["status", "updated_at"])
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        text = extract_pdf_text(response.content)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("DRAFT", text)

    @override_settings(QUOTATION_LOGO_PATH="https://example.com/logo.png", QUOTATION_PDF_ALLOW_REMOTE_IMAGES=False)
    @patch("quotations.pdf.urlopen")
    def test_pdf_does_not_fetch_remote_images_by_default(self, mock_urlopen):
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        mock_urlopen.assert_not_called()
        self.assertNotIn("Al Ameen Pharmacy", extract_pdf_text(response.content))

    @override_settings(
        QUOTATION_LOGO_PATH="https://res.cloudinary.com/demo/logo.png",
        QUOTATION_PDF_ALLOW_REMOTE_IMAGES=False,
        QUOTATION_PDF_ALLOWED_REMOTE_IMAGE_HOSTS=["res.cloudinary.com"],
    )
    @patch("quotations.pdf.urlopen")
    def test_pdf_can_fetch_allowlisted_cloudinary_images(self, mock_urlopen):
        class MockImageResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def read(self, size=-1):
                return make_png_bytes()

        mock_urlopen.return_value = MockImageResponse()
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        mock_urlopen.assert_called_once()

    def test_pdf_wraps_long_customer_name_in_metadata_table(self):
        self.company.name = "Makharaafi International Technical Contracting and Facilities Management Services LLC"
        self.company.save(update_fields=["name"])
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        text = " ".join(extract_pdf_text(response.content).split())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("Makharaafi International Technical Contracting", text)
        self.assertIn(quotation.quotation_number, text)

    def test_pdf_shows_purchaser_contact_details(self):
        contact = CompanyContact.objects.create(
            company=self.company,
            name="Ahmed Khan",
            role="Purchase Officer",
            department="Procurement",
            phone="+971501234567",
            email="ahmed@example.com",
        )
        quotation = Quotation.objects.create(company=self.company, contact=contact, created_by=self.staff)
        QuotationLine.objects.create(
            quotation=quotation,
            product=self.product,
            item_name_snapshot="Settings Item",
            quantity=Decimal("1.000"),
            unit="box",
            unit_price=Decimal("25.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
        )
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        text = " ".join(extract_pdf_text(response.content).split())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("Attention Ahmed Khan", text)
        self.assertNotIn("Position Purchase Officer", text)
        self.assertNotIn("Department Procurement", text)
        self.assertIn("Contact No. +971501234567", text)
        self.assertIn("Contact Email ahmed@example.com", text)

    def test_pdf_hides_empty_optional_customer_fields_and_shows_address_trn(self):
        self.company.billing_address = "Office 12, Dubai Healthcare City, Dubai"
        self.company.trn = "100234567800003"
        self.company.save(update_fields=["billing_address", "trn"])
        contact = CompanyContact.objects.create(company=self.company, name="-", phone="-", email="-")
        quotation = Quotation.objects.create(company=self.company, contact=contact, created_by=self.staff)
        QuotationLine.objects.create(
            quotation=quotation,
            product=self.product,
            item_name_snapshot="Settings Item",
            quantity=Decimal("1.000"),
            unit="box",
            unit_price=Decimal("25.00"),
            match_status=QuotationLine.MATCH_CONFIRMED,
        )
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        text = " ".join(extract_pdf_text(response.content).split())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("Customer Address Office 12, Dubai Healthcare City, Dubai", text)
        self.assertIn("Customer TRN 100234567800003", text)
        self.assertNotIn("Attention", text)
        self.assertNotIn("Contact No.", text)
        self.assertNotIn("Contact Email", text)

    def test_excel_hides_empty_optional_customer_fields_and_shows_address_trn(self):
        self.company.billing_address = "Office 12, Dubai Healthcare City, Dubai"
        self.company.trn = "100234567800003"
        self.company.save(update_fields=["billing_address", "trn"])
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-excel", args=[quotation.id]))
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        values = [cell.value for row in workbook["Quotation"].iter_rows() for cell in row if cell.value is not None]

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("Customer Address", values)
        self.assertIn("Office 12, Dubai Healthcare City, Dubai", values)
        self.assertIn("Customer TRN", values)
        self.assertIn("100234567800003", values)
        self.assertNotIn("Attention", values)
        self.assertNotIn("Contact No.", values)
        self.assertNotIn("Contact Email", values)

    def test_pdf_hides_internal_line_notes_and_keeps_double_digit_serials_together(self):
        quotation = Quotation.objects.create(company=self.company, created_by=self.staff)
        for index in range(1, 13):
            QuotationLine.objects.create(
                quotation=quotation,
                product=self.product,
                item_name_snapshot=f"Item {index}",
                quantity=Decimal("1.000"),
                unit="each",
                unit_price=Decimal("10.00"),
                vat_rate=Decimal("0.00"),
                match_status=QuotationLine.MATCH_CONFIRMED,
                sort_order=index,
                notes="Quantity and unit detected, confirm accuracy.",
            )
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        text = extract_pdf_text(response.content)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertNotIn("Quantity and unit detected", text)
        self.assertIn("Item 10", text)
        self.assertNotIn("\n1\n0\n", text)
        self.assertNotIn("&nbsp;", text)
        self.assertIn("Subtotal", text)
        self.assertIn("VAT", text)
        self.assertIn("Grand Total", text)

    def test_pdf_generation_works_with_uploaded_logo_and_stamp_images(self):
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                settings_obj = QuotationSettings.get_solo()
                settings_obj.company_name = "Image Settings Pharmacy"
                settings_obj.logo.save("logo.png", ContentFile(make_png_bytes()), save=False)
                settings_obj.signature_image.save("signature.png", ContentFile(make_png_bytes()), save=False)
                settings_obj.stamp_image.save("stamp.png", ContentFile(make_png_bytes(color=(212, 160, 65, 255))), save=False)
                settings_obj.save()
                quotation = self.create_valid_quote()
                self.client.force_authenticate(self.staff)

                response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response["Content-Type"], "application/pdf")
                self.assertTrue(response.content.startswith(b"%PDF"))

    def test_pdf_full_logo_only_does_not_repeat_company_name_text(self):
        storage_settings = {
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root, STORAGES=storage_settings):
                settings_obj = QuotationSettings.get_solo()
                settings_obj.company_name = "Full Lockup Pharmacy"
                settings_obj.logo_layout = QuotationSettings.LOGO_LAYOUT_FULL
                settings_obj.logo.save("logo.png", ContentFile(make_png_bytes()), save=False)
                settings_obj.save()
                quotation = self.create_valid_quote()
                self.client.force_authenticate(self.staff)

                response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertNotIn("Full Lockup Pharmacy", extract_pdf_text(response.content))

    def test_pdf_no_logo_uses_company_name_text(self):
        settings_obj = QuotationSettings.get_solo()
        settings_obj.company_name = "No Logo Pharmacy"
        settings_obj.logo_layout = QuotationSettings.LOGO_LAYOUT_NONE
        settings_obj.save()
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("No Logo Pharmacy", extract_pdf_text(response.content))

    def test_pdf_missing_signature_and_stamp_use_placeholders(self):
        settings_obj = QuotationSettings.get_solo()
        settings_obj.logo_layout = QuotationSettings.LOGO_LAYOUT_NONE
        settings_obj.signature_label = "Signature"
        settings_obj.stamp_label = "Stamp"
        settings_obj.show_signature_area = True
        settings_obj.show_stamp_area = True
        settings_obj.save()
        quotation = self.create_valid_quote()
        self.client.force_authenticate(self.staff)

        response = self.client.get(reverse("quotation-pdf", args=[quotation.id]))
        text = " ".join(extract_pdf_text(response.content).split())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("Authorized Signature", text)
        self.assertIn("Company Stamp", text)

    def test_manual_inquiry_flow_still_works(self):
        self.client.force_authenticate(self.staff)
        response = self.client.post(
            reverse("quotation-inquiry-list"),
            {
                "company": self.company.id,
                "subject": "Manual inquiry still works",
                "lines": [{"raw_name": "Manual Item", "quantity": "1.000", "unit": "box"}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        inquiry = Inquiry.objects.get(id=response.data["id"])
        self.assertEqual(inquiry.source, Inquiry.SOURCE_MANUAL)
        self.assertEqual(inquiry.source_type, Inquiry.SOURCE_TYPE_MANUAL)
        self.assertEqual(inquiry.lines.count(), 1)
